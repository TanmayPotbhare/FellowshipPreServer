from flask import Flask, render_template, Response, make_response, request, redirect, session, url_for, jsonify, \
    send_file, flash, send_from_directory
import folium
import hashlib  # Import hashlib for password hashing
from openpyxl import Workbook
from io import BytesIO
from fpdf import FPDF
# from reportlab.pdfgen import canvas
import uuid
import base64
import mysql.connector
from werkzeug.utils import secure_filename
import re, os, csv
from flask_mail import Mail, Message
from itsdangerous import URLSafeTimedSerializer
import random
import json
from datetime import datetime, timedelta, date

# ----------- FLASK APP CONFIGURATION -------------------
app = Flask(__name__)
app.config['SECRET_KEY'] = 'rootTanmay'
# ---------------------xxx-------------------------------


UPLOAD_FOLDER = 'static/uploads/'
# ALLOWED_EXTENSIONS = {'pdf', 'png', 'jpg', 'jpeg', 'gif'}
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['PDF_STORAGE_PATH'] = '/var/www/icswebapp/icswebapp/static/pdf_application_form/'

# -------------- Bilingual Configuration -------------
app.config['SUPPORTED_LANGUAGES'] = ['en', 'mr']

serializer = URLSafeTimedSerializer('SECRET_KEY')

# ------------ DATABASE CONFIGURATION -------------------
hostserver = '43.240.64.151'
localserver = 'localhost'
host = localserver

cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',  # --------  DATABASE CONNECTION
                              host=host,
                              database='ICSApplication')
cursor = cnx.cursor()
# --------------- END DATABASE ---------------------------


# ------------ MAIL CONFIGURATION -------------------
app.config['MAIL_SERVER'] = 'us2.smtp.mailhostbox.com'
app.config['MAIL_PORT'] = 587
app.config['MAIL_USERNAME'] = 'helpdesk@trti-maha.in'  # --------  E-MAIL CONNECTION
app.config['MAIL_PASSWORD'] = 'FOtIEzp9'
app.config['MAIL_USE_TLS'] = True
app.config['MAIL_USE_SSL'] = False
app.config['MAIL_DEBUG'] = True  # Enable debugging
mail = Mail(app)
# ---------------------xxx-------------------------------


# ---------------------- Language Array ------------------
multilingual_content = {
    'english': {
        # ---------- Title -------------
        'title_home': 'Fellowship | Home',
        # ----- Top Menu -----
        'skip-to-main-content': 'Skip to Main Content',
        'mobile-no': '020-2633 2380, 020-2636 0941',
        'a-': 'A-',
        'a': 'A',
        'a+': 'A+',
        'english': 'English',
        'marathi': 'Marathi',
        'trti_header': 'TRIBAL RESEARCH & TRAINING INSTITUTE GOVERNMENT OF MAHARASHTRA',
        'admin_login_btn': 'Admin Login',
        'fellowship': 'Fellowship',

        # -----------------------Leaders ----------------------
        'eknath_shinde': 'Shri. Eknath Shinde',
        'eknath_shinde_desgn': "Hon'ble Chief Minister, Maharashtra State",
        'dev_phad': 'Shri. Devendra Fadnavis',
        'dev_phad_desgn': "'Hon'ble Deputy Chief Minister, Maharashtra State",
        'ajit_pawar': 'Shri. Ajit Pawar',
        'ajit_pawar_desgn': "Hon'ble Deputy Chief Minister, Maharashtra State",
        'vijay_gavit': 'Dr. Vijaykumar Gavit',
        'vijay_gavit_desgn': "Hon'ble Minister, Tribal Development Department",
        'vijay_waghmare': "Shri. Vijay Waghmare I.A.S.",
        'vijay_waghmare_desgn': "Hon'ble Additional Chief Secretary, Tribal Development Department",
        'rajendra_bharud': "Dr. Rajendra Bharud I.A.S.",
        'rajendra_bharud_desgn': "Hon'ble Commissioner, Tribal Research and Training Institute, Pune",

        # --------------- Menubar -----------------
        'home': 'Home',
        'about': "About",
        'government_regulation': 'Government Regulation',
        'contact_us': 'Contact Us',
        'charts': 'Reports',
        'login_sign_btn': 'Login | Sign up',

        # ------------------ Banner Section ---------------------
        'total_applications': "Total Application",
        'total_applications_2022': "Total Application 2022",
        'total_applications_2021': "Total Application 2021",
        'fellowship_awarded': 'Fellowship Awarded',
        'fellowship_completed': 'Fellowship Completed',
        'this_yrs_application': 'This Year Applications',
        'get_your_fellowship_now': 'Get Your Fellowship Now',
        'enroll_here': 'Enroll Here',

        # ------------- About ------------------
        'about_the_institute': 'About The Institution',
        'about_title': 'Tribal Research and Training Institute, Maharashtra State, Pune',
        'read_more': "Read More",
        'criteria': "Criteria",
        'scheme_benefits': 'Scheme Benefits',
        'eligibility_criteria': 'Eligibility Criteria',
        'updates': 'Updates',
        'how_to_enroll': 'How to enroll?',
        'step': 'Step',
        'step_1_content': 'Create Your Login or Sign on site',
        'step_2_content': 'Fill Application Form',
        'step_3_content': 'Get Selected by Admin',
        'step_4_content': 'Upload Joining Letter of PhD',
        'step_5_content': 'Get your Fellowship',

        'faq': 'Frequently Asked Questions',
        'q1': 'How many Fellows do you select?',
        'ans1': 'We hire 100 students every year',
        'q2': 'I wasn’t selected, can I receive feedback on my application?',
        'ans2': 'We will send you feedback through email, and it will be reflected on your dashboard as well.',
        'related_links': 'Related Links',
        'security_brands': 'SECURITY & BRANDS',
        'address': '28,  Bund Garden Rd, Near Old Circuit House, Camp, Pune, Maharashtra 411001 ',
        'phone': 'Phone',
        'email': 'Email',
        'national_portal_of_india': 'National Portal of India',
        'tribal_development_department': 'Tribal development Department',
        'tribal_commissionerate': 'Tribal Commissionerate',
        'aaple_sarkar': 'Aaple Sarkar',
        'ministry_of_tribal_affairs': 'Ministry of Tribal Affairs- Govt of India',
        'my_gov': 'My Gov',
        'security_brand': 'SECURITY & BRAND',
        'hyperlink_policy': 'Hyperlink Policy',
        'terms_and_conditions': 'Terms and Conditions',
        'privacy_policy': 'Privacy Policy',
        'copyright_policy': 'Copyright Policy',
        'website_monitoring_policy': 'Website Monitoring Policy',
        'sitemap': 'Sitemap',
        'footer_text': 'Website Content Owned & Managed by Tribal Research And Training Institute | MoTA | Maharashtra, India',
        'desgn_dev': 'Designed and Developed by',
        'ics': 'Integrated Consultancy Services ',
        'you_name': 'Your Name',
        'you_email': 'Your Email',
        'subject': 'Subject',
        'message': 'Message',
        'send_message': 'Send Message',
        'address': 'Address',
        'address_text': '28, BUND GARDEN RD, NEAR OLD CIRCUIT HOUSE, CAMP, PUNE, MAHARASHTRA 411001',
        'email_us': 'Email us',
        'open_hours': 'Open hours',
        'monday_friday': 'Monday - Friday'

    },
    'marathi': {
        # ----------- Title -------------
        'title_home': 'अधिछात्रवृत्ती | मुख्यपृष्ठ',

        # ----- Top Menu -----
        'skip-to-main-content': 'मुख्य सामग्रीवर जा',
        'mobile-no': '०२०-२६३३ २३८०, ०२०-२६३६ ०९४१',
        'a-': 'ए-',
        'a': 'ए',
        'a+': 'ए+',
        'english': 'इंग्रजी',
        'marathi': 'मराठी',
        'trti_header': 'आदिवासी संशोधन व प्रशिक्षण संस्था, महाराष्ट्र शासन',
        'admin_login_btn': 'प्रशासक लॉगिन',
        'fellowship': 'अधिछात्रवृत्ती',

        # -----------------------Leaders ----------------------
        'eknath_shinde': 'श्री. एकनाथ शिंदे',
        'eknath_shinde_desgn': "माननीय मुख्यमंत्री, महाराष्ट्र राज्य",
        'dev_phad': 'श्री. देवेंद्र फडणवीस',
        'dev_phad_desgn': "माननीय उपमुख्यमंत्री, महाराष्ट्र राज्य",
        'ajit_pawar': 'श्री. अजित पवार',
        'ajit_pawar_desgn': 'माननीय उपमुख्यमंत्री, महाराष्ट्र राज्य',
        'vijay_gavit': 'डॉ. विजयकुमार गावित',
        'vijay_gavit_desgn': "माननीय मंत्री, आदिवासी विकास विभाग",
        'vijay_waghmare': "श्री. विजय वाघमारे भा.प्र.से",
        'vijay_waghmare_desgn': "माननीय अतिरिक्त मुख्य सचिव, आदिवासी विकास विभाग",
        'rajendra_bharud': "डॉ. राजेंद्र भारूड भा.प्र.से.",
        'rajendra_bharud_desgn': "माननीय आयुक्त्त, आदिवासी संशोधन व प्रशिक्ष",

        # --------------- Menubar -----------------
        'home': 'मुख्यपृष्ठ',
        'about': "आमच्याविषयी",
        'government_regulation': 'सरकारी नियमन',
        'contact_us': 'संपर्क',
        'charts': 'अहवाल',
        'login_sign_btn': 'लॉगिन | साइन अप करा',

        # ------------------ Banner Section ---------------------
        'total_applications': "एकूण अर्ज",
        'total_applications_2022': "एकूण अर्ज 2022",
        'total_applications_2021': "एकूण अर्ज 2021",
        'fellowship_awarded': 'अधिछात्रवृत्ती प्रदान',
        'fellowship_completed': 'अधिछात्रवृत्ती पूर्ण केलेले',
        'this_yrs_application': 'या वर्षाचे अर्ज',
        'enroll_here': 'येथे नावनोंदणी करा',
        'get_your_fellowship_now': 'आता आपली अधिछात्रवृत्ती मिळवा',

        # ----------------------------- About institute ---------------------
        'about_the_institute': 'संस्थेबद्दल',
        'about_title': 'आदिवासी संशोधन व प्रशिक्षण संस्था, महाराष्ट्र राज्य, पुणे',
        'read_more': "पुढे वाचा",

        # ---------------------------
        'criteria': "निकष",
        'scheme_benefits': 'योजनेचे फायदे',
        'eligibility_criteria': 'पात्रता निकष',
        'updates': 'अपडेट्स',
        'how_to_enroll': 'नोंदणी कशी करावी?',
        'step': 'स्टेप',
        'step_1_content': 'साइटवर आपले लॉगिन तयार करा किंवा साइन इन करा',
        'step_2_content': 'अर्ज भरा',
        'step_3_content': 'प्रशासकाद्वारे निवड मिळवा',
        'step_4_content': 'पीएचडीचे सामील पत्र अपलोड करा',
        'step_5_content': 'तुमची फेलोशिप मिळवा',

        'faq': 'वारंवार विचारले जाणारे प्रश्न',
        'q1': 'तुम्ही किती विद्यार्थी निवडता?',
        'ans1': 'आम्ही दरवर्षी 100 विद्यार्थी निवडतो',
        'q2': 'मी निवडलो नाही, मला अभिप्राय मिळू शकेल का माझ्या अर्जावर?',
        'ans2': 'आम्ही आपल्याला ईमेलद्वारे अभिप्राय पाठवू आणि ते आपल्या डॅशबोर्डवर देखील प्रतिबिंबित होईल.',
        'related_links': 'संबंधित दुवे',
        'security_brands': 'सुरक्षा आणि ब्रँड',

        'address': '२८, बंद गार्डन रोड, जवळ जुन्या सर्किट हाऊस, कॅम्प, पुणे, महाराष्ट्र ४११००१',
        'phone': 'फोन',
        'email': 'ईमेल',
        'national_portal_of_india': 'भारताचा राष्ट्रीय पोर्टल',
        'tribal_development_department': 'आदिवासी विकास विभाग',
        'tribal_commissionerate': 'आदिवासी आयुक्त',
        'aaple_sarkar': 'आपले सरकार',
        'ministry_of_tribal_affairs': 'आदिवासी प्रकरण मंत्रालय- भारत सरकार',
        'my_gov': 'माय गोव्ह',
        'security_brand': 'सुरक्षा आणि ब्रँड',
        'hyperlink_policy': 'हायपरलिंक धोरण',
        'terms_and_conditions': 'नियम आणि अटी',
        'privacy_policy': 'गोपनीयता धोरण',
        'copyright_policy': 'कॉपीराइट धोरण',
        'website_monitoring_policy': 'वेबसाइट देखरेख धोरण',
        'sitemap': 'साइटमॅप',
        'footer_text': 'वेबसाइट सामग्री आदिवासी संशोधन आणि प्रशिक्षण संस्थेच्या मालकीची आणि व्यवस्थापित | जनजातीय कार्य मंत्रालय | महाराष्ट्र शासन, भारत',
        'desgn_dev': ' डिझाइन आणि विकसित बाय',
        'ics': 'इंटिग्रेटएड कन्सल्टन्सी सरवीसेस',
        'you_name': 'तुमचे नाव',
        'you_email': 'आपला ई - मेल',
        'subject': 'विषय',
        'message': 'संदेश',
        'send_message': 'संदेश पाठवा',
        'address': 'पत्ता',
        'call_us': 'आम्हाला कॉल करा',
        'trti': 'आदिवासी संशोधन आणि प्रशिक्षण संस्था',
        'address_text': '28, बंड गार्डन आरडी, ओल्ड  सर्किट हाऊसजवळ, कॅम्प, पुणे, महाराष्ट्र 411001',
        'email_us': 'आम्हाला ईमेल करा',
        'open_hours': 'उघडे तास',
        'monday_friday': 'सोमवार-शुक्रवार'
    },
    # Add more languages and content as needed.
}


# --------------------- Set Session ---------------
@app.route('/set_session/<value>')
def set_session(value):
    session['language'] = value
    return redirect(request.referrer)


# ---------------- HOME PAGE ---------------------------
@app.route('/', methods=['GET', 'POST'])
def home_page():
    if 'language' in session:
        language = session['language']
    else:
        language = 'marathi'

    # --------------------------  HOME PAGE
    total_count = applications_today()
    old_user_21 = old_users_count_2021()
    old_user_22 = old_users_count_2022()
    print(total_count)
    news_record = news()
    return render_template('home-page.html', total_count=total_count, old_user_21=old_user_21, old_user_22=old_user_22,
                           language=language, multilingual_content=multilingual_content, news_record=news_record)


@app.route('/charts')
def reports():
    if 'language' in session:
        language = session['language']
    else:
        language = 'marathi'
    twentyone_count = year_twentyone_count()
    twentytwo_count = year_twentytwo_count()
    twentythree_count = year_twentythree_count()
    male_count = male_count_report()
    female_count = female_count_report()
    trans_count = trans_count_report()
    disability_yes = disability_yes_count_report()
    disability_no = disability_no_count_report()
    return render_template('report_homepage.html', twentyone_count=twentyone_count,
                           language=language, twentytwo_count=twentytwo_count, twentythree_count=twentythree_count,
                           male_count=male_count, female_count=female_count, multilingual_content=multilingual_content,
                           trans_count=trans_count, disability_yes=disability_yes, disability_no=disability_no)


# ---------------- END HOME PAGE -----------------------


def send_email_rejection(email, full_name, status, applicant_id):
    # Email content in HTML format
    msg = Message('Application Status Changed', sender='icstrti@outlook.com', recipients=[email])
    msg.body = msg.body = "Hi, " + full_name + "\n Your Status for Fellowship : " + status + \
                          "\n Unfortunately the status of your application has changed to Rejected!!" + \
                          "\n Please login to view the status as Accepted for Fellowship" + \
                          "\n Your Application ID = " + applicant_id
    mail.send(msg)


# ----------------- ABOUT US ---------------------------
@app.route('/aboutus')
def aboutus():
    if 'language' in session:
        language = session['language']
    else:
        language = 'marathi'

    return render_template('AboutUs.html', language=language, multilingual_content=multilingual_content)


# ----------------- END ABOUT US -----------------------


# ----------------- GOVERNMENT RELATIONS ---------------
@app.route('/gr')
def gr():  # --------------------------  GOVERNMENT RELATIONS PAGE
    return render_template('gr.html')


# ----------------- END GOVERNMENT RELATIONS ----------


# ----------------- CONTACT US ---------------------------
@app.route('/contact')
def contact_us():  # --------------------------  CONTACT US PAGE
    if 'language' in session:
        language = session['language']
    else:
        language = 'marathi'
    # Create a folium map centered at Pune, India
    m = folium.Map(location=[18.5204, 73.8567], zoom_start=12)
    # Add a marker at Pune, India
    folium.Marker(location=[18.5204, 73.8567], popup="Pune, India").add_to(m)
    # Render the map in the template
    map_html = m._repr_html_()
    # Pass the map HTML to the template
    return render_template('contact.html', map=map_html, language=language, multilingual_content=multilingual_content)


# ----------------- END CONTACT US -----------------------


# ----------------- LOGIN --------------------------------
def new_applicant_incomplete_form(email, form_filled):  # -------------- CHECK IF USER HAS FILLED THE FORM
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor()
    sql = "SELECT form_filled FROM application_page WHERE email = %s AND form_filled = %s;"
    cursor.execute(sql, (email, form_filled))
    result = cursor.fetchone()
    cursor.close()
    cnx.close()
    return result is not None


def is_form_filled(email):  # -------------- CHECK IF USER HAS FILLED THE FORM
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor()
    sql = "SELECT form_filled='1' FROM application_page WHERE email=%s"
    cursor.execute(sql, (email,))
    result = cursor.fetchone()
    cursor.close()
    cnx.close()
    return result is not None


def check_final_approval(email):  # ----------- CHECK IF USER IS FINALLY APPROVED
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor()
    sql = "SELECT final_approval FROM application_page WHERE email = %s AND final_approval = 'accepted'"
    cursor.execute(sql, (email,))
    result = cursor.fetchone()
    cursor.close()
    cnx.close()
    print('Final Approval Result:', result)  # Add this line for debugging
    return result is not None


def get_id_by_email(email):  # --------------- GET ID for EMAIL IN SESSION
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor()
    sql = """ SELECT id FROM application_page WHERE email=%s"""
    cursor.execute(sql, (email,))
    result = cursor.fetchone()
    print(result[0])
    cursor.close()
    cnx.close()
    return result[0]


def is_withdrawn(email):
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor()
    sql = """ SELECT fellowship_withdrawn='withdrawn' FROM signup WHERE email=%s"""
    cursor.execute(sql, (email,))
    result = cursor.fetchone()
    print(result[0])
    cursor.close()
    cnx.close()
    return result[0]


def old_user(email):
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor()
    # Corrected SQL query
    sql = """SELECT * FROM old_users WHERE phd_registration_year IN ('2021', '2022') AND email = %s"""
    cursor.execute(sql, (email,))
    result = cursor.fetchone()
    if result:
        print(result[0])
    cursor.close()
    cnx.close()
    return result[0] if result else None


@app.route('/login', methods=['GET', 'POST'])
def login():  # --------------------------  LOGIN PAGE
    if 'language' in session:
        language = session['language']
    else:
        language = 'marathi'
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor(dictionary=True)

    if request.method == 'POST':
        email = request.form['email']
        password = request.form['password']

        # Check if email and password are valid from signup table
        sql = "SELECT email, password, first_name, year FROM signup WHERE email=%s AND password=%s"
        cursor.execute(sql, (email, password))
        user = cursor.fetchone()
        if user:
            print(f'Email: {email}')  # Add this line to check the value of email
            session['email'] = email
            session['user_name'] = user['first_name']
            if is_withdrawn(email):
                error = 'You Have withdrawn from Fellowship. Please contact Us to report your issue.'
                return render_template('login.html', error=error)
            elif old_user(email):
                return redirect(url_for('preview'))
            elif new_applicant_incomplete_form(email, 0):
                print('New Applicant and Incomplete form')
                return redirect(url_for('section1'))
            elif check_final_approval(email):
                session['final_approval'] = "accepted"
                print('I have checked approval function')
                return redirect(url_for('main_page'))
            elif is_form_filled(email):
                session['final_approval'] = "pending"
                id = get_id_by_email(email)
                # print(id)
                print('Checked viewform function')
                return redirect(url_for('viewform', id=id))
            else:
                print('User is a new applicant')
                return redirect(url_for('section1'))
        else:
            error = 'Invalid Email or Password. Please enter valid Credentials'
            return render_template('login.html', error=error, language=language,
                                   multilingual_content=multilingual_content)
    return render_template('login.html', language=language, multilingual_content=multilingual_content)


# @app.route('/application_form', methods=['GET', 'POST'])
# def render_application_form():                                     # ---------------- RENDER APPLICATION FORM
#     email = session['email']
#     cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
#                                   host=host,
#                                   database='ICSApplication')
#     cursor = cnx.cursor(dictionary=True)
#     sql = """SELECT * FROM old_users WHERE email = %s"""
#     cursor.execute(sql, (email,))
#     # Fetch all records matching the query
#     records = cursor.fetchall()
#     print(records)
#     # Close the cursor and database connection
#     cursor.close()
#     cnx.close()
#     return render_template('application-form.html', records=records)


# VIEW STUDENT FORM
@app.route('/viewform/<int:id>')
def viewform(id):  # -------------- VIEW STUDENT FORM
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor(dictionary=True)
    sql = """SELECT * FROM application_page WHERE id = %s"""
    cursor.execute(sql, (id,))
    # Fetch all records matching the query
    records = cursor.fetchall()
    print(records)
    # Close the cursor and database connection
    cursor.close()
    cnx.close()
    return render_template('form-view.html', records=records)


@app.route('/mainpage')
def main_page():  # -------------- APPLICATION LIST PAGE
    email = session['email']
    print(" User Email: " + email)
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor(dictionary=True)
    sql = """SELECT first_name, middle_name, last_name, email, applicant_id, application_date, id, 
    phd_registration_date, phd_registration_year, adhaar_number FROM application_page WHERE email = %s"""
    cursor.execute(sql, (email,))
    # Fetch all records matching the query
    records = cursor.fetchall()
    # Close the cursor and database connection
    cursor.close()
    cnx.close()
    return render_template('application-list.html', records=records)


# ----------------- END LOGIN -----------------------


# ----------------- SIGN UP --------------------------------
def is_user_registered(email):  # ---------------- CHECK IF EMAIL IS IN THE DATABASE
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor()
    sql = "SELECT verified FROM signup WHERE email = %s"
    cursor.execute(sql, (email,))
    result = cursor.fetchone()
    cursor.close()
    cnx.close()
    return result


def send_email_verification(email, first_name, otp):
    msg_body = f'''
               <!DOCTYPE html>
           <html lang="en">

           <head>
               <meta charset="UTF-8">
               <meta name="viewport" content="width=device-width,initial-scale=1">
               <title>Congratulations</title>
               <link rel="preconnect" href="https://fonts.googleapis.com">
               <link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
               <link href="https://fonts.googleapis.com/css2?family=Poppins:ital,wght@0,100;0,200;0,300;0,400;0,500;0,600;0,700;0,800;1,100;1,200;1,300;1,400;1,500;1,600;1,700;1,800&display=swap" rel="stylesheet">
               <link rel="preconnect" href="https://fonts.googleapis.com">
               <link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
               <link href="https://fonts.googleapis.com/css2?family=Montserrat:ital,wght@0,100;0,200;0,300;0,400;0,500;0,600;0,700;0,800;0,900;1,200;1,300;1,400;1,500;1,600;1,700;1,800;1,900&family=Roboto:ital,wght@0,100;0,300;0,400;0,500;0,700;0,900;1,100;1,300;1,400;1,500;1,700;1,900&display=swap" rel="stylesheet">
               <link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.4.2/css/all.min.css" integrity="sha512-z3gLpd7yknf1YoNbCzqRKc4qyor8gaKU1qmn+CShxbuBusANI9QpRohGBreCFkKxLhei6S9CQXFEbbKuqLg0DA==" crossorigin="anonymous" referrerpolicy="no-referrer">
               <link rel="preconnect" href="https://fonts.googleapis.com">
               <link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
               <link href="https://fonts.googleapis.com/css2?family=Charm:wght@400;700&display=swap" rel="stylesheet">
           </head>

           <body style="font-family: Montserrat, sans-serif; text-align: center; margin: 0; padding: 0;">
               <table style="width: 100%; background-color: transparent;">
                   <tr>
                       <td style="padding: 0; width: 60%; margin: auto; border-radius: 5px; box-shadow: 0 5px 15px #55efc4; margin-top: 100px; position: relative; overflow: hidden; background-color: transparent !important;">
                           <table style="width: 100%;">
                               <tr>
                                   <td style="position: relative; margin-bottom: 80px; height: 40vh; width: 40vh; margin: auto;">
                                       <img src="http://fellowship.trti-maha.in/static/Images/award.png" class="image_award" alt="image_award" style="width: 300px; height: 300px; object-fit: contain; position: absolute; width: 30vh; height: 50vh; top: 0; left: -50; transform: translateX(-50%); z-index: 100;">

                                   </td>
                               </tr>
                               <tr>
                                   <td style="background-color: #fff; border-radius: 5px; z-index: 10 !important; background: 0 0; background-color: transparent !important;">
                                       <h1 style="font-family: Charm, cursive; font-size: 49px; padding: 0; font-weight: 400; margin-top: -20px; color: #0c2e8a; margin:0 !important;"></h1>
                                       <h2 style="margin: 0; font-size: 32px; color: #0c2e8a; text-transform: uppercase; font-weight: 700; position: relative; padding-bottom: 3px; margin-top: 20px; border-bottom: 3px solid #50d8af; width: fit-content; margin: auto;"></h2>
                                       <p style="color: #0c2e8a; font-weight: 700; line-height: 30px; text-transform: capitalize;">Dear,  {first_name} </p>
                                       <p style="color: #0c2e8a; font-weight: 700; line-height: 30px; text-transform: capitalize;">Thank you for your interest in creating a user account for Online Fellowship Portal System</p>
                                       <p style="color: #0c2e8a; font-weight: 700; line-height: 30px; text-transform: capitalize;">To activate your account, please confirm your email address by entering the OTP given below.</p>
                                       <div style="width: 100%; margin-top: 30px;"><a style="padding: 10px 20px; text-decoration: none; color: #0c2e8a; font-weight: 700; background-color: #50d8af; border-radius: 5px; border: 1px solid #0c2e8a; display: inline-block; letter-spacing: 5px;">{otp}</a></div>
                                       <p style="color: #0c2e8a; font-weight: 700; line-height: 30px; text-transform: capitalize;">In case of any technical issue while filling online application form, please contact on toll 
    free helpline number 18002330444 (9:45 am to 6:30 pm) Monday to Friday</p>     
                                   </td>
                               </tr>
                           </table>
                       </td>
                   </tr>
               </table>
           </body>

           </html>

           '''
    msg = Message('Verify Email', sender='helpdesk@trti-maha.in', recipients=[email])
    msg.html = msg_body
    mail.send(msg)


@app.route('/signup', methods=['GET', 'POST'])
def signup():  # --------------------------  SIGN UP PAGE
    if request.method == 'POST':
        first_name = request.form['first_name']
        middle_name = request.form['middle_name']
        last_name = request.form['last_name']
        email = request.form['email']
        password = request.form['password']
        confirm_password = request.form['confirm_password']
        year = request.form['year']

        # Check if the passwords match
        if password != confirm_password:
            return render_template('login.html', error='Passwords do not match',
                                   multilingual_content=multilingual_content)

        if old_user(email):
            error = (
                'Please Login with the registered email ID and Password for your login will be Fellowship@XXXX (Last four digits of your registred Mobile Number).'
                'You can change the password once you login')
            return render_template('login.html', error=error, multilingual_content=multilingual_content)

        if is_user_registered(email):
            error = ('This email is already registered. Please use a different email or log in with an existing one.',
                     'error')
            return render_template('login.html', error=error, multilingual_content=multilingual_content)
        else:
            unique_id = random.randint(100000, 999999)
            global otp
            otp = random.randint(100000, 999999)

            # Store user registration data in a session for verification
            session['registration_data'] = {
                'first_name': first_name,
                'middle_name': middle_name,
                'last_name': last_name,
                'email': email,
                'password': password,
                'confirm_password': confirm_password,
                'year': year,
                'unique_id': unique_id
            }

            send_email_verification(email, first_name, otp)
            return render_template('email_verify.html', email=email)
    return render_template('login.html', multilingual_content=multilingual_content)


# Define a function to insert user registration data into the database
def insert_user_data(registration_data):
    try:
        cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                      host=host,
                                      database='ICSApplication')
        cursor = cnx.cursor()

        # Define your INSERT SQL statement with %s placeholders
        sql = "INSERT INTO signup (first_name, middle_name, last_name, email, password, confirm_password, year, unique_id) " \
              "VALUES (%s, %s, %s, %s, %s, %s, %s, %s)"

        # Extract the data from the 'registration_data' dictionary
        data = (
            registration_data['first_name'],
            registration_data['middle_name'],
            registration_data['last_name'],
            registration_data['email'],
            registration_data['password'],
            registration_data['confirm_password'],
            registration_data['year'],
            registration_data['unique_id']
        )

        # Execute the SQL statement with the data
        cursor.execute(sql, data)

        # Commit the changes to the database
        cnx.commit()

        # Close the cursor and database connection
        cursor.close()
        cnx.close()

        return True  # Return True to indicate a successful insertion

    except mysql.connector.Error as err:
        print("MySQL Error: {}".format(err))
        return False  # Return False to indicate an error occurred during insertion


@app.route('/email_verify', methods=['GET', 'POST'])
def email_verify():
    if 'registration_data' not in session:
        flash('Session data not found. Please sign up again.', 'error')
        return redirect(url_for('signup'))

    user_otp = request.form['otp']

    if otp == int(user_otp):
        registration_data = session.get('registration_data')
        insert_user_data(registration_data)
        flash('Your email is verified and registration is successful.')
        return redirect(url_for('login'))
    else:
        error = 'You have entered the wrong OTP. Please enter the OTP again sent to your email'
        return render_template('email_verify.html', error=error)


# ----------------- END SIGN UP -----------------------


# ----------------- ADMIN LOGIN -----------------------
@app.route('/adminlogin', methods=['GET', 'POST'])
def admin_login():  # ------------------ ADMIN LOGIN
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor()

    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']

        # Check if the user is an admin
        sql = "SELECT * FROM admin WHERE username=%s AND password=%s"
        cursor.execute(sql, (username, password))
        user = cursor.fetchone()
        cnx.commit()

        if user:
            # Set session variable to indicate user is logged in
            session['logged_in'] = True
            # Close the connection and cursor
            cursor.close()
            cnx.close()
            return redirect(url_for('index'))
        else:
            error = 'Invalid username or password'
            # Close the connection and cursor
            cursor.close()
            cnx.close()
            return render_template('adminlogin.html', error=error)
    return render_template('adminlogin.html')


# -------------------- END ADMIN LOGIN ------------------------------


# --------------------- INDEX - ADMIN DASHBOARD PAGE ---------------
# @app.route('/index')
# def index():                                                        # ---------------- ADMIN DASHBOARD - INDEX PAGE
#     return render_template('index.html')
# ----------------- END INDEX - ADMIN DASHBOARD PAGE -------------------


# --------------------- USERS PROFILE USER SIDE PAGE ---------------
@app.route('/user_profile', methods=['GET', 'POST'])
def user_profile():  # ----------------  USERS PROFILE USER SIDE PAGE
    email = session['email']

    if request.method == 'POST':
        if 'edit_profile' in request.form:
            submit_edit_profile()  # Call the edit profile function
        elif 'change_password' in request.form:
            change_password_user()  # Call the change password function
            flash("Password changed successfully")  # Flash a success message

    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor(dictionary=True)

    sql = """SELECT * FROM application_page WHERE email = %s"""
    cursor.execute(sql, (email,))
    records = cursor.fetchall()

    cursor.close()
    cnx.close()

    return render_template('users-profile.html', records=records)


def submit_edit_profile():  # ------------- SUBMIT EDIT PROFILE FOR MY PROFILE
    if request.method == 'POST':
        first_name = request.form['first_name']
        address = request.form['add_1']
        phone = request.form['mobile_number']
        email = session['email']
        cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                      host=host,
                                      database='ICSApplication')
        cursor = cnx.cursor(dictionary=True)
        # Construct the SQL update query
        sql = f"UPDATE application_page SET first_name = '{first_name}', add_1 = '{address}', mobile_number = '{phone}' WHERE email = '{email}'"
        cursor.execute(sql)
        cnx.commit()
        cursor.close()
        cnx.close()


def change_password_user():  # -------------- CHANGE PASSWORD FOR USER
    if request.method == 'POST':
        current_password = request.form['current_password']
        print(current_password)
        new_password = request.form['new_password']
        confirm_password = request.form['confirm_password']
        email = session['email']
        cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                      host=host,
                                      database='ICSApplication')
        cursor = cnx.cursor(dictionary=True)

        query = 'SELECT password FROM signup WHERE email = %s'
        print("SQL Query:", query)
        cursor.execute(query, (email,))
        result = cursor.fetchone()

        if result:
            # Get the stored hashed password from the result dictionary
            stored_password = result['password']

            # Hash the entered current password for comparison
            hashed_current_password = hashlib.sha256(current_password.encode()).hexdigest()

            if stored_password == hashed_current_password:
                # Passwords match, update the password
                hashed_new_password = hashlib.sha256(new_password.encode()).hexdigest()
                sql = f"UPDATE signup SET password = '{hashed_new_password}', confirm_password = '{confirm_password}' WHERE email = '{email}'"
                cursor.execute(sql)
                cnx.commit()
        else:
            print("User not found")


# ----------------- END  USERS PROFILE USER SIDE PAGE  -------------------


#                                                ---- FINALLY APPROVED USER FUNCTIONALITY
# ---------------- USER AFTER FINAL APPROVAL -----------------------------
@app.route('/manage_profile_AA')
def manage_profile():  # ---------- MANAGE PROFILE
    email = session['email']

    if request.method == 'POST':
        if 'edit_profile' in request.form:
            submit_edit_profile()  # Call the edit profile function
        elif 'change_password' in request.form:
            change_password_user()  # Call the change password function
            flash("Password changed successfully")  # Flash a success message

    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor(dictionary=True)

    sql = """SELECT * FROM application_page WHERE email = %s"""
    cursor.execute(sql, (email,))
    records = cursor.fetchall()

    cursor.close()
    cnx.close()

    return render_template('users-profile.html', records=records)


def submit_edit_profile():  # ------------- SUBMIT EDIT PROFILE FOR MY PROFILE
    if request.method == 'POST':
        first_name = request.form['first_name']
        address = request.form['add_1']
        phone = request.form['mobile_number']
        email = session['email']
        cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                      host=host,
                                      database='ICSApplication')
        cursor = cnx.cursor(dictionary=True)
        # Construct the SQL update query
        sql = f"UPDATE application_page SET first_name = '{first_name}', add_1 = '{address}', mobile_number = '{phone}' WHERE email = '{email}'"
        cursor.execute(sql)
        cnx.commit()
        cursor.close()
        cnx.close()


def change_password_user():  # -------------- CHANGE PASSWORD FOR USER
    if request.method == 'POST':
        current_password = request.form['current_password']
        print(current_password)
        new_password = request.form['new_password']
        confirm_password = request.form['confirm_password']
        email = session['email']
        cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                      host=host,
                                      database='ICSApplication')
        cursor = cnx.cursor(dictionary=True)

        query = 'SELECT password FROM signup WHERE email = %s'
        print("SQL Query:", query)
        cursor.execute(query, (email,))
        result = cursor.fetchone()

        if result:
            # Get the stored hashed password from the result dictionary
            stored_password = result['password']

            # Hash the entered current password for comparison
            hashed_current_password = hashlib.sha256(current_password.encode()).hexdigest()

            if stored_password == hashed_current_password:
                # Passwords match, update the password
                hashed_new_password = hashlib.sha256(new_password.encode()).hexdigest()
                sql = f"UPDATE signup SET password = '{hashed_new_password}', confirm_password = '{confirm_password}' WHERE email = '{email}'"
                cursor.execute(sql)
                cnx.commit()
        else:
            print("User not found")
    return render_template('users-profile.html')


# -------------------- END MANAGE PROFILE ---------------------------------


# --------------------- AWARD LETTER ----------------------------------
@app.route('/award_letter_AA')
def award_letter_AA():
    return render_template('award_letter_AA.html')


# -------------------- END AWARD LETTER --------------------------------


# --------------------- JOINING LETTER ----------------------------------
@app.route('/joining_report_AA', methods=['GET', 'POST'])
def joining_report_AA():
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor(dictionary=True)
    records = None
    joining_report = None
    joining_date = None  # Initialize with a default value
    email = session.get('email', None)

    # Fetch the joining date from the database
    cursor.execute("SELECT joining_date FROM application_page WHERE email = %s", (email,))
    result = cursor.fetchone()

    if result:
        joining_date = result['joining_date']

    if request.method == 'POST':
        joining_report = request.files['joining_report']
        joining_report_path = save_file(joining_report)
        sql = """INSERT INTO application_page (email, joining_date, joining_report) 
                                 VALUES (%s, %s, %s)"""
        cursor.execute(sql, (email, joining_date, joining_report_path))
        cnx.commit()

    cursor.close()
    cnx.close()
    return render_template('joining_report_AA.html', records=records, joining_date=joining_date,
                           joining_report=joining_report)


# -------------------- END JOINING LETTER --------------------------------


# --------------------- PRESENTY LETTER ----------------------------------
@app.route('/presenty_AA')
def presenty_AA():
    return render_template('presenty_AA.html')


# -------------------- END PRESENTY LETTER --------------------------------


# --------------------- HALF YEARLY REPORT ----------------------------------
@app.route('/half_yearly_rep_AA', methods=['GET', 'POST'])
def half_yearly_rep_AA():
    if 'email' not in session:
        return redirect('/login')

    email = session['email']
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor(dictionary=True)
    existing_reports = []  # Initialize existing_reports here

    if request.method == 'POST':
        # Handle file uploads and save them to the database
        report_paths = []
        for i in range(1, 11):
            report = request.files.get(f'half_yearly_report{i}')
            if report:
                # Save the uploaded report to a directory
                # You can use your own logic to save the report and get the file path
                report_path = save_file(report)
                report_paths.append(report_path)

        # Update the database with the report paths
        for i, report_path in enumerate(report_paths, start=1):
            cursor.execute(f"UPDATE application_page SET half_yearly_report{i} = %s WHERE email = %s",
                           (report_path, email))
        cnx.commit()

    # Fetch the saved reports for the user
    cursor.execute(
        f"SELECT half_yearly_report1, half_yearly_report2, half_yearly_report3, half_yearly_report4, half_yearly_report5, "
        f"half_yearly_report6, half_yearly_report7, half_yearly_report8, half_yearly_report9, half_yearly_report10 "
        f"FROM application_page WHERE email = %s",
        (email,))
    reports = cursor.fetchone()
    # Count the number of submitted reports
    submitted_count = sum([1 for i in range(1, 6) if reports[f'half_yearly_report{i}']])

    # Fetch the joining date for the user
    cursor.execute("SELECT joining_date FROM application_page WHERE email = %s", (email,))
    result = cursor.fetchone()

    if result:
        joining_date = result['joining_date']
        start_dates = [joining_date + timedelta(days=i * 30 * 6) for i in range(10)]
        end_dates = [start_date + timedelta(days=30 * 6) for start_date in start_dates]
    else:
        joining_date = None
        start_dates = []
        end_dates = []

    cursor.close()
    cnx.close()
    return render_template('half_yearly_rep_AA.html', reports=reports, joining_date=joining_date,
                           start_dates=start_dates, end_dates=end_dates, submitted_count=submitted_count)


# -------------------- END HALF YEARLY REPORT --------------------------------


# --------------------- HRA RENT AGREEMENT ----------------------------------
@app.route('/rent_agreement_AA', methods=['GET', 'POST'])
def rent_agreement_AA():
    if 'email' not in session:
        # Redirect to the login page if the user is not logged in
        return redirect('/login')

    email = session['email']
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor(dictionary=True)
    existing_reports = []  # Initialize existing_reports here

    if request.method == 'POST':
        # Handle file uploads and save them to the database
        report_paths = []
        for i in range(1, 6):
            report = request.files.get(f'rent_agreement{i}')
            if report:
                # Save the uploaded report to a directory
                # You can use your own logic to save the report and get the file path
                report_path = save_file(report)
                report_paths.append(report_path)

        # Update the database with the report paths
        for i, report_path in enumerate(report_paths, start=1):
            cursor.execute(f"UPDATE application_page SET rent_agreement{i} = %s WHERE email = %s",
                           (report_path, email))
        cnx.commit()

    # Fetch the saved reports for the user
    cursor.execute(
        f"SELECT rent_agreement1, rent_agreement2, rent_agreement3, rent_agreement4, rent_agreement5 FROM application_page WHERE email = %s",
        (email,))
    reports = cursor.fetchone()
    # Count the number of submitted reports
    submitted_count = sum([1 for i in range(1, 6) if reports[f'rent_agreement{i}']])

    # Fetch the joining date for the user
    cursor.execute("SELECT joining_date FROM application_page WHERE email = %s", (email,))
    result = cursor.fetchone()

    if result:
        joining_date = result['joining_date']
        start_dates = [joining_date + timedelta(days=i * 365) for i in range(5)]
        end_dates = [start_date + timedelta(days=365) for start_date in start_dates]
    else:
        joining_date = None
        start_dates = []
        end_dates = []

    cursor.close()
    cnx.close()

    return render_template('rent_agreement_AA.html', reports=reports, joining_date=joining_date,
                           start_dates=start_dates, end_dates=end_dates, submitted_count=submitted_count)


# -------------------- END HRA RENT AGREEMENT --------------------------------


# --------------------- WITHDRAW FELLOWSHIP ----------------------------------
@app.route('/with_from_fellowship_AA')
def with_from_fellowship_AA():
    return render_template('with_from_fellowship_AA.html')


# -------------------- END WITHDRAW FELLOWSHIP --------------------------------


# --------------------- CHANGE GUIDE ----------------------------------
@app.route('/change_guide_AA', methods=['GET', 'POST'])
def change_guide_AA():
    if 'email' not in session:
        # Redirect to the login page if the user is not logged in
        return redirect('/login')

        # Update the database to change the name of the guide
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')

    if request.method == 'POST':
        # Get the user's email from the session
        email = session['email']

        new_guide_name = request.form.get('new_guide_name')
        cursor = cnx.cursor()

        # Execute an SQL query to update the giude's name
        update_query = "UPDATE application_page SET name_of_guide = %s WHERE email = %s"
        cursor.execute(update_query, (new_guide_name, email))

        # Commit the transaction and close the cursor and connection
        cnx.commit()
        cursor.close()
        return redirect('/change_guide_AA')

    cursor = cnx.cursor()

    # SQL query to fetch the present guide name from database
    cursor.execute("SELECT name_of_guide FROM application_page")
    result = cursor.fetchone()

    if result is not None:
        guide_name = result[0]
    else:
        guide_name = "No Guide Found"

    # cursor.close()
    cnx.close()

    return render_template('change_guide_AA.html', guide_name=guide_name)


# -------------------- CHANGE GUIDE  --------------------------------


# --------------------- CHANGE CENTER ----------------------------------
@app.route('/change_center_AA', methods=['GET', 'POST'])
def change_center_AA():
    if 'email' not in session:
        # Redirect to the login page if the user is not logged in

        return redirect('/login')

        # Update the database to change the name of the center
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')

    if request.method == 'POST':
        # Get the user's email from the session
        email = session['email']
        new_center_name = request.form.get('new_center_name')
        cursor = cnx.cursor()
        update_query = "UPDATE application_page SET name_of_college = %s WHERE email = %s"
        cursor.execute(update_query, (new_center_name, email))

        # Commit the transaction and close the cursor and connection
        cnx.commit()
        cursor.close()
        return redirect('/change_center_AA')
    cursor = cnx.cursor()

    # SQL query to fetch the present center name from database
    cursor.execute("SELECT name_of_college FROM application_page")
    result = cursor.fetchone()

    if result is not None:
        center_name = result[0]
    else:
        center_name = "No Center Found"

    # cursor.close()
    cnx.close()
    return render_template('change_center_AA.html', center_name=center_name)


# -------------------- CHANGE CENTER  ---------------------------------


# --------------------- PHD AWARD ----------------------------------
@app.route('/phd_award_AA')
def phd_award_AA():
    return render_template('phd_award_AA.html')


# -------------------- END PHD AWARD -----------------------------


# --------------------- UPLOAD THESIS ----------------------------------
@app.route('/upload_thesis_AA')
def upload_thesis_AA():
    return render_template('upload_thesis_AA.html')


# -------------------- END UPLOAD THESIS -----------------------------


def save_file(file):
    if file and allowed_file(file.filename):
        filename = secure_filename(file.filename)
        file_path = os.path.join(UPLOAD_FOLDER, filename)
        file.save(file_path)


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in {'png', 'jpg', 'jpeg', 'gif', 'bmp', 'pdf'}


#                                               ---- ADMIN LEVEL 1, 2, 3
# ------------------ ADMIN LEVEL 1 ----------------
@app.route('/adminPage', methods=['GET', 'POST'])
def admin_level_1():
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor(dictionary=True)

    if request.method == 'POST':
        # Check if a form was submitted
        if 'accept' in request.form:
            # Handle accepting an application
            applicant_id = request.form['accept']
            update_status_admin(applicant_id, 'accepted')
        elif 'reject' in request.form:
            # Handle rejecting an application
            applicant_id = request.form['reject']
            update_status_admin(applicant_id, 'rejected')
            cursor.execute("SELECT email, first_name, status FROM application_page WHERE applicant_id = %s",
                           (applicant_id,))
            user_data = cursor.fetchone()

            if user_data:
                email = user_data[0]
                full_name = f"{user_data[1]} {user_data[2]}"

                send_email_rejection(email, full_name, 'Rejected', applicant_id)
        # Commit the changes to the database
        cnx.commit()

    cursor.execute("SELECT * FROM application_page where form_filled='1'")
    data = cursor.fetchall()
    print(data)
    cursor.close()
    cnx.close()
    return render_template('admin_level_one.html', data=data)


# ------------------ END ADMIN LEVEL 1 ----------------


# ----------------- ADMIN LEVEL 2 ----------------------
@app.route('/level_two_admin', methods=['GET', 'POST'])
def level_two_admin():
    if not session.get('logged_in'):
        # Redirect to the admin login page if the user is not logged in
        return redirect(url_for('admin_login'))
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor(dictionary=True)
    if request.method == 'POST':
        # Check if a form was submitted
        if 'accept' in request.form:
            # Handle accepting an application
            applicant_id = request.form['accept']
            update_scrutiny_admin(applicant_id, 'accepted')
        elif 'reject' in request.form:
            # Handle rejecting an application
            applicant_id = request.form['reject']
            update_scrutiny_admin(applicant_id, 'rejected')
            cursor.execute("SELECT email, first_name, scrutiny_status FROM application_page WHERE applicant_id = %s",
                           (applicant_id,))
            user_data = cursor.fetchall()

            if user_data:
                email = user_data[0]
                full_name = f"{user_data[1]} {user_data[2]}"

                # Send an email to the user
                send_email_rejection(email, full_name, 'Rejected', applicant_id)
        # Commit the changes to the database
        cnx.commit()

    cursor.execute("SELECT * FROM application_page WHERE status='accepted'")
    data = cursor.fetchall()
    print(data)
    cursor.close()
    cnx.close()
    return render_template('admin_level_two.html', data=data)


# ------------------- END ADMIN LEVEL 2 ------------------


# -------------------- ADMIN LEVEL 3 ----------------------
@app.route('/level_three_admin', methods=['GET', 'POST'])
def level_three_admin():
    if not session.get('logged_in'):
        # Redirect to the admin login page if the user is not logged in
        return redirect(url_for('admin_login'))
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor(dictionary=True)
    final_approval = request.get_data()
    if request.method == 'POST':
        # Check if a form was submitted
        if 'accept' in request.form:
            # Handle accepting an application
            applicant_id = request.form['accept']
            today = date.today()
            print("Today's date:", today)
            update_final_appr_admin(applicant_id, 'accepted', today)
            # Fetch the applicant's email and full name from the database
            cursor.execute(
                "SELECT email, first_name, last_name, final_approval FROM application_page WHERE applicant_id = %s",
                (applicant_id,))
            user_data = cursor.fetchone()

            if user_data:
                email = user_data['email']
                full_name = f"{user_data['first_name']} {user_data['last_name']}"

                # Send an email to the user
                send_email_approval(email, full_name, 'Accepted', applicant_id)

        elif 'reject' in request.form:
            # Handle rejecting an application
            applicant_id = request.form['reject']
            today = date.today()
            update_final_appr_admin(applicant_id, 'rejected', today)
            cursor.execute("SELECT email, first_name, final_approval FROM application_page WHERE applicant_id = %s",
                           (applicant_id,))
            user_data = cursor.fetchone()

            if user_data:
                email = user_data[0]
                full_name = f"{user_data[1]} {user_data[2]}"

                # Send an email to the user
                send_email_approval(email, full_name, 'Rejected', applicant_id)
        # Commit the changes to the database
        cnx.commit()

    cursor.execute("SELECT * FROM application_page WHERE scrutiny_status='accepted'")
    data = cursor.fetchall()
    print(data)
    cursor.close()
    cnx.close()
    return render_template('admin_level_three.html', data=data)


# ------------------ END ADMIN LEVEL 3 -------------------


# ----------------- SEND EMAIL REJECTION & ACCEPTANCE -------------------
def send_email_rejection(email, full_name, status, applicant_id):
    # Email content in HTML format
    msg = Message('Application Status Changed', sender='icstrti@outlook.com', recipients=[email])
    msg.body = msg.body = "Hi, " + full_name + "\n Your Status for Fellowship : " + status + \
                          "\n Unfortunately the status of your application has changed to Rejected!!" + \
                          "\n Please login to view the status as Accepted for Fellowship" + \
                          "\n Your Application ID = " + applicant_id
    mail.send(msg)


def send_email_approval(email, full_name, status, applicant_id):
    base_url = request.url_root
    # email_body = render_template('email_template.html', full_name=full_name, status=status,  applicant_id=applicant_id)
    # Construct the HTML email body
    msg_body = f'''
        <!DOCTYPE html>
    <html lang="en">

    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width,initial-scale=1">
        <title>Congratulations</title>
        <link rel="preconnect" href="https://fonts.googleapis.com">
        <link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
        <link href="https://fonts.googleapis.com/css2?family=Poppins:ital,wght@0,100;0,200;0,300;0,400;0,500;0,600;0,700;0,800;1,100;1,200;1,300;1,400;1,500;1,600;1,700;1,800&display=swap" rel="stylesheet">
        <link rel="preconnect" href="https://fonts.googleapis.com">
        <link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
        <link href="https://fonts.googleapis.com/css2?family=Montserrat:ital,wght@0,100;0,200;0,300;0,400;0,500;0,600;0,700;0,800;0,900;1,200;1,300;1,400;1,500;1,600;1,700;1,800;1,900&family=Roboto:ital,wght@0,100;0,300;0,400;0,500;0,700;0,900;1,100;1,300;1,400;1,500;1,700;1,900&display=swap" rel="stylesheet">
        <link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.4.2/css/all.min.css" integrity="sha512-z3gLpd7yknf1YoNbCzqRKc4qyor8gaKU1qmn+CShxbuBusANI9QpRohGBreCFkKxLhei6S9CQXFEbbKuqLg0DA==" crossorigin="anonymous" referrerpolicy="no-referrer">
        <link rel="preconnect" href="https://fonts.googleapis.com">
        <link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
        <link href="https://fonts.googleapis.com/css2?family=Charm:wght@400;700&display=swap" rel="stylesheet">
    </head>

    <body style="font-family: Montserrat, sans-serif; text-align: center; margin: 0; padding: 0;">
        <table style="width: 100%; background-color: transparent;">
            <tr>
                <td style="padding: 0; width: 60%; margin: auto; border-radius: 5px; box-shadow: 0 5px 15px #55efc4; margin-top: 100px; position: relative; overflow: hidden; background-color: transparent !important;">
                    <table style="width: 100%;">
                        <tr>
                            <td style="position: relative; margin-bottom: 80px; height: 40vh; width: 40vh; margin: auto;">
                                <img src="http://fellowship.trti-maha.in/static/Images/award.png" class="image_award" alt="image_award" style="width: 300px; height: 300px; object-fit: contain; position: absolute; width: 30vh; height: 50vh; top: 0; left: -50; transform: translateX(-50%); z-index: 100;">

                            </td>
                        </tr>
                        <tr>
                            <td style="background-color: #fff; border-radius: 5px; z-index: 10 !important; background: 0 0; background-color: transparent !important;">
                                <h1 style="font-family: Charm, cursive; font-size: 49px; padding: 0; font-weight: 400; margin-top: -20px; color: #0c2e8a; margin:0 !important;">Congratulations</h1>
                                <h2 style="margin: 0; font-size: 32px; color: #0c2e8a; text-transform: uppercase; font-weight: 700; position: relative; padding-bottom: 3px; margin-top: 20px; border-bottom: 3px solid #50d8af; width: fit-content; margin: auto;">Hello, {full_name}</h2>
                                <p style="color: #0c2e8a; font-weight: 700; line-height: 30px; text-transform: capitalize;">Congratulations the status of your application has changed to Accepted!!</p>
                                <p style="color: #0c2e8a; font-weight: 700; line-height: 30px; text-transform: capitalize;">Please login to view the status as Accepted for Fellowship</p>
                                <p style="color: #0c2e8a; font-weight: 700; line-height: 30px; text-transform: capitalize;">Your Application ID: </p>
                                <div style="width: 100%; margin-top: 30px;"><a style="padding: 10px 20px; text-decoration: none; color: #0c2e8a; font-weight: 700; background-color: #50d8af; border-radius: 5px; border: 1px solid #0c2e8a; display: inline-block; letter-spacing: 5px;">{applicant_id}</a></div>
                            </td>
                        </tr>
                    </table>
                </td>
            </tr>
        </table>
    </body>

    </html>

    '''

    # Email content in HTML format
    msg = Message('Application Status Changed', sender='helpdesk@trti-maha.in', recipients=[email])
    msg.html = msg_body
    mail.send(msg)


# --------------------- END SEND EMAIL --------------------------


# --------------------- UPDATE STATUS ADMIN -------------------------
# STATUS UPDATE ON STUDENT RECORDS FUNCTIONALITY
# ----------------------------------------------
def update_status_admin(applicant_id, status):
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor()

    # Update the status for the specified applicant ID
    update_query = "UPDATE application_page SET status = %s WHERE applicant_id = %s"
    cursor.execute(update_query, (status, applicant_id))

    # Commit the changes to the database
    cnx.commit()

    # Close the cursor and database connection
    cursor.close()
    cnx.close()


def update_scrutiny_admin(applicant_id, scrutiny_status):
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor()

    # Update the status for the specified applicant ID
    update_query = "UPDATE application_page SET scrutiny_status = %s WHERE applicant_id = %s"
    cursor.execute(update_query, (scrutiny_status, applicant_id))

    # Commit the changes to the database
    cnx.commit()

    # Close the cursor and database connection
    cursor.close()
    cnx.close()


def update_final_appr_admin(applicant_id, final_approval, joining_date):
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor()

    # Update the status for the specified applicant ID
    update_query = "UPDATE application_page SET final_approval = %s, joining_date=%s WHERE applicant_id = %s"
    cursor.execute(update_query, (final_approval, joining_date, applicant_id))

    # Commit the changes to the database
    cnx.commit()

    # Close the cursor and database connection
    cursor.close()
    cnx.close()


# ------------------------- END UPDATE STATUS ADMIN ----------------------------


@app.route('/export_data', methods=['GET', 'POST'])
def export_data():
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor()
    cursor.execute("SELECT * FROM application_page")
    data = cursor.fetchall()
    # Create a workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Add headers to the worksheet
    ws.append(['applicant_photo', 'applicant_id', 'first_name', 'middle_name', 'last_name', 'mobile_number', 'email',
               'date_of_birth', 'gender', 'age', 'caste', 'your_caste', 'marital_status', 'dependents', 'add_1',
               'add_2',
               'pincode', 'village', 'taluka', 'district', 'state', 'qualification_level', 'phd_registration_date',
               'concerned_university', 'topic_of_phd', 'name_of_guide', 'name_of_college', 'stream', 'board_university',
               'admission_year', 'passing_year', 'percentage', 'family_annual_income', 'income_certificate_number',
               'issuing_authority', 'domicile', 'domicile_certificate', 'relation', 'domicile_number', 'caste',
               'caste_category', 'caste_certificate_number', 'issuing_district', 'caste_applicant_name',
               'caste_issuing_authority', 'salaried', 'disability', 'father_alive', 'father_name', 'mother_alive',
               'mother_name', 'work_in_government', 'bank_name', 'account_number', 'ifsc_code', 'account_holder_name',
               'documentfile1', 'documentfile2', 'documentfile3', 'documentfile4', 'documentfile5', 'documentfile6',
               'documentfile7', 'documentfile8', 'documentfile9', 'current_date', 'formatted_datetime'])

    # Add data to the worksheet
    for row in data:
        ws.append(row)

    # Save the workbook in memory as bytes
    data = BytesIO()
    wb.save(data)
    data.seek(0)

    # Create a response object and attach the workbook as a file
    response = make_response(data.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=ApplicantData.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    return response


@app.route('/export')
def export():
    export_data()


#                                           ------- ADMIN PAYMENT SHEET
@app.route('/payment_sheet', methods=['GET', 'POST'])
def payment_sheet():
    user_records = []
    if request.method == 'GET':
        # Establish a database connection
        cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                      host=host,
                                      database='ICSApplication')
        cursor = cnx.cursor(dictionary=True)
        print('I have made connection')

        # Fetch user data based on the email
        cursor.execute("SELECT * FROM application_page where final_approval='accepted'")
        user_data = cursor.fetchall()  # Use fetchall to retrieve all rows
        print(user_data)

        for row in user_data:
            # Calculate values based on user data
            applicant_id = row['applicant_id']
            faculty = row["faculty"]
            joining_date = row["phd_registration_date"]
            city = row['city']
            print(joining_date)
            print(faculty)
            # Calculate Count Yearly
            if faculty == "Arts":
                count_yearly = 20500
            elif faculty == "Science":
                count_yearly = 25000
            else:
                count_yearly = 0  # Handle other faculty values as needed

            if city in ['AHMEDABAD', 'BENGALURU', 'CHENNAI', 'DELHI', 'HYDERABAD', 'KOLKATA', 'MUMBAI', 'PUNE']:
                rate = '24%'
            elif city in ['AGRA', 'AJMER', 'ALIGARH', 'AMRAVATI', 'AMRITSAR', 'ANAND', 'ASANSOL', 'AURANGABAD',
                          'BAREILLY', 'BELAGAVI', 'BRAHMAPUR', 'BHAVNAGAR', 'BHIWANDI', 'BHOPAL', 'BHUBANESWAR',
                          'BIKANER', 'BILASPUR', 'BOKARO STEEL CITY', 'BURDWAN', 'CHANDIGARH', 'COIMBATORE', 'CUTTACK',
                          'DAHOD', 'DEHRADUN', 'DOMBIVLI', 'DHANBAD', 'BHILAI', 'DURGAPUR', 'ERODE', 'FARIDABAD',
                          'GHAZIABAD', 'GORAKHPUR', 'GUNTUR', 'GURGAON', 'GUWAHATI', 'GWALIOR', 'HAMIRPUR',
                          'HUBBALLI–DHARWAD', 'INDORE', 'JABALPUR', 'JAIPUR', 'JALANDHAR', 'JALGAON', 'JAMMU',
                          'JAMSHEDPUR', 'JHANSI', 'JODHPUR', 'KALABURAGI', 'KAKINADA', 'KANNUR', 'KANPUR', 'KARNAL',
                          'KOCHI', 'KOLHAPUR', 'KOLLAM', 'KOTA', 'KOZHIKODE', 'KUMBAKONAM', 'KURNOOL', 'LUDHIANA',
                          'LUCKNOW', 'MADURAI', 'MALAPPURAM', 'MATHURA', 'MANGALURU', 'MEERUT', 'MORADABAD', 'MYSURU',
                          'NAGPUR', 'NANDED', 'NADIAD', 'NASHIK', 'NELLORE', 'NOIDA', 'PATNA', 'PUDUCHERRY', 'PURLIA',
                          'PRAYAGRAJ', 'RAIPUR', 'RAJKOT', 'RAJAMAHENDRAVARAM', 'RANCHI', 'ROURKELA', 'RATLAM',
                          'SAHARANPUR', 'SALEM', 'SANGLI', 'SHIMLA', 'SILIGURI', 'SOLAPUR', 'SRINAGAR', 'SURAT',
                          'THANJAVUR', 'THIRUVANANTHAPURAM', 'THRISSUR', 'TIRUCHIRAPPALLI', 'TIRUNELVELI',
                          'TIRUVANNAMALAI', 'UJJAIN', 'VIJAYAPURA', 'VADODARA', 'VARANASI', 'VASAI-VIRAR', 'VIJAYAWADA',
                          'VISAKHAPATNAM', 'VELLORE', 'WARANGAL']:
                rate = '16%'
            else:
                rate = '8%'

            print("Rate:", rate)

            # Initialize the "from" and "to" date to empty strings
            duration_date_from = ""
            duration_date_to = ""

            if joining_date:  # Check if joining_date is not None
                # Calculate Duration Date (adding 3 months to joining date)
                duration_date_from = joining_date.strftime('%Y-%m-%d')  # "from" date is the joining date
                duration_date_to = (joining_date + timedelta(days=90)).strftime('%Y-%m-%d')  # 90 days = 3 months

            # Calculate Total Months
            total_months = 3

            # Calculate Fellowship
            fellowship = 31000  # Fixed value for 3 months

            # Calculate Total Fellowship
            total_fellowship = fellowship * total_months

            rate_str = float(rate.rstrip('%'))
            convert_rate = (rate_str / 100)
            hra_amount = convert_rate * fellowship

            months = total_months

            total_hra = hra_amount * months

            total = total_fellowship + total_hra

            # Create a record dictionary for the user
            record = {
                "applicant_id": row['applicant_id'],
                "full_name": row['first_name'] + ' ' + row['last_name'],
                "first_name": row['first_name'],
                "last_name": row['last_name'],
                "email": row["email"],
                "faculty": row['faculty'],
                "joining_date": row['joining_date'],
                "city": row['city'],
                "duration": duration_date_from + '' + duration_date_to,
                "rate": rate,
                "count": count_yearly,
                "amount": hra_amount,
                "months": months,
                "total_hra": total_hra,
                "total": total,
                "duration_date_from": duration_date_from,
                "duration_date_to": duration_date_to,
                "total_months": total_months,
                "fellowship": fellowship,
                "to_fellowship": total_fellowship,
                "phd_registration_year": row['phd_registration_year'],
                "id": row['id']
            }
            print(record)
            user_records.append(record)

            email = record['email']
            print(email)

            if email:
                print("Existing Record:", email)
                # Record already exists, do not insert again
            else:
                print("Record not found, proceeding with the INSERT query")
                # Insert values into the payment_sheet table
                cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                              host=host,
                                              database='ICSApplication')
                cursor = cnx.cursor()

                insert_query = """
                    INSERT INTO payment_sheet (
                        full_name, faculty, city, duration_date_from, duration_date_to,
                        rate, count, amount, months, total_hra, total,
                        total_months, fellowship,
                        to_fellowship, email
                    )
                    VALUES (%(full_name)s, %(faculty)s, %(city)s, %(rate)s, %(count)s,
                            %(amount)s, %(months)s, %(total_hra)s, %(total)s,
                            %(duration_date_from)s, %(duration_date_to)s,
                            %(total_months)s, %(fellowship)s, %(to_fellowship)s, %(email)s)
                """
                # Execute the INSERT query
                cursor.execute(insert_query, record)

                # Commit the changes to the database
                cnx.commit()

            # Close the database cursor and connection
            cursor.close()
            cnx.close()
        # Close the database cursor and connection
        cursor.close()
        cnx.close()

    return render_template('payment_sheet.html', user_records=user_records)


@app.route('/fellowship_details/<string:email>', methods=['GET', 'POST'])
def fellowship_details(email):
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor(dictionary=True)
    cursor.execute("SELECT * FROM application_page where email=%s ", (email,))
    result = cursor.fetchall()
    print("result" + str(result))

    cursor.execute("SELECT * FROM payment_sheet where email=%s ", (email,))
    record = cursor.fetchall()
    print("record" + str(record))

    cursor.execute("SELECT fellowship_withdrawn FROM signup where email=%s ", (email,))
    output = cursor.fetchall()
    print("record" + str(output))

    cnx.commit()
    cursor.close()
    cnx.close()
    return render_template('fellowship_details.html', result=result, record=record, output=output)


#                                                     ---- ADMIN DASHBOARD
@app.route('/index', methods=['GET', 'POST'])
def index():
    session['user_name'] = 'Admin'
    total_count = applications_today()
    accepted_count = accepted_applications()
    rejected_count = rejected_applications()
    male_count = male_count_report()
    female_count = female_count_report()
    trans_count = trans_count_report()

    twentyone_count = year_twentyone_count()
    twentytwo_count = year_twentytwo_count()
    twentythree_count = year_twentythree_count()
    olduser_count = old_users_count_2021()
    olduser22_count = old_users_count_2022()

    disability_yes = disability_yes_count_report()
    disability_no = disability_no_count_report()
    physical_disability = type_disability_physically()
    visually_disability = type_disability_visually()
    hearing_disability = type_disability_hearing()
    male_report = male_record_report()
    female_report = female_record_report()
    priority_people_caste = priority_by_caste()
    news_record = news()

    if request.method == 'POST':
        user = request.form['user']
        title = request.form['title']
        subtitle = request.form['subtitle']
        cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                      host=host,
                                      database='ICSApplication')
        cursor = cnx.cursor()

        sql = """INSERT INTO news_and_updates(user, title, subtitle) VALUES(%s, %s, %s) """
        data = (user, title, subtitle)
        cursor.execute(sql, data)
        # Commit the changes
        cnx.commit()
        # Close the cursor and connection
        cursor.close()
        cnx.close()
        return render_template('index.html', total_count=total_count, accepted_count=accepted_count,
                               rejected_count=rejected_count,
                               male_count=male_count, female_count=female_count, trans_count=trans_count,
                               disability_yes=disability_yes,
                               disability_no=disability_no, male_report=male_report, female_report=female_report,
                               priority_people_caste=priority_people_caste,
                               twentyone_count=twentyone_count, twentytwo_count=twentytwo_count,
                               olduser_count=olduser_count,
                               olduser22_count=olduser22_count, twentythree_count=twentythree_count,
                               news_record=news_record,
                               physical_disability=physical_disability, hearing_disability=hearing_disability,
                               visually_disability=visually_disability)
    return render_template('index.html', total_count=total_count, accepted_count=accepted_count,
                           rejected_count=rejected_count,
                           male_count=male_count, female_count=female_count, trans_count=trans_count,
                           disability_yes=disability_yes,
                           disability_no=disability_no, male_report=male_report, female_report=female_report,
                           priority_people_caste=priority_people_caste,
                           twentyone_count=twentyone_count, twentytwo_count=twentytwo_count,
                           olduser_count=olduser_count,
                           olduser22_count=olduser22_count, twentythree_count=twentythree_count,
                           news_record=news_record,
                           physical_disability=physical_disability, hearing_disability=hearing_disability,
                           visually_disability=visually_disability)


# ---------------------  DISABILITY FUNCTIONALITY ---------------------------------
# ---------------------------------------------------------------------------------
@app.route('/disability_report_yes', methods=['GET', 'POST'])
def disability_report_yes():
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor(dictionary=True)
    cursor.execute("SELECT * FROM application_page WHERE disability='Yes'")
    result = cursor.fetchall()
    cnx.commit()
    cursor.close()
    cnx.close()
    print(result)
    return render_template('disability_report_yes.html', result=result)


# ----------------------  EXPORT TO EXCEL Applicants with Disability -------------------------------------------
@app.route('/export_disability_report_yes', methods=['GET', 'POST'])
def export_disability_report_yes():
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor()
    cursor.execute("SELECT * from application_page WHERE disability ='Yes' ")
    data = cursor.fetchall()
    # Create a workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Add headers to the worksheet
    ws.append(['applicant_id', 'first_name', 'middle_name', 'last_name', 'mobile_number', 'email',
               'date_of_birth', 'gender', 'age', 'caste', 'your_caste', 'marital_status', 'dependents', 'add_1',
               'add_2', 'pincode', 'village', 'taluka', 'district', 'state', 'qualification_level',
               'phd_registration_date',
               'concerned_university', 'topic_of_phd', 'name_of_guide', 'name_of_college', 'stream', 'board_university',
               'admission_year', 'passing_year', 'percentage', 'family_annual_income', 'income_certificate_number',
               'issuing_authority', 'domicile', 'domicile_certificate', 'relation', 'domicile_number', 'caste',
               'caste_category', 'caste_certificate_number', 'issuing_district', 'caste_applicant_name',
               'caste_issuing_authority', 'salaried', 'disability', 'father_alive', 'father_name', 'mother_alive',
               'mother_name', 'work_in_government', 'bank_name', 'account_number', 'ifsc_code', 'account_holder_name'])

    # Add data to the worksheet
    for row in data:
        ws.append(row)

    # Save the workbook in memory as bytes
    data = BytesIO()
    wb.save(data)
    data.seek(0)

    # Create a response object and attach the workbook as a file
    response = make_response(data.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=Applicants_with_disability.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    return response


@app.route('/disability_report_no', methods=['GET', 'POST'])
def disability_report_no():
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor(dictionary=True)
    cursor.execute("SELECT * FROM application_page WHERE disability='No'")
    result = cursor.fetchall()
    cnx.commit()
    cursor.close()
    cnx.close()
    print(result)
    return render_template('disability_report_no.html', result=result)


# ----------------------  EXPORT TO EXCEL Applicants WITHOUT Disability -------------------------------------------
@app.route('/export_disability_report_no', methods=['GET', 'POST'])
def export_disability_report_no():
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor()
    cursor.execute("SELECT * from application_page WHERE disability ='No' ")
    data = cursor.fetchall()
    # Create a workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Add headers to the worksheet
    ws.append(['applicant_id', 'first_name', 'middle_name', 'last_name', 'mobile_number', 'email',
               'date_of_birth', 'gender', 'age', 'caste', 'your_caste', 'marital_status', 'dependents', 'add_1',
               'add_2', 'pincode', 'village', 'taluka', 'district', 'state', 'qualification_level',
               'phd_registration_date',
               'concerned_university', 'topic_of_phd', 'name_of_guide', 'name_of_college', 'stream', 'board_university',
               'admission_year', 'passing_year', 'percentage', 'family_annual_income', 'income_certificate_number',
               'issuing_authority', 'domicile', 'domicile_certificate', 'relation', 'domicile_number', 'caste',
               'caste_category', 'caste_certificate_number', 'issuing_district', 'caste_applicant_name',
               'caste_issuing_authority', 'salaried', 'disability', 'father_alive', 'father_name', 'mother_alive',
               'mother_name', 'work_in_government', 'bank_name', 'account_number', 'ifsc_code', 'account_holder_name'])

    # Add data to the worksheet
    for row in data:
        ws.append(row)

    # Save the workbook in memory as bytes
    data = BytesIO()
    wb.save(data)
    data.seek(0)

    # Create a response object and attach the workbook as a file
    response = make_response(data.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=Applicants_without_disability.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    return response


def disability_yes_count_report():  # ----- To count users with NO disability  ------
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor()
    cursor.execute("SELECT COUNT(*) FROM application_page WHERE disability='Yes'")
    result = cursor.fetchone()
    print(result)
    return result[0]


def disability_no_count_report():  # ----- To count users with NO disability  ------
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor()
    cursor.execute("SELECT COUNT(*) FROM application_page WHERE disability='no'")
    result = cursor.fetchone()
    print(result)
    return result[0]


def type_disability_physically():  # ----- To count users with NO disability  ------
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor()
    cursor.execute("SELECT COUNT(*) FROM application_page WHERE type_of_disability='Physically Handicapped'")
    result = cursor.fetchone()
    print(result)
    return result[0]


def type_disability_visually():  # ----- To count users with NO disability  ------
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor()
    cursor.execute("SELECT COUNT(*) FROM application_page WHERE type_of_disability='Visually Impaired'")
    result = cursor.fetchone()
    print(result)
    return result[0]


def type_disability_hearing():  # ----- To count users with NO disability  ------
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor()
    cursor.execute("SELECT COUNT(*) FROM application_page WHERE type_of_disability='Hearing Impaired'")
    result = cursor.fetchone()
    print(result)
    return result[0]


@app.route('/disabilty_report_physical', methods=['GET', 'POST'])
def disabilty_report_physical():
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor(dictionary=True)
    cursor.execute("SELECT * FROM application_page WHERE type_of_disability='Physically Handicapped'")
    result = cursor.fetchall()
    cnx.commit()
    cursor.close()
    cnx.close()
    print(result)
    return render_template('disabilty_report_physical.html', result=result)


# ----------------------  EXPORT TO EXCEL Physically Handicapped Applicants -------------------------------------------
@app.route('/export_physically_handicapped_report', methods=['GET', 'POST'])
def export_physically_handicapped_report():
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor()
    cursor.execute("SELECT * from application_page WHERE type_of_disability='Physically Handicapped' ")
    data = cursor.fetchall()
    # Create a workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Add headers to the worksheet
    ws.append(['applicant_id', 'first_name', 'middle_name', 'last_name', 'mobile_number', 'email',
               'date_of_birth', 'gender', 'age', 'caste', 'your_caste', 'marital_status', 'dependents', 'add_1',
               'add_2', 'pincode', 'village', 'taluka', 'district', 'state', 'qualification_level',
               'phd_registration_date',
               'concerned_university', 'topic_of_phd', 'name_of_guide', 'name_of_college', 'stream', 'board_university',
               'admission_year', 'passing_year', 'percentage', 'family_annual_income', 'income_certificate_number',
               'issuing_authority', 'domicile', 'domicile_certificate', 'relation', 'domicile_number', 'caste',
               'caste_category', 'caste_certificate_number', 'issuing_district', 'caste_applicant_name',
               'caste_issuing_authority', 'salaried', 'disability', 'father_alive', 'father_name', 'mother_alive',
               'mother_name', 'work_in_government', 'bank_name', 'account_number', 'ifsc_code', 'account_holder_name'])

    # Add data to the worksheet
    for row in data:
        ws.append(row)

    # Save the workbook in memory as bytes
    data = BytesIO()
    wb.save(data)
    data.seek(0)

    # Create a response object and attach the workbook as a file
    response = make_response(data.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=Physically_handicapped_applicants.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    return response


@app.route('/disabilty_report_visual', methods=['GET', 'POST'])
def disabilty_report_visual():
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor(dictionary=True)
    cursor.execute("SELECT * FROM application_page WHERE type_of_disability='Visually Impaired'")
    result = cursor.fetchall()
    cnx.commit()
    cursor.close()
    cnx.close()
    print(result)
    return render_template('disabilty_report_visual.html', result=result)


# ----------------------  EXPORT TO EXCEL Visually Impaired Applicants -------------------------------------------
@app.route('/export_visually_impaired_report', methods=['GET', 'POST'])
def export_visually_impaired_report():
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor()
    cursor.execute("SELECT * from application_page WHERE type_of_disability='Visually Impaired'")
    data = cursor.fetchall()
    # Create a workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Add headers to the worksheet
    ws.append(['applicant_id', 'first_name', 'middle_name', 'last_name', 'mobile_number', 'email',
               'date_of_birth', 'gender', 'age', 'caste', 'your_caste', 'marital_status', 'dependents', 'add_1',
               'add_2', 'pincode', 'village', 'taluka', 'district', 'state', 'qualification_level',
               'phd_registration_date',
               'concerned_university', 'topic_of_phd', 'name_of_guide', 'name_of_college', 'stream', 'board_university',
               'admission_year', 'passing_year', 'percentage', 'family_annual_income', 'income_certificate_number',
               'issuing_authority', 'domicile', 'domicile_certificate', 'relation', 'domicile_number', 'caste',
               'caste_category', 'caste_certificate_number', 'issuing_district', 'caste_applicant_name',
               'caste_issuing_authority', 'salaried', 'disability', 'father_alive', 'father_name', 'mother_alive',
               'mother_name', 'work_in_government', 'bank_name', 'account_number', 'ifsc_code', 'account_holder_name'])

    # Add data to the worksheet
    for row in data:
        ws.append(row)

    # Save the workbook in memory as bytes
    data = BytesIO()
    wb.save(data)
    data.seek(0)

    # Create a response object and attach the workbook as a file
    response = make_response(data.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=Visually_impaired_applicants.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    return response


@app.route('/disabilty_report_hearing', methods=['GET', 'POST'])
def disabilty_report_hearing():
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor(dictionary=True)
    cursor.execute("SELECT * FROM application_page WHERE type_of_disability='Hearing Impaired'")
    result = cursor.fetchall()
    cnx.commit()
    cursor.close()
    cnx.close()
    print(result)
    return render_template('disabilty_report_hearing.html', result=result)


# ----------------------  EXPORT TO EXCEL Hearing Impaired Applicants -------------------------------------------
@app.route('/export_hearing_impaired_report', methods=['GET', 'POST'])
def export_hearing_impaired_report():
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor()
    cursor.execute("SELECT applicant_id, first_name, middle_name, last_name, mobile_number, email, date_of_birth,"
                   "gender, age, caste, your_caste, marital_status, add_1, add_2, pincode, village, taluka, district,"
                   "state, city, qualification_level, phd_registration_date, concerned_university, topic_of_phd, "
                   "name_of_guide, name_of_college, stream, board_university, admission_year, passing_year, percentage,"
                   "family_annual_income, income_certificate_number, issuing_authority, domicile, domicile_certificate,"
                   "domicile_number, caste_certf, issuing_district, caste_issuing_authority, salaried, disability,"
                   "father_name, mother_name, work_in_government, bank_name, account_number, ifsc_code,"
                   "account_holder_name FROM application_page WHERE type_of_disability='Hearing Impaired'")
    data = cursor.fetchall()
    # Create a workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Add headers to the worksheet
    ws.append(['applicant_id', 'first_name', 'middle_name', 'last_name', 'mobile_number', 'email',
               'date_of_birth', 'gender', 'age', 'caste', 'your_caste', 'marital_status', 'dependents', 'add_1',
               'add_2', 'pincode', 'village', 'taluka', 'district', 'state', 'qualification_level',
               'phd_registration_date',
               'concerned_university', 'topic_of_phd', 'name_of_guide', 'name_of_college', 'stream', 'board_university',
               'admission_year', 'passing_year', 'percentage', 'family_annual_income', 'income_certificate_number',
               'issuing_authority', 'domicile', 'domicile_certificate', 'domicile_number', 'caste_certf',
               'issuing_district',
               'caste_issuing_authority', 'salaried', 'disability', 'father_name', 'mother_name', 'work_in_government',
               'bank_name', 'account_number', 'ifsc_code', 'account_holder_name'])

    # Add data to the worksheet
    for row in data:
        ws.append(row)

    # Save the workbook in memory as bytes
    data = BytesIO()
    wb.save(data)
    data.seek(0)

    # Create a response object and attach the workbook as a file
    response = make_response(data.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=Hearing_impaired_applicants.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    return response


# ------------------------------- END DISABILITY -----------------------------------------
# ----------------------------------------------------------------------------------------


# ------------------------------- MALE AND FEMALE FUNCTIONALITY --------------------------
# ----------------------------------------------------------------------------------------
@app.route('/male_record_report', methods=['GET', 'POST'])
def male_record_report():
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor(dictionary=True)
    cursor.execute("SELECT * FROM application_page WHERE gender='male'")
    result = cursor.fetchall()
    cnx.commit()
    cursor.close()
    cnx.close()
    # print(result)
    return render_template('male_record_report.html', result=result)


# ----------------------  EXPORT TO EXCEL Male Applications -------------------------------------------
@app.route('/export_male_applications', methods=['GET', 'POST'])
def export_male_applications():
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor()
    cursor.execute("SELECT applicant_id, first_name, middle_name, last_name, mobile_number, email, date_of_birth,"
                   "gender, age, caste, your_caste, marital_status, add_1, add_2, pincode, village, taluka, district,"
                   "state, city, qualification_level, phd_registration_date, concerned_university, topic_of_phd, "
                   "name_of_guide, name_of_college, stream, board_university, admission_year, passing_year, percentage,"
                   "family_annual_income, income_certificate_number, issuing_authority, domicile, domicile_certificate,"
                   "domicile_number, caste_certf, issuing_district, caste_issuing_authority, salaried, disability,"
                   "father_name, mother_name, work_in_government, bank_name, account_number, ifsc_code,"
                   "account_holder_name FROM application_page WHERE gender ='male' ")
    data = cursor.fetchall()
    # Create a workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Add headers to the worksheet
    ws.append(['applicant_id', 'first_name', 'middle_name', 'last_name', 'mobile_number', 'email',
               'date_of_birth', 'gender', 'age', 'caste', 'your_caste', 'marital_status', 'dependents', 'add_1',
               'add_2', 'pincode', 'village', 'taluka', 'district', 'state', 'qualification_level',
               'phd_registration_date',
               'concerned_university', 'topic_of_phd', 'name_of_guide', 'name_of_college', 'stream', 'board_university',
               'admission_year', 'passing_year', 'percentage', 'family_annual_income', 'income_certificate_number',
               'issuing_authority', 'domicile', 'domicile_certificate', 'domicile_number', 'caste_certf',
               'issuing_district',
               'caste_issuing_authority', 'salaried', 'disability', 'father_name', 'mother_name', 'work_in_government',
               'bank_name', 'account_number', 'ifsc_code', 'account_holder_name'])

    # Add data to the worksheet
    for row in data:
        ws.append(row)

    # Save the workbook in memory as bytes
    data = BytesIO()
    wb.save(data)
    data.seek(0)

    # Create a response object and attach the workbook as a file
    response = make_response(data.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=Male_Applicants_Data.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    return response


@app.route('/female_record_report', methods=['GET', 'POST'])
def female_record_report():
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor(dictionary=True)
    cursor.execute("SELECT * FROM application_page WHERE gender='female'")
    result = cursor.fetchall()
    cnx.commit()
    cursor.close()
    cnx.close()
    print(result)
    return render_template('female_record_report.html', result=result)


# ----------------------  EXPORT TO EXCEL Female Applications -------------------------------------------
@app.route('/export_female_applications', methods=['GET', 'POST'])
def export_female_applications():
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor()
    cursor.execute("SELECT applicant_id, first_name, middle_name, last_name, mobile_number, email, date_of_birth,"
                   "gender, age, caste, your_caste, marital_status, add_1, add_2, pincode, village, taluka, district,"
                   "state, city, qualification_level, phd_registration_date, concerned_university, topic_of_phd, "
                   "name_of_guide, name_of_college, stream, board_university, admission_year, passing_year, percentage,"
                   "family_annual_income, income_certificate_number, issuing_authority, domicile, domicile_certificate,"
                   "domicile_number, caste_certf, issuing_district, caste_issuing_authority, salaried, disability,"
                   "father_name, mother_name, work_in_government, bank_name, account_number, ifsc_code,"
                   "account_holder_name FROM application_page WHERE gender ='female' ")
    data = cursor.fetchall()
    # Create a workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Add headers to the worksheet
    ws.append(['applicant_id', 'first_name', 'middle_name', 'last_name', 'mobile_number', 'email',
               'date_of_birth', 'gender', 'age', 'caste', 'your_caste', 'marital_status', 'dependents', 'add_1',
               'add_2', 'pincode', 'village', 'taluka', 'district', 'state', 'qualification_level',
               'phd_registration_date',
               'concerned_university', 'topic_of_phd', 'name_of_guide', 'name_of_college', 'stream', 'board_university',
               'admission_year', 'passing_year', 'percentage', 'family_annual_income', 'income_certificate_number',
               'issuing_authority', 'domicile', 'domicile_certificate', 'domicile_number', 'caste_certf',
               'issuing_district',
               'caste_issuing_authority', 'salaried', 'disability', 'father_name', 'mother_name', 'work_in_government',
               'bank_name', 'account_number', 'ifsc_code', 'account_holder_name'])

    # Add data to the worksheet
    for row in data:
        ws.append(row)

    # Save the workbook in memory as bytes
    data = BytesIO()
    wb.save(data)
    data.seek(0)

    # Create a response object and attach the workbook as a file
    response = make_response(data.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=Female_Applicants_Data.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    return response


def applications_today():
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor()
    cursor.execute("SELECT COUNT(*) FROM old_users where phd_registration_year='2023'")
    result = cursor.fetchone()
    print(result)
    return result[0]


def accepted_applications():  # ----- To count accepted applications  ------
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor()
    cursor.execute("SELECT COUNT(*) FROM application_page WHERE final_approval='accepted'")
    result = cursor.fetchone()
    print(result)
    return result[0]


def rejected_applications():  # ----- To count rejected applications  ------
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor()
    cursor.execute("SELECT COUNT(*) FROM application_page WHERE final_approval='rejected'")
    result = cursor.fetchone()
    print(result)
    return result[0]


def male_count_report():  # ----- To count male users  ------
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor()
    cursor.execute("SELECT COUNT(*) FROM application_page WHERE gender='male'")
    result = cursor.fetchone()
    print(result)
    return result[0]


def female_count_report():  # ----- To count female users  ------
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor()
    cursor.execute("SELECT COUNT(*) FROM application_page WHERE gender='female'")
    result = cursor.fetchone()
    print(result)
    return result[0]


def trans_count_report():  # ----- To count trans_gender users  ------
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor()
    cursor.execute("SELECT COUNT(*) FROM application_page WHERE gender='other'")
    result = cursor.fetchone()
    print(result)
    return result[0]


def year_twentyone_count():  # ----- To count users from 2021
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor()
    cursor.execute("SELECT COUNT(*) FROM old_users WHERE phd_registration_year='2021'")
    result = cursor.fetchone()
    print(result)
    return result[0]


def year_twentytwo_count():  # ----- To count users from 2022
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor()
    cursor.execute("SELECT COUNT(*) FROM old_users WHERE phd_registration_year='2022'")
    result = cursor.fetchone()
    print(result)
    return result[0]


def year_twentythree_count():  # ----- To count users from 2023
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor()
    cursor.execute("SELECT COUNT(*) FROM old_users WHERE phd_registration_year='2023'")
    result = cursor.fetchone()
    print(result)
    return result[0]


def old_users_count_2021():  # ----- To count old users
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor()
    cursor.execute("SELECT COUNT(*) FROM old_users where phd_registration_year='2021'")
    result = cursor.fetchone()
    print(result)
    return result[0]


def old_users_count_2022():  # ----- To count old users
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor()
    cursor.execute("SELECT COUNT(*) FROM old_users where phd_registration_year='2022'")
    result = cursor.fetchone()
    print(result)
    return result[0]


# ---------------------  Export Old Users list in Excel  -----------------------
@app.route('/export_old_users_applications', methods=['GET', 'POST'])
def export_old_users_applications():
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor()
    cursor.execute("SELECT * FROM old_users")
    data = cursor.fetchall()
    # Create a workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Add headers to the worksheet
    ws.append(
        ['first_name', 'middle_name', 'last_name', 'mobile_number', 'email', 'phd_registration_date', 'day', 'month',
         'year', 'age'])

    # Add data to the worksheet
    for row in data:
        ws.append(row)

    # Save the workbook in memory as bytes
    data = BytesIO()
    wb.save(data)
    data.seek(0)

    # Create a response object and attach the workbook as a file
    response = make_response(data.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=Old_Users_Data.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    return response


def priority_by_caste():  # ----- To count users with caste "Katkari , Kolam , madia"------
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor(dictionary=True)
    cursor.execute("SELECT * FROM application_page WHERE your_caste IN ('katkari', 'kolam', 'madia')")
    result = cursor.fetchall()
    return result


def news():
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor(dictionary=True)
    cursor.execute("SELECT * FROM news_and_updates LIMIT 4")
    result = cursor.fetchall()
    return result


# ----------------------------------------------------------------------------
# ------------------- BEGIN APPLICATIONS 2023 REPORTS ------------------------------
# ----------------------------------------------------------------------------
@app.route('/total_application_report', methods=['GET', 'POST'])
def total_application_report():
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor(dictionary=True)
    cursor.execute("SELECT * FROM application_page")
    result = cursor.fetchall()
    cnx.commit()
    cursor.close()
    cnx.close()
    # print(result)
    return render_template('total_application_report.html', result=result)


# ----------------------  EXPORT TO EXCEL Total applications 2023 -------------------------------------------
@app.route('/export_total_applications_23', methods=['GET', 'POST'])
def export_total_applications_23():
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor()
    cursor.execute("SELECT applicant_id, first_name, middle_name, last_name, mobile_number, email, date_of_birth,"
                   "gender, age, caste, your_caste, marital_status, add_1, add_2, pincode, village, taluka, district,"
                   "state, city, qualification_level, phd_registration_date, concerned_university, topic_of_phd, "
                   "name_of_guide, name_of_college, stream, board_university, admission_year, passing_year, percentage,"
                   "family_annual_income, income_certificate_number, issuing_authority, domicile, domicile_certificate,"
                   "domicile_number, caste_certf, issuing_district, caste_issuing_authority, salaried, disability,"
                   "father_name, mother_name, work_in_government, bank_name, account_number, ifsc_code,"
                   "account_holder_name FROM application_page WHERE phd_registration_year = '2023' ")
    data = cursor.fetchall()
    # Create a workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Add headers to the worksheet
    # ws.append(['applicant_id','email','first_name','last_name','application_date'])

    ws.append(['applicant_id', 'first_name', 'middle_name', 'last_name', 'mobile_number', 'email',
               'date_of_birth', 'gender', 'age', 'caste', 'your_caste', 'marital_status', 'dependents', 'add_1',
               'add_2', 'pincode', 'village', 'taluka', 'district', 'state', 'qualification_level',
               'phd_registration_date',
               'concerned_university', 'topic_of_phd', 'name_of_guide', 'name_of_college', 'stream', 'board_university',
               'admission_year', 'passing_year', 'percentage', 'family_annual_income', 'income_certificate_number',
               'issuing_authority', 'domicile', 'domicile_certificate', 'domicile_number', 'caste_certf',
               'issuing_district',
               'caste_issuing_authority', 'salaried', 'disability', 'father_name', 'mother_name', 'work_in_government',
               'bank_name', 'account_number', 'ifsc_code', 'account_holder_name'])

    # Add data to the worksheet
    for row in data:
        ws.append(row)

    # Save the workbook in memory as bytes
    data = BytesIO()
    wb.save(data)
    data.seek(0)

    # Create a response object and attach the workbook as a file
    response = make_response(data.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=Total_Applications_2023.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    return response


@app.route('/total_accepted_report', methods=['GET', 'POST'])
def total_accepted_report():
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor(dictionary=True)
    cursor.execute("SELECT * FROM application_page WHERE final_approval='accepted'")
    result = cursor.fetchall()
    cnx.commit()
    cursor.close()
    cnx.close()
    # print(result)
    return render_template('total_accepted_report.html', result=result)


# ----------------------  EXPORT TO EXCEL Accepted Applications 2023 -------------------------------------------
@app.route('/export_accepted_applications_23', methods=['GET', 'POST'])
def export_accepted_applications_23():
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor()
    cursor.execute("SELECT applicant_id, first_name, middle_name, last_name, mobile_number, email, date_of_birth,"
                   "gender, age, caste, your_caste, marital_status, add_1, add_2, pincode, village, taluka, district,"
                   "state, city, qualification_level, phd_registration_date, concerned_university, topic_of_phd, "
                   "name_of_guide, name_of_college, stream, board_university, admission_year, passing_year, percentage,"
                   "family_annual_income, income_certificate_number, issuing_authority, domicile, domicile_certificate,"
                   "domicile_number, caste_certf, issuing_district, caste_issuing_authority, salaried, disability,"
                   "father_name, mother_name, work_in_government, bank_name, account_number, ifsc_code,"
                   "account_holder_name FROM application_page WHERE phd_registration_year = '2023' AND final_approval = 'accepted' ")
    data = cursor.fetchall()
    # Create a workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Add headers to the worksheet
    ws.append(['applicant_id', 'first_name', 'middle_name', 'last_name', 'mobile_number', 'email',
               'date_of_birth', 'gender', 'age', 'caste', 'your_caste', 'marital_status', 'dependents', 'add_1',
               'add_2', 'pincode', 'village', 'taluka', 'district', 'state', 'qualification_level',
               'phd_registration_date',
               'concerned_university', 'topic_of_phd', 'name_of_guide', 'name_of_college', 'stream', 'board_university',
               'admission_year', 'passing_year', 'percentage', 'family_annual_income', 'income_certificate_number',
               'issuing_authority', 'domicile', 'domicile_certificate', 'domicile_number', 'caste_certf',
               'issuing_district',
               'caste_issuing_authority', 'salaried', 'disability', 'father_name', 'mother_name', 'work_in_government',
               'bank_name', 'account_number', 'ifsc_code', 'account_holder_name'])

    # Add data to the worksheet
    for row in data:
        ws.append(row)

    # Save the workbook in memory as bytes
    data = BytesIO()
    wb.save(data)
    data.seek(0)

    # Create a response object and attach the workbook as a file
    response = make_response(data.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=Accepted_Applications_2023.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    return response


@app.route('/total_rejected_report', methods=['GET', 'POST'])
def total_rejected_report():
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor(dictionary=True)
    cursor.execute("SELECT * FROM application_page WHERE final_approval='rejected'")
    result = cursor.fetchall()
    cnx.commit()
    cursor.close()
    cnx.close()
    # print(result)
    return render_template('total_rejected_report.html', result=result)


# ----------------------  EXPORT TO EXCEL Rejected Applications 2023 -------------------------------------------
@app.route('/export_rejected_applications_23', methods=['GET', 'POST'])
def export_rejected_applications_23():
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor()
    cursor.execute("SELECT applicant_id, first_name, middle_name, last_name, mobile_number, email, date_of_birth,"
                   "gender, age, caste, your_caste, marital_status, add_1, add_2, pincode, village, taluka, district,"
                   "state, city, qualification_level, phd_registration_date, concerned_university, topic_of_phd, "
                   "name_of_guide, name_of_college, stream, board_university, admission_year, passing_year, percentage,"
                   "family_annual_income, income_certificate_number, issuing_authority, domicile, domicile_certificate,"
                   "domicile_number, caste_certf, issuing_district, caste_issuing_authority, salaried, disability,"
                   "father_name, mother_name, work_in_government, bank_name, account_number, ifsc_code,"
                   "account_holder_name FROM application_page WHERE phd_registration_year = '2023' AND final_approval = 'rejected' ")
    data = cursor.fetchall()
    # Create a workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Add headers to the worksheet
    ws.append(['applicant_id', 'first_name', 'middle_name', 'last_name', 'mobile_number', 'email',
               'date_of_birth', 'gender', 'age', 'caste', 'your_caste', 'marital_status', 'dependents', 'add_1',
               'add_2', 'pincode', 'village', 'taluka', 'district', 'state', 'qualification_level',
               'phd_registration_date',
               'concerned_university', 'topic_of_phd', 'name_of_guide', 'name_of_college', 'stream', 'board_university',
               'admission_year', 'passing_year', 'percentage', 'family_annual_income', 'income_certificate_number',
               'issuing_authority', 'domicile', 'domicile_certificate', 'domicile_number', 'caste_certf',
               'issuing_district',
               'caste_issuing_authority', 'salaried', 'disability', 'father_name', 'mother_name', 'work_in_government',
               'bank_name', 'account_number', 'ifsc_code', 'account_holder_name'])

    # Add data to the worksheet
    for row in data:
        ws.append(row)

    # Save the workbook in memory as bytes
    data = BytesIO()
    wb.save(data)
    data.seek(0)

    # Create a response object and attach the workbook as a file
    response = make_response(data.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=Rejected_Applications_2023.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    return response


# ----------------------------------------------------------------------------
# ------------------- END APPLICATIONS 2023 REPORTS ------------------------------
# ----------------------------------------------------------------------------

# ----------------------------------------------------------------------------
# ------------------- BEGIN APPLICATIONS 2022 REPORTS ------------------------------
# ----------------------------------------------------------------------------
@app.route('/total_application_report_22', methods=['GET', 'POST'])
def total_application_report_22():
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor(dictionary=True)
    cursor.execute("SELECT * FROM old_users WHERE phd_registration_year='2022'")
    result = cursor.fetchall()
    cnx.commit()
    cursor.close()
    cnx.close()
    # print(result)
    return render_template('total_application_report_22.html', result=result)


# ----------------------  EXPORT TO EXCEL Total applications 2022 -------------------------------------------
@app.route('/export_total_applications_22', methods=['GET', 'POST'])
def export_total_applications_22():
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor()
    # cursor.execute("SELECT applicant_id,first_name,middle_name,last_name,mobile_number,email,date_of_birth,gender,age,caste,your_caste,marital_status,dependents,add_1,add_2,pincode,village,taluka,district,state,qualification_level,phd_registration_date,concerned_university,topic_of_phd,name_of_guide,name_of_college,stream,board_university,admission_year,passing_year,percentage,family_annual_income,income_certificate_number,issuing_authority,domicile,domicile_certificate,relation,domicile_number,caste,caste_category,caste_certificate_number,issuing_district,caste_applicant_name,caste_issuing_authority,salaried,disability,father_alive,father_name,mother_alive,mother_name,work_in_government,bank_name,account_number,ifsc_code,account_holder_name FROM application_page")
    cursor.execute("SELECT * from application_page WHERE phd_registration_year = '2022' ")
    data = cursor.fetchall()
    # Create a workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Add headers to the worksheet
    # ws.append(['applicant_id','email','first_name','last_name','application_date'])

    ws.append(['applicant_id', 'first_name', 'middle_name', 'last_name', 'mobile_number', 'email',
               'date_of_birth', 'gender', 'age', 'caste', 'your_caste', 'marital_status', 'dependents', 'add_1',
               'add_2', 'pincode', 'village', 'taluka', 'district', 'state', 'qualification_level',
               'phd_registration_date',
               'concerned_university', 'topic_of_phd', 'name_of_guide', 'name_of_college', 'stream', 'board_university',
               'admission_year', 'passing_year', 'percentage', 'family_annual_income', 'income_certificate_number',
               'issuing_authority', 'domicile', 'domicile_certificate', 'relation', 'domicile_number', 'caste',
               'caste_category', 'caste_certificate_number', 'issuing_district', 'caste_applicant_name',
               'caste_issuing_authority', 'salaried', 'disability', 'father_alive', 'father_name', 'mother_alive',
               'mother_name', 'work_in_government', 'bank_name', 'account_number', 'ifsc_code', 'account_holder_name'])

    # Add data to the worksheet
    for row in data:
        ws.append(row)

    # Save the workbook in memory as bytes
    data = BytesIO()
    wb.save(data)
    data.seek(0)

    # Create a response object and attach the workbook as a file
    response = make_response(data.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=Total_Applications_2022.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    return response


@app.route('/total_accepted_report_22', methods=['GET', 'POST'])
def total_accepted_report_22():
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor(dictionary=True)
    cursor.execute("SELECT * FROM old_users WHERE phd_registration_year='2022'")
    result = cursor.fetchall()
    cnx.commit()
    cursor.close()
    cnx.close()
    # print(result)
    return render_template('total_accepted_report_22.html', result=result)


# ----------------------  EXPORT TO EXCEL Accepted Applications 2022 -------------------------------------------
@app.route('/export_accepted_applications_22', methods=['GET', 'POST'])
def export_accepted_applications_22():
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor()
    cursor.execute("SELECT * from application_page WHERE final_approval = 'accepted' ")
    data = cursor.fetchall()
    # Create a workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Add headers to the worksheet
    ws.append(['applicant_id', 'first_name', 'middle_name', 'last_name', 'mobile_number', 'email',
               'date_of_birth', 'gender', 'age', 'caste', 'your_caste', 'marital_status', 'dependents', 'add_1',
               'add_2', 'pincode', 'village', 'taluka', 'district', 'state', 'qualification_level',
               'phd_registration_date',
               'concerned_university', 'topic_of_phd', 'name_of_guide', 'name_of_college', 'stream', 'board_university',
               'admission_year', 'passing_year', 'percentage', 'family_annual_income', 'income_certificate_number',
               'issuing_authority', 'domicile', 'domicile_certificate', 'relation', 'domicile_number', 'caste',
               'caste_category', 'caste_certificate_number', 'issuing_district', 'caste_applicant_name',
               'caste_issuing_authority', 'salaried', 'disability', 'father_alive', 'father_name', 'mother_alive',
               'mother_name', 'work_in_government', 'bank_name', 'account_number', 'ifsc_code', 'account_holder_name'])

    # Add data to the worksheet
    for row in data:
        ws.append(row)

    # Save the workbook in memory as bytes
    data = BytesIO()
    wb.save(data)
    data.seek(0)

    # Create a response object and attach the workbook as a file
    response = make_response(data.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=Accepted_Applications_2023.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    return response


@app.route('/total_rejected_report_22', methods=['GET', 'POST'])
def total_rejected_report_22():
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor(dictionary=True)
    cursor.execute("SELECT rejected FROM old_users WHERE phd_registration_year='2022'")
    result = cursor.fetchall()
    cnx.commit()
    cursor.close()
    cnx.close()
    # print(result)
    return render_template('total_rejected_report_22.html', result=result)


# ----------------------  EXPORT TO EXCEL Rejected Applications 2022 -------------------------------------------
@app.route('/export_rejected_applications_22', methods=['GET', 'POST'])
def export_rejected_applications_22():
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor()
    cursor.execute(
        "SELECT * from application_page WHERE phd_registration_year = '2022' AND final_approval = 'rejected' ")
    data = cursor.fetchall()
    # Create a workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Add headers to the worksheet
    ws.append(['applicant_id', 'first_name', 'middle_name', 'last_name', 'mobile_number', 'email',
               'date_of_birth', 'gender', 'age', 'caste', 'your_caste', 'marital_status', 'dependents', 'add_1',
               'add_2', 'pincode', 'village', 'taluka', 'district', 'state', 'qualification_level',
               'phd_registration_date',
               'concerned_university', 'topic_of_phd', 'name_of_guide', 'name_of_college', 'stream', 'board_university',
               'admission_year', 'passing_year', 'percentage', 'family_annual_income', 'income_certificate_number',
               'issuing_authority', 'domicile', 'domicile_certificate', 'relation', 'domicile_number', 'caste',
               'caste_category', 'caste_certificate_number', 'issuing_district', 'caste_applicant_name',
               'caste_issuing_authority', 'salaried', 'disability', 'father_alive', 'father_name', 'mother_alive',
               'mother_name', 'work_in_government', 'bank_name', 'account_number', 'ifsc_code', 'account_holder_name'])

    # Add data to the worksheet
    for row in data:
        ws.append(row)

    # Save the workbook in memory as bytes
    data = BytesIO()
    wb.save(data)
    data.seek(0)

    # Create a response object and attach the workbook as a file
    response = make_response(data.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=Rejected_Applications_2022.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    return response


# ----------------------------------------------------------------------------
# ------------------- END APPLICATIONS 2022 REPORTS ------------------------------
# ----------------------------------------------------------------------------

# ----------------------------------------------------------------------------
# ------------------- BEGIN APPLICATIONS 2021 REPORTS ------------------------------
# ----------------------------------------------------------------------------
@app.route('/total_application_report_21', methods=['GET', 'POST'])
def total_application_report_21():
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor(dictionary=True)
    cursor.execute("SELECT * FROM old_users WHERE phd_registration_year='2021'")
    result = cursor.fetchall()
    cnx.commit()
    cursor.close()
    cnx.close()
    # print(result)
    return render_template('total_application_report_21.html', result=result)


# ----------------------  EXPORT TO EXCEL Total applications 2021 ----------------
@app.route('/export_total_applications_21', methods=['GET', 'POST'])
def export_total_applications_21():
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor()
    # cursor.execute("SELECT applicant_id,first_name,middle_name,last_name,mobile_number,email,date_of_birth,gender,age,caste,your_caste,marital_status,dependents,add_1,add_2,pincode,village,taluka,district,state,qualification_level,phd_registration_date,concerned_university,topic_of_phd,name_of_guide,name_of_college,stream,board_university,admission_year,passing_year,percentage,family_annual_income,income_certificate_number,issuing_authority,domicile,domicile_certificate,relation,domicile_number,caste,caste_category,caste_certificate_number,issuing_district,caste_applicant_name,caste_issuing_authority,salaried,disability,father_alive,father_name,mother_alive,mother_name,work_in_government,bank_name,account_number,ifsc_code,account_holder_name FROM application_page")
    cursor.execute("SELECT * from application_page WHERE phd_registration_year = '2021' ")
    data = cursor.fetchall()
    # Create a workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Add headers to the worksheet
    # ws.append(['applicant_id','email','first_name','last_name','application_date'])

    ws.append(['applicant_id', 'first_name', 'middle_name', 'last_name', 'mobile_number', 'email',
               'date_of_birth', 'gender', 'age', 'caste', 'your_caste', 'marital_status', 'dependents', 'add_1',
               'add_2', 'pincode', 'village', 'taluka', 'district', 'state', 'qualification_level',
               'phd_registration_date',
               'concerned_university', 'topic_of_phd', 'name_of_guide', 'name_of_college', 'stream', 'board_university',
               'admission_year', 'passing_year', 'percentage', 'family_annual_income', 'income_certificate_number',
               'issuing_authority', 'domicile', 'domicile_certificate', 'relation', 'domicile_number', 'caste',
               'caste_category', 'caste_certificate_number', 'issuing_district', 'caste_applicant_name',
               'caste_issuing_authority', 'salaried', 'disability', 'father_alive', 'father_name', 'mother_alive',
               'mother_name', 'work_in_government', 'bank_name', 'account_number', 'ifsc_code', 'account_holder_name'])

    # Add data to the worksheet
    for row in data:
        ws.append(row)

    # Save the workbook in memory as bytes
    data = BytesIO()
    wb.save(data)
    data.seek(0)

    # Create a response object and attach the workbook as a file
    response = make_response(data.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=Total_Applications_2021.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    return response


@app.route('/total_accepted_report_21', methods=['GET', 'POST'])
def total_accepted_report_21():
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor(dictionary=True)
    cursor.execute("SELECT * FROM old_users WHERE phd_registration_year='2021'")
    result = cursor.fetchall()
    cnx.commit()
    cursor.close()
    cnx.close()
    # print(result)
    return render_template('total_accepted_report_21.html', result=result)


# ----------------------  EXPORT TO EXCEL Accepted Applications 2021 -------------------------------------------
@app.route('/export_accepted_applications_21', methods=['GET', 'POST'])
def export_accepted_applications_21():
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor()
    cursor.execute(
        "SELECT * from application_page WHERE phd_registration_year = '2021' AND final_approval = 'accepted' ")
    data = cursor.fetchall()
    # Create a workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Add headers to the worksheet
    ws.append(['applicant_id', 'first_name', 'middle_name', 'last_name', 'mobile_number', 'email',
               'date_of_birth', 'gender', 'age', 'caste', 'your_caste', 'marital_status', 'dependents', 'add_1',
               'add_2', 'pincode', 'village', 'taluka', 'district', 'state', 'qualification_level',
               'phd_registration_date',
               'concerned_university', 'topic_of_phd', 'name_of_guide', 'name_of_college', 'stream', 'board_university',
               'admission_year', 'passing_year', 'percentage', 'family_annual_income', 'income_certificate_number',
               'issuing_authority', 'domicile', 'domicile_certificate', 'relation', 'domicile_number', 'caste',
               'caste_category', 'caste_certificate_number', 'issuing_district', 'caste_applicant_name',
               'caste_issuing_authority', 'salaried', 'disability', 'father_alive', 'father_name', 'mother_alive',
               'mother_name', 'work_in_government', 'bank_name', 'account_number', 'ifsc_code', 'account_holder_name'])

    # Add data to the worksheet
    for row in data:
        ws.append(row)

    # Save the workbook in memory as bytes
    data = BytesIO()
    wb.save(data)
    data.seek(0)

    # Create a response object and attach the workbook as a file
    response = make_response(data.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=Accepted_Applications_2021.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    return response


@app.route('/total_rejected_report_21', methods=['GET', 'POST'])
def total_rejected_report_21():
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor(dictionary=True)
    cursor.execute("SELECT rejected FROM old_users WHERE phd_registration_year='2021'")
    result = cursor.fetchall()
    cnx.commit()
    cursor.close()
    cnx.close()
    # print(result)
    return render_template('total_rejected_report_21.html', result=result)


# ----------------------  EXPORT TO EXCEL Rejected Applications 2021 -------------------------------------------
@app.route('/export_rejected_applications_21', methods=['GET', 'POST'])
def export_rejected_applications_21():
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor()
    cursor.execute(
        "SELECT * from application_page WHERE phd_registration_year = '2021' AND final_approval = 'rejected' ")
    data = cursor.fetchall()
    # Create a workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Add headers to the worksheet
    ws.append(['applicant_id', 'first_name', 'middle_name', 'last_name', 'mobile_number', 'email',
               'date_of_birth', 'gender', 'age', 'caste', 'your_caste', 'marital_status', 'dependents', 'add_1',
               'add_2', 'pincode', 'village', 'taluka', 'district', 'state', 'qualification_level',
               'phd_registration_date',
               'concerned_university', 'topic_of_phd', 'name_of_guide', 'name_of_college', 'stream', 'board_university',
               'admission_year', 'passing_year', 'percentage', 'family_annual_income', 'income_certificate_number',
               'issuing_authority', 'domicile', 'domicile_certificate', 'relation', 'domicile_number', 'caste',
               'caste_category', 'caste_certificate_number', 'issuing_district', 'caste_applicant_name',
               'caste_issuing_authority', 'salaried', 'disability', 'father_alive', 'father_name', 'mother_alive',
               'mother_name', 'work_in_government', 'bank_name', 'account_number', 'ifsc_code', 'account_holder_name'])

    # Add data to the worksheet
    for row in data:
        ws.append(row)

    # Save the workbook in memory as bytes
    data = BytesIO()
    wb.save(data)
    data.seek(0)

    # Create a response object and attach the workbook as a file
    response = make_response(data.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=Rejected_Applications_2021.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    return response


# ----------------------------------------------------------------------------
# ------------------- END APPLICATIONS 2021 REPORTS ------------------------------
# ----------------------------------------------------------------------------

@app.route('/priority_people_report', methods=['GET', 'POST'])
def priority_people_report():
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor(dictionary=True)
    cursor.execute("SELECT * FROM application_page WHERE your_caste IN ('katkari', 'kolam', 'madia')")
    result = cursor.fetchall()
    cnx.commit()
    cursor.close()
    cnx.close()
    # print(result)
    return render_template('priority_people_report.html', result=result)


# ---------------------  Export Disability Report Yes list to Excel  -----------------------
@app.route('/export_priority_caste_report', methods=['GET', 'POST'])
def export_priority_caste_report():
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor()
    cursor.execute(
        "SELECT applicant_id, email, first_name, last_name, application_date  FROM application_page WHERE your_caste IN ('katkari', 'kolam', 'madia')")
    data = cursor.fetchall()
    # Create a workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Add headers to the worksheet
    ws.append(['Applicant id', 'Email', 'First Name', 'Last Name', 'Application Date'])

    # Add data to the worksheet
    for row in data:
        ws.append(row)

    # Save the workbook in memory as bytes
    data = BytesIO()
    wb.save(data)
    data.seek(0)

    # Create a response object and attach the workbook as a file
    response = make_response(data.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=Priority_Caste_Report.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    return response


@app.route('/all_news', methods=['GET', 'POST'])
def all_news():
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor(dictionary=True)
    cursor.execute("SELECT * FROM news_and_updates")
    result = cursor.fetchall()
    cnx.commit()
    cursor.close()
    cnx.close()
    # print(result)
    return render_template('all_news_updates.html', result=result)


# ---------------------  Export Rejected Users list in Excel  -----------------------
@app.route('/export_rejected_applications', methods=['GET', 'POST'])
def export_rejected_applications():
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor()
    cursor.execute(
        "SELECT first_name,middle_name,last_name, mobile_number, email FROM application_page WHERE final_approval = 'rejected'")
    data = cursor.fetchall()
    # Create a workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Add headers to the worksheet
    ws.append(['first_name', 'middle_name', 'last_name', 'mobile_number', 'email', 'day', 'month', 'year', 'age'])

    # Add data to the worksheet
    for row in data:
        ws.append(row)

    # Save the workbook in memory as bytes
    data = BytesIO()
    wb.save(data)
    data.seek(0)

    # Create a response object and attach the workbook as a file
    response = make_response(data.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=Rejected_Users_Data.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    return response


# --------------------------------------------------------------------------------------------------------------------
# --------------------------------------------------------------------------------------------------------------------
#                        ----------------------- APPLICATION FORM -------------------------------
# --------------------------------------------------------------------------------------------------------------------
# --------------------------------------------------------------------------------------------------------------------
# Define form classes for each section
# class Section1Form(FlaskForm):
#     applicant_photo = FileField('applicant_photo', validators=[FileRequired()])
#     adhaar_number = IntegerField('adhaar_number', validators=[InputRequired()])
#     first_name = StringField('first_name', validators=[InputRequired()])
#     middle_name = StringField('middle_name', validators=[InputRequired()])
#     last_name = StringField('last_name', validators=[InputRequired()])
#     mobile_number = IntegerField('mobile_number', validators=[InputRequired()])
#     email = StringField('email', validators=[InputRequired()])
#     date_of_birth = StringField('date_of_birth', validators=[InputRequired()])
#     gender = StringField('gender', validators=[InputRequired()])
#     age = StringField('age', validators=[InputRequired()])
#     caste = StringField('caste', validators=[InputRequired()])
#     your_caste = StringField('your_caste', validators=[InputRequired()])
#     marital_status = StringField('marital_status', validators=[InputRequired()])
#     add_1 = StringField('add_1', validators=[InputRequired()])
#     add_2 = StringField('add_2', validators=[InputRequired()])
#     pincode = IntegerField('pincode', validators=[InputRequired()])
#     village = StringField('village', validators=[InputRequired()])
#     taluka = StringField('taluka', validators=[InputRequired()])
#     district = StringField('district', validators=[InputRequired()])
#     city = StringField('city', validators=[InputRequired()])
#     state = StringField('state', validators=[InputRequired()])
#     # Add more fields for this section
#
#
# class Section2Form(FlaskForm):
#     ssc_passing_year = IntegerField('ssc_passing_year', validators=[InputRequired()])
#     ssc_percentage = IntegerField('ssc_percentage', validators=[InputRequired()])
#     ssc_school_name = StringField('ssc_school_name', validators=[InputRequired()])
#     hsc_passing_year = IntegerField('hsc_passing_year', validators=[InputRequired()])
#     hsc_percentage = IntegerField('hsc_percentage', validators=[InputRequired()])
#     hsc_school_name = StringField('hsc_school_name', validators=[InputRequired()])
#     graduation_passing_year = IntegerField('graduation_passing_year', validators=[InputRequired()])
#     graduation_percentage = IntegerField('graduation_percentage', validators=[InputRequired()])
#     graduation_school_name = StringField('graduation_school_name', validators=[InputRequired()])
#     phd_passing_year = IntegerField('phd_passing_year', validators=[InputRequired()])
#     phd_percentage = IntegerField('phd_percentage', validators=[InputRequired()])
#     phd_school_name = StringField('phd_school_name', validators=[InputRequired()])
#     have_you_qualified = StringField('have_you_qualified', validators=[InputRequired()])
#     concerned_university = StringField('concerned_university', validators=[InputRequired()])
#     topic_of_phd = StringField('topic_of_phd', validators=[InputRequired()])
#     name_of_guide = StringField('name_of_guide', validators=[InputRequired()])
#     board_university = StringField('board_university', validators=[InputRequired()])
#     stream = StringField('stream', validators=[InputRequired()])
#     faculty = StringField('faculty', validators=[InputRequired()])
#     admission_year = DateField('admission_year', validators=[InputRequired()])
#     passing_year = DateField('passing_year', validators=[InputRequired()])
#     percentage = IntegerField('percentage', validators=[InputRequired()])
#     family_annual_income =IntegerField('family_annual_income', validators=[InputRequired()])
#     income_certificate_number = IntegerField('income_certificate_number', validators=[InputRequired()])
#     issuing_authority = StringField('issuing_authority', validators=[InputRequired()])
#     # Add more fields for this section
#
#
# class Section3Form(FlaskForm):
#     domicile = StringField('Field 4', validators=[InputRequired()])
#     domicile_certificate = StringField('Field 4', validators=[InputRequired()])
#     domicile_number = StringField('Field 4', validators=[InputRequired()])
#     caste_certf = StringField('Field 4', validators=[InputRequired()])
#     issuing_district = StringField('Field 4', validators=[InputRequired()])
#     caste_issuing_authority = StringField('Field 4', validators=[InputRequired()])
#     salaried = StringField('Field 4', validators=[InputRequired()])
#     disability = StringField('Field 4', validators=[InputRequired()])
#     type_of_disability = StringField('Field 4', validators=[InputRequired()])
#
#
# class Section4Form(FlaskForm):
#     father_name = StringField('father_name', validators=[InputRequired()])
#     mother_name = StringField('mother_name', validators=[InputRequired()])
#     work_in_government = StringField('work_in_government', validators=[InputRequired()])
#     bank_name = StringField('bank_name', validators=[InputRequired()])
#     account_number = StringField('account_number', validators=[InputRequired()])
#     ifsc_code = StringField('ifsc_code', validators=[InputRequired()])
#     account_holder_name = StringField('account_holder_name', validators=[InputRequired()])
#     # Add more fields for this section
#
#
# class Section5Form(FlaskForm):
#     documentfile1 = FileField('documentfile1', validators=[FileRequired()])
#     documentfile2 = FileField('documentfile1', validators=[FileRequired()])
#     documentfile3 = FileField('documentfile1', validators=[FileRequired()])
#     documentfile4 = FileField('documentfile1', validators=[FileRequired()])
#     documentfile5 = FileField('documentfile1', validators=[FileRequired()])
#     documentfile6 = FileField('documentfile1', validators=[FileRequired()])
#     documentfile7 = FileField('documentfile1', validators=[FileRequired()])
#     documentfile8 = FileField('documentfile1', validators=[FileRequired()])
#     documentfile9 = FileField('documentfile1', validators=[FileRequired()])
#     # Add more fields for this section
#
#
# @app.route('/section/<int:section_number>', methods=['GET', 'POST'])
# def section(section_number):
#     if 'data' not in session:
#         session['data'] = {}
#
#     if request.method == 'POST':
#         # Simulate saving form data to the session (you should replace this with database operations)
#         form_data = request.form.to_dict()
#         session['data'][section_number] = form_data
#
#         # For simplicity, after saving, redirect to the next section (or preview if it's the last section)
#         if section_number < 5:
#             return redirect(url_for('section', section_number=section_number + 1))
#         else:
#             return redirect(url_for('preview'))
#
#     # Render the HTML form for the section
#     return render_template('section.html', section_number=section_number, data=session['data'].get(section_number, {}))

# @app.route('/preview', methods=['GET', 'POST'])
# def preview():
#     if request.method == 'POST':
#         # Process the final submission and insert data into the database
#         for section_number, section_data in session['data'].items():
#             # Insert data into the database using MySQL queries
#
#         # For simplicity, create a PDF and redirect to the submit route
#         create_pdf(session['data'])
#         return redirect(url_for('submit'))
#
#     # Render the preview page with the submitted data
#     return render_template('preview.html', data=session['data'])


@app.route('/section1', methods=['GET', 'POST'])
def section1():
    email = session['email']
    print('I am in section 1:' + email)
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor(dictionary=True)

    # Check if a record already exists for this user
    cursor.execute("SELECT applicant_photo, adhaar_number, first_name, "
                   "middle_name, last_name, mobile_number, email, date_of_birth, gender, age, caste, your_caste, "
                   "marital_status, add_1, add_2, pincode, village, taluka, district, state, city"
                   " FROM application_page WHERE email = %s", (email,))
    record = cursor.fetchone()
    print(record)

    # Initialize an empty dictionary if no record is found
    if record is None:
        record = {}

    if request.method == 'POST':
        #
        adhaar_number = request.form['adhaar_number']
        first_name = request.form['first_name']
        middle_name = request.form['middle_name']
        last_name = request.form['last_name']
        mobile_number = request.form['mobile_number']
        email = session['email']
        date_of_birth = request.form['date_of_birth']
        gender = request.form['gender']
        age = request.form['age']
        caste = request.form['caste']
        your_caste = request.form['your_caste']
        marital_status = request.form['marital_status']
        add_1 = request.form['add_1']
        add_2 = request.form['add_2']
        pincode = request.form['pincode']
        village = request.form['village']
        taluka = request.form['taluka']
        district = request.form['district']
        state = request.form['state']
        city = request.form['city']

        # Access other fields in a similar manner

        # Handle file upload (applicant's photo)
        photo = request.files['applicant_photo']
        photo_path = save_file(photo)

        if not record:
            # Save the form data to the database
            print('Inserting new record for:' + email)
            cursor.execute("INSERT INTO application_page (applicant_photo, adhaar_number, first_name, "
                           "middle_name, last_name, mobile_number, email, date_of_birth, gender, age, caste, your_caste, "
                           "marital_status, add_1, add_2, pincode, village, taluka, district, state, city) "
                           "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)",
                           (photo_path, adhaar_number, first_name, middle_name, last_name, mobile_number,
                            email, date_of_birth, gender, age, caste, your_caste, marital_status, add_1, add_2,
                            pincode, village, taluka, district, state, city))
            cnx.commit()
            return redirect(url_for('section2'))
    return render_template('AForm_section1.html', record=record)


#
#
@app.route('/section2', methods=['GET', 'POST'])
def section2():
    email = session['email']
    print('I am in section 2:' + email)
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor(dictionary=True)

    # Check if a record already exists for this user
    cursor.execute("SELECT ssc_passing_year, ssc_percentage, ssc_school_name, "
                   "hsc_passing_year, hsc_percentage, hsc_school_name, graduation_passing_year, "
                   "graduation_percentage, graduation_school_name, phd_passing_year, phd_percentage, "
                   "phd_school_name, have_you_qualified, phd_registration_date, phd_registration_year, "
                   "concerned_university, topic_of_phd, name_of_guide, board_university, name_of_college, "
                   "stream, faculty, admission_year, passing_year, percentage, family_annual_income,"
                   "income_certificate_number, issuing_authority FROM application_page WHERE email = %s", (email,))
    existing_record = cursor.fetchone()
    print(existing_record)

    # Initialize an empty dictionary if no record is found
    if existing_record is None:
        existing_record = {}

    view_mode = 'view'
    # Determine whether the user is in "view" or "edit" mode
    if not any(existing_record.values()):
        view_mode = 'edit'
    else:
        view_mode = 'view'

    if request.method == 'POST':
        ssc_passing_year = request.form['ssc_passing_year']
        ssc_percentage = request.form['ssc_percentage']
        ssc_school_name = request.form['ssc_school_name']
        hsc_passing_year = request.form['hsc_passing_year']
        hsc_percentage = request.form['hsc_percentage']
        hsc_school_name = request.form['hsc_school_name']
        graduation_passing_year = request.form['graduation_passing_year']
        graduation_percentage = request.form['graduation_percentage']
        graduation_school_name = request.form['graduation_school_name']
        phd_passing_year = request.form['phd_passing_year']
        phd_percentage = request.form['phd_percentage']
        phd_school_name = request.form['phd_school_name']
        have_you_qualified = request.form['have_you_qualified']
        phd_registration_date = request.form['phd_registration_date']
        phd_registration_year = request.form['phd_registration_year']
        concerned_university = request.form['concerned_university']
        topic_of_phd = request.form['topic_of_phd']
        name_of_guide = request.form['name_of_guide']
        board_university = request.form['board_university']
        name_of_college = request.form['name_of_college']
        stream = request.form['stream']
        faculty = request.form['faculty']
        admission_year = request.form['admission_year']
        passing_year = request.form['passing_year']
        percentage = request.form['percentage']
        family_annual_income = request.form['family_annual_income']
        income_certificate_number = request.form['income_certificate_number']
        issuing_authority = request.form['issuing_authority']
        print(request.form)
        print("I am breaking before the if condition")

        if all(value is None for value in existing_record.values()):
            # Save the form data to the database
            print('I am breaking while insertion')
            print('Inserting new record for:' + email)
            if view_mode == 'edit':
                cursor.execute(
                    "UPDATE application_page SET ssc_passing_year = %s, ssc_percentage = %s, ssc_school_name = %s, "
                    "hsc_passing_year = %s, hsc_percentage = %s, hsc_school_name = %s, graduation_passing_year = %s, "
                    "graduation_percentage = %s, graduation_school_name = %s, phd_passing_year = %s, phd_percentage = %s, "
                    "phd_school_name = %s, have_you_qualified = %s, phd_registration_date = %s, phd_registration_year = %s, "
                    "concerned_university = %s, topic_of_phd = %s, name_of_guide = %s, board_university = %s, name_of_college = %s, "
                    "stream = %s, faculty = %s, admission_year = %s, passing_year = %s, percentage = %s, family_annual_income = %s, "
                    "income_certificate_number = %s, issuing_authority = %s WHERE email = %s",
                    (ssc_passing_year, ssc_percentage, ssc_school_name, hsc_passing_year, hsc_percentage,
                     hsc_school_name,
                     graduation_passing_year, graduation_percentage, graduation_school_name, phd_passing_year,
                     phd_percentage,
                     phd_school_name, have_you_qualified, phd_registration_date, phd_registration_year,
                     concerned_university,
                     topic_of_phd, name_of_guide, board_university, name_of_college, stream, faculty, admission_year,
                     passing_year, percentage, family_annual_income, income_certificate_number, issuing_authority,
                     email))
                cnx.commit()
            return redirect(url_for('section3'))
        # Select records again to verify the insertion
        cursor.execute("SELECT ssc_passing_year, ssc_percentage, ssc_school_name, "
                       "hsc_passing_year, hsc_percentage, hsc_school_name, graduation_passing_year, "
                       "graduation_percentage, graduation_school_name, phd_passing_year, phd_percentage, "
                       "phd_school_name, have_you_qualified, phd_registration_date, phd_registration_year, "
                       "concerned_university, topic_of_phd, name_of_guide, board_university, name_of_college, "
                       "stream, faculty, admission_year, passing_year, percentage, family_annual_income,"
                       "income_certificate_number, issuing_authority FROM application_page WHERE email = %s", (email,))
        existing_record = cursor.fetchone()
        print(existing_record)
        # Initialize view_mode with a default value

    return render_template('AForm_section2.html', existing_record=existing_record, view_mode=view_mode)


@app.route('/section3', methods=['GET', 'POST'])
def section3():
    email = session['email']
    print('I am in section 3:' + email)
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor(dictionary=True)

    # Check if a record already exists for this user
    cursor.execute("SELECT domicile, domicile_certificate, domicile_number, caste_certf, issuing_district, "
                   "caste_issuing_authority, salaried, disability, type_of_disability "
                   "FROM application_page WHERE email = %s", (email,))
    existing_record = cursor.fetchone()
    # print(existing_record)
    # Check if existing_record is not None
    if existing_record is not None:
        # Determine whether the user is in "view" or "edit" mode
        if not any(existing_record.values()):
            view_mode = 'edit'
        else:
            view_mode = 'view'
    else:
        # Handle the case when existing_record is None
        view_mode = 'edit'  # or 'view', depending on your logic for this case

    if request.method == 'POST':
        # Save form data in the session for this section
        if request.method == 'POST':
            domicile = request.form['domicile']
            domicile_certificate = request.form['domicile_certificate']
            domicile_number = request.form['domicile_number']
            caste_certf = request.form['caste_certf']
            issuing_district = request.form['issuing_district']
            caste_issuing_authority = request.form['caste_issuing_authority']
            salaried = request.form['salaried']
            disability = request.form['disability']
            type_of_disability = request.form['type_of_disability']
            print(request.form)
            print("I am breaking before the if condition")
            if all(value is None for value in existing_record.values()):
                # Save the form data to the database
                print('I am breaking while insertion')
                print('Inserting new record for:' + email)
                if view_mode == 'edit':
                    cursor.execute(
                        "UPDATE application_page SET domicile = %s, domicile_certificate = %s, domicile_number = %s, "
                        "caste_certf = %s, issuing_district = %s, caste_issuing_authority = %s, salaried = %s, "
                        "disability = %s, type_of_disability = %s WHERE email = %s",
                        (domicile, domicile_certificate, domicile_number, caste_certf, issuing_district,
                         caste_issuing_authority, salaried, disability, type_of_disability, email))
                    cnx.commit()
            return redirect(url_for('section4'))
        # Select records again to verify the insertion
        cursor.execute("SELECT domicile, domicile_certificate, domicile_number, "
                       "caste_certf, issuing_district, caste_issuing_authority, salaried, disability,"
                       " type_of_disability FROM application_page WHERE email = %s", (email,))
        existing_record = cursor.fetchone()
        print(existing_record)
    return render_template('AForm_section3.html', existing_record=existing_record, view_mode=view_mode)


@app.route('/section4', methods=['GET', 'POST'])
def section4():
    email = session['email']
    print('I am in section 4:' + email)
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor(dictionary=True)

    # Check if a record already exists for this user
    cursor.execute("SELECT father_name, mother_name, work_in_government, bank_name, account_number, "
                   "ifsc_code, account_holder_name FROM application_page WHERE email = %s", (email,))
    existing_record = cursor.fetchone()
    print(existing_record)

    if existing_record is not None:
        # Determine whether the user is in "view" or "edit" mode
        if not any(existing_record.values()):
            view_mode = 'edit'
        else:
            view_mode = 'view'
    else:
        # Handle the case when existing_record is None
        view_mode = 'edit'  # or 'view', depending on your logic for this case

    if request.method == 'POST':
        # Save form data in the session for this section
        father_name = request.form['father_name']
        mother_name = request.form['mother_name']
        work_in_government = request.form['work_in_government']
        bank_name = request.form['bank_name']
        account_number = request.form['account_number']
        ifsc_code = request.form['ifsc_code']
        account_holder_name = request.form['account_holder_name']
        # Process the fields as needed
        print(request.form)
        print("I am breaking before the if condition")
        if all(value is None for value in existing_record.values()):
            # Save the form data to the database
            print('I am breaking while insertion')
            print('Inserting new record for:' + email)
            if view_mode == 'edit':
                cursor.execute(
                    "UPDATE application_page SET father_name = %s, mother_name = %s, work_in_government = %s, "
                    "bank_name = %s, account_number = %s, ifsc_code = %s, account_holder_name = %s WHERE email = %s",
                    (father_name, mother_name, work_in_government, bank_name, account_number, ifsc_code,
                     account_holder_name,
                     email))
                cnx.commit()
            return redirect(url_for('section5'))
        # Select records again to verify the insertion
        cursor.execute("SELECT father_name, mother_name, work_in_government, bank_name, account_number, "
                       "ifsc_code, account_holder_name FROM application_page WHERE email = %s", (email,))
        existing_record = cursor.fetchone()
        print(existing_record)
    return render_template('AForm_section4.html', existing_record=existing_record, view_mode=view_mode)


@app.route('/section5', methods=['GET', 'POST'])
def section5():
    email = session['email']
    print('I am in section 5:' + email)
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor(dictionary=True)

    # Check if a record already exists for this user
    cursor.execute("SELECT documentfile1, documentfile2, documentfile3, documentfile4, documentfile5, documentfile6,"
                   "documentfile7, documentfile8, documentfile9 FROM application_page WHERE email = %s", (email,))
    existing_record = cursor.fetchone()
    print(existing_record)
    if request.method == 'POST':
        # Process the form data and files
        document1 = save_file(request.files['documentfile1'])
        document2 = save_file(request.files['documentfile2'])
        document3 = save_file(request.files['documentfile3'])
        document4 = save_file(request.files['documentfile4'])
        document5 = save_file(request.files['documentfile5'])
        document6 = save_file(request.files['documentfile6'])
        document7 = save_file(request.files['documentfile7'])
        document8 = save_file(request.files['documentfile8'])
        document9 = save_file(request.files['documentfile9'])
    else:
        document1 = ''
        document2 = ''
        document3 = ''
        document4 = ''
        document5 = ''
        document6 = ''
        document7 = ''
        document8 = ''
        document9 = ''
    # Check if any of the documents have been uploaded
    if any(doc for doc in
           [document1, document2, document3, document4, document5, document6, document7, document8, document9]):
        # If any document is uploaded, update or insert into the database
        if all(value is None for value in existing_record.values()):
            # Insert new records
            print('Inserting new record for:' + email)
            cursor.execute(
                "INSERT INTO application_page (email, documentfile1, documentfile2, documentfile3, documentfile4, documentfile5, "
                "documentfile6, documentfile7, documentfile8, documentfile9) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)",
                (email, document1, document2, document3, document4, document5, document6, document7, document8,
                 document9))
        else:
            # Update existing records
            print('Updating records for:' + email)
            cursor.execute(
                "UPDATE application_page SET documentfile1 = %s, documentfile2 = %s, documentfile3 = %s, documentfile4 = %s, "
                "documentfile5 = %s, documentfile6 = %s, documentfile7 = %s, documentfile8 = %s, documentfile9 = %s WHERE email = %s",
                (document1, document2, document3, document4, document5, document6, document7, document8, document9,
                 email))
        cnx.commit()
        return redirect(url_for('preview'))
    else:
        # Handle case where no files were uploaded
        # You can add appropriate logic here (e.g., show an error message)
        print('No files were uploaded')
    return render_template('AForm_section5.html', existing_record=existing_record)


@app.route('/preview', methods=['GET', 'POST'])
def preview():
    email = session['email']
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor(dictionary=True)
    # Check if the user's email is in old_users for 2021 or 2022
    cursor.execute("SELECT * FROM old_users WHERE email = %s AND phd_registration_year IN ('2021', '2022')", (email,))
    old_user_data = cursor.fetchone()

    if old_user_data:
        # User exists in old_users for 2021 or 2022, use this data
        return render_template('preview.html', records=old_user_data, editable=True)
    else:
        # User does not exist in old_users for 2021 or 2022, fetch from application_page
        cursor.execute("SELECT * FROM application_page WHERE email = %s", (email,))
        application_page_data = cursor.fetchone()
        return render_template('preview.html', records=application_page_data, editable=True)

    cursor.close()
    cnx.close()
    return render_template('preview.html', records=records, editable=True)


@app.route('/submit_form', methods=['GET', 'POST'])
def submit_form():
    email = session['email']
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor(dictionary=True)

    if request.method == 'POST':
        # Generate the applicant_id
        cursor.execute("SELECT phd_registration_year, id FROM application_page WHERE email = %s", (email,))
        result = cursor.fetchone()

        if result:
            phd_registration_year = result['phd_registration_year']
            id = result['id']
            applicant_id = f"TRTI/{phd_registration_year}/{id}"

            # Update the 'applicant_id', 'form_filled', and 'application_date' fields in the database
            cursor.execute(
                "UPDATE application_page SET applicant_id = %s, form_filled = 1, application_date = %s WHERE email = %s",
                (applicant_id, datetime.now(), email))
            cnx.commit()

    cursor.execute("SELECT * FROM application_page WHERE email = %s", (email,))
    records = cursor.fetchone()
    print(records)
    return render_template('submit_form.html', records=records)


# Function to generate PDF content based on data
# Function to generate a styled PDF
def generate_pdf_with_styling(data, filename):
    class PDF(FPDF):
        header_added = False  # To track whether the header is added to the first page

        def header(self):
            if not self.header_added:
                # Add a header
                self.set_font("Arial", "B", 12)
                self.cell(0, 10, "Fellowship ", align="C",
                          ln=True)  # Add space by changing the second parameter (e.g., 20)
                # Insert an image (symbol) at the center of the header
                self.image('static/Images/trti.jpeg', 10, 10, 20)  # Replace with the path to your symbol image
                # Insert an image (symbol) at the right of the header
                self.image('static/Images/satya.png', 155, 10, 20)  # Replace with the path to your small image
                self.image('static/Images/maharashtra_shasn.png', 175, 10,
                           20)  # Replace with the path to your symbol image
                self.cell(0, 10, "Tribal Research & Training Institute, Pune ", align="C", ln=True)
                self.cell(0, 1, "Government of Maharashtra ", align="C", ln=True)
                self.set_font("Arial", "B", size=8)
                self.cell(0, 10,
                          "28, Queen's Garden, Bund Garden Rd, near Old Circuit House, Camp, Pune, Maharashtra 411001 ",
                          align="C", ln=True)
                self.ln(0)  # Reduce the space below the address
                self.line(10, self.get_y(), 200, self.get_y())  # Draw a line from left (10) to right (200)
                self.ln(5)  # Adjust this value to control the space after the line

                self.cell(0, 10,
                          " Fellowship Application Form 2023 - 2024",
                          align="C", ln=True)
                self.ln(2)  # Adjust this value to control the space after the line
                self.line(10, self.get_y(), 200, self.get_y())  # Draw a line from left (10) to right (200)
                self.header_added = True  # Set to True after adding the header

        def image_and_date(self, data):
            # Date and Applicant ID
            self.set_font("Arial", "B", size=10)
            current_date = datetime.now().strftime("%Y-%m-%d")  # You can change the date format as needed
            self.cell(50, 10, "Date: " + current_date, ln=False)

            if 'applicant_id' and 'application_date' in data:
                data['applicant_id'] = 'TRTI' + '/' + str(data['phd_registration_year']) + '/' + str(data['id'])
                self.cell(50, 10, "Applicant ID: " + data['applicant_id'], ln=True)
                self.cell(50, 10, "Application Date : " + str(data['application_date']), ln=True)

            if 'applicant_photo' in data:
                photo = data['applicant_photo']
                print(photo)
                # Insert the applicant photo (adjust the coordinates and size as needed)
                self.image(photo, 165, 60, 30, 35)
                self.rect(165, 60, 30, 35)
                self.ln(10)  # Space between Date/Applicant ID and the main content

        def footer(self):
            # Add a footer
            self.set_y(-15)
            self.set_font("Arial", "B", 8)
            self.cell(0, 10, f" {self.page_no()} ", align="C")

            # Center-align the "TRTI" text
            self.cell(0, 10, " TRTI  |  Fellowship | 2023 - 2024 ", align="R")

    personal_details = {
        "adhaar_number": data['adhaar_number'],
        "first_name": data['first_name'],
        "middle_name": data['middle_name'],
        "last_name": data['last_name'],
        "mobile_number": data['mobile_number'],
        "email": data['email'],
        "date_of_birth": data['date_of_birth'],
        "Gender": data['gender'],
        "Age": data['age'],
        "Caste": data['caste'],
        "Your Caste ": data['your_caste'],
        "Marital Status": data['marital_status']
        # Add more fields as needed
    }

    address_details = {
        "Address Line 1": data['add_1'],
        "Address Line 2": data['add_2'],
        "Pincode": data['pincode'],
        "State": data['state'],
        "City": data['city'],
        "District": data['district'],
        "Taluka": data['taluka'],
        "Village": data['village']
        # Add more fields as needed
    }

    qualification_details = {
        "qualification_level": data['qualification_level'],
        "ssc_passing_year": data['ssc_passing_year'],
        "ssc_percentage": data['ssc_percentage'],
        "ssc_school_name": data['ssc_school_name'],
        "hsc_passing_year": data['hsc_passing_year'],
        "hsc_percentage": data['hsc_percentage'],
        "hsc_school_name": data['hsc_school_name'],
        "graduation_passing_year": data['graduation_passing_year'],
        "graduation_percentage": data['graduation_percentage'],
        "graduation_school_name": data['graduation_school_name'],
        "phd_passing_year": data['phd_passing_year'],
        "phd_percentage": data['phd_percentage'],
        "phd_school_name": data['phd_school_name'],
        "have_you_qualified": data['have_you_qualified']
        # Add more fields as needed
    }

    phd_details = {
        "phd_registration_date": data['phd_registration_date'],
        "phd_registration_year": data['phd_registration_year'],
        "faculty": data['faculty'],
        "concerned_university": data['concerned_university'],
        "topic_of_phd": data['topic_of_phd'],
        "name_of_guide": data['name_of_guide'],
        "name_of_college": data['name_of_college'],
        "Board / University": data['board_university'],
        "Stream": data['stream'],
        "Admission Year": data['admission_year'],
        "Passing Year": data['passing_year'],
        "Percentage": data['percentage']
        # Add more fields as needed
    }

    income_dom_caste_details = {
        "Family Annual Income": data['family_annual_income'],
        "Income Certificate Number": data['income_certificate_number'],
        "Issuing Authority": data['issuing_authority'],
        "domicile": data['domicile'],
        "Domicile Certificate": data['domicile_certificate'],
        "Domicile Number": data['domicile_number'],
        "Caste Certificate": data['caste_certf'],
        "Caste Issuing District": data['issuing_district'],
        "Caste Issuing Authority": data['caste_issuing_authority']
    }

    parent_details = {
        "Salaried": data['salaried'],
        "Disability": data['disability'],
        "Type of Disability": data['type_of_disability'],
        "Father Name": data['father_name'],
        "Mother Name": data['mother_name'],
        "Anyone Work in Government": data['work_in_government'],
    }

    bank_details = {
        "Bank Name": data['bank_name'],
        "Account Number": data['account_number'],
        "IFSC Code": data['ifsc_code'],
        "Account Holder Name": data['account_holder_name']
    }

    pdf = PDF()
    pdf.add_page()
    pdf.header()
    pdf.image_and_date(data)

    # Personal Details

    pdf.cell(0, 10, "Personal Details", ln=True)
    pdf.ln(2)  # Adjust this value to control the space after the line
    pdf.line(10, pdf.get_y(), 200, pdf.get_y())  # Draw a line from left (10) to right (200)
    pdf.header_added = True  # Set to True after adding the header
    pdf.set_font("Arial", size=10)
    for field, value in personal_details.items():
        pdf.cell(40, 10, str(field), ln=False)
        pdf.cell(40, 10, str(value), ln=True)
    pdf.ln(10)  # Adjust this value to control the space after each section

    # Personal Details
    pdf.cell(0, 10, "Address Details", ln=True)
    pdf.ln(2)  # Adjust this value to control the space after the line
    pdf.line(10, pdf.get_y(), 200, pdf.get_y())  # Draw a line from left (10) to right (200)
    pdf.header_added = True  # Set to True after adding the header
    pdf.set_font("Arial", size=10)
    for field, value in address_details.items():
        pdf.cell(40, 10, str(field), ln=False)
        pdf.cell(40, 10, str(value), ln=True)
    pdf.ln(10)  # Adjust this value to control the space after each section

    # Personal Details
    pdf.cell(0, 10, "Qualification Details", ln=True)
    pdf.ln(2)  # Adjust this value to control the space after the line
    pdf.line(10, pdf.get_y(), 200, pdf.get_y())  # Draw a line from left (10) to right (200)
    pdf.header_added = True  # Set to True after adding the header
    pdf.set_font("Arial", size=10)
    for field, value in qualification_details.items():
        pdf.cell(40, 10, str(field), ln=False)
        pdf.cell(40, 10, str(value), ln=True)
    pdf.ln(10)  # Adjust this value to control the space after each section

    # Personal Details
    pdf.cell(0, 10, "P.H.D Details", ln=True)
    pdf.ln(2)  # Adjust this value to control the space after the line
    pdf.line(10, pdf.get_y(), 200, pdf.get_y())  # Draw a line from left (10) to right (200)
    pdf.header_added = True  # Set to True after adding the header
    pdf.set_font("Arial", size=10)
    for field, value in phd_details.items():
        pdf.cell(40, 10, str(field), ln=False)
        pdf.cell(40, 10, str(value), ln=True)
    pdf.ln(10)  # Adjust this value to control the space after each section

    # Personal Details
    pdf.cell(0, 10, "Income / Domicile / Caste Details", ln=True)
    pdf.ln(2)  # Adjust this value to control the space after the line
    pdf.line(10, pdf.get_y(), 200, pdf.get_y())  # Draw a line from left (10) to right (200)
    pdf.header_added = True  # Set to True after adding the header
    pdf.set_font("Arial", size=10)
    for field, value in income_dom_caste_details.items():
        pdf.cell(40, 10, str(field), ln=False)
        pdf.cell(40, 10, str(value), ln=True)
    pdf.ln(10)  # Adjust this value to control the space after each section

    # Personal Details
    pdf.cell(0, 10, "Parent Details", ln=True)
    pdf.ln(2)  # Adjust this value to control the space after the line
    pdf.line(10, pdf.get_y(), 200, pdf.get_y())  # Draw a line from left (10) to right (200)
    pdf.header_added = True  # Set to True after adding the header
    pdf.set_font("Arial", size=10)
    for field, value in parent_details.items():
        pdf.cell(40, 10, str(field), ln=False)
        pdf.cell(40, 10, str(value), ln=True)
    pdf.ln(10)  # Adjust this value to control the space after each section

    # Personal Details
    pdf.cell(0, 10, "Bank Details", ln=True)
    pdf.ln(2)  # Adjust this value to control the space after the line
    pdf.line(10, pdf.get_y(), 200, pdf.get_y())  # Draw a line from left (10) to right (200)
    pdf.header_added = True  # Set to True after adding the header
    pdf.set_font("Arial", size=10)
    for field, value in bank_details.items():
        pdf.cell(40, 10, str(field), ln=False)
        pdf.cell(40, 10, str(value), ln=True)
    pdf.ln(10)  # Adjust this value to control the space after each section

    # Save the PDF to a file
    pdf.output(filename)


@app.route('/generate_pdf', methods=['GET', 'POST'])
def generate_pdf():
    email = session['email']
    output_filename = 'static/pdf_application_form/pdfform.pdf'
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac', host=host, database='ICSApplication')
    cursor = cnx.cursor(dictionary=True)
    cursor.execute(
        "SELECT id, applicant_photo, applicant_id, adhaar_number, first_name, last_name, middle_name, mobile_number,"
        " email, date_of_birth, gender, age, caste, your_caste, marital_status, dependents, state, district,"
        " taluka, village, city, add_1, add_2, pincode, qualification_level, ssc_passing_year,"
        " ssc_percentage, ssc_school_name, hsc_passing_year, hsc_percentage, hsc_school_name,"
        " graduation_passing_year, graduation_percentage, graduation_school_name, phd_passing_year,"
        " phd_percentage, phd_school_name,have_you_qualified, name_of_college, name_of_guide, topic_of_phd,"
        " concerned_university, faculty, phd_registration_date, phd_registration_year, stream,"
        " board_university, admission_year, passing_year, percentage, family_annual_income, "
        "income_certificate_number, issuing_authority, domicile, domicile_certificate, domicile_number,"
        " caste_certf, issuing_district, caste_issuing_authority, salaried, disability, type_of_disability,"
        " father_name, mother_name, work_in_government, bank_name, account_number, ifsc_code,"
        " account_holder_name, application_date FROM application_page WHERE email = %s", (email,))
    data = cursor.fetchone()
    print(data)
    # Generate a styled PDF
    generate_pdf_with_styling(data, output_filename)

    # Serve the generated PDF as a response
    with open(output_filename, "rb") as pdf_file:
        response = Response(pdf_file.read(), content_type="application/pdf")
        response.headers['Content-Disposition'] = 'inline; filename=pdfform.pdf'

    return response


def save_file(file):
    if file:
        filename = file.filename
        file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
        return os.path.join(app.config['UPLOAD_FOLDER'], filename)
    else:
        return "Save File"


def allowed_file(filename):
    ALLOWED_EXTENSIONS = {'pdf', 'txt', 'jpg', 'jpeg', 'png'}
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def insert_into_old_users(email, applicant_id, phd_registration_date, date_of_birth, age):
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor()
    # Update the specific record with the matching email
    with cursor:
        sql = "UPDATE old_users SET applicant_id = %s, phd_registration_date = %s, date_of_birth = %s, age = %s WHERE email = %s"
        cursor.execute(sql, (applicant_id, phd_registration_date, date_of_birth, age, email))
        cnx.commit()
        cursor.close()
        cnx.close()


# --------------------------------------------------------------------
# -------------      FOOTER FUNCTIONALITY     -----------
# --------------------------------------------------------------------
@app.route('/hyperlink_policy')
def hyperlink_policy():
    if 'language' in session:
        language = session['language']
    else:
        language = 'marathi'
    return render_template('hyperlink_policy.html', multilingual_content=multilingual_content, language=language)


@app.route('/t_and_c')
def t_and_c():
    if 'language' in session:
        language = session['language']
    else:
        language = 'marathi'
    return render_template('t_and_c.html', multilingual_content=multilingual_content, language=language)


@app.route('/privacy_policy')
def privacy_policy():
    if 'language' in session:
        language = session['language']
    else:
        language = 'marathi'
    return render_template('privacy_policy.html', multilingual_content=multilingual_content, language=language)


@app.route('/copyright_policy')
def copyright_policy():
    if 'language' in session:
        language = session['language']
    else:
        language = 'marathi'
    return render_template('copyright_policy.html', multilingual_content=multilingual_content, language=language)


@app.route('/wmp_policy')
def wmp_policy():
    if 'language' in session:
        language = session['language']
    else:
        language = 'marathi'
    return render_template('wmp_policy.html', multilingual_content=multilingual_content, language=language)


@app.route('/sitemap')
def sitemap():
    if 'language' in session:
        language = session['language']
    else:
        language = 'marathi'
    return render_template('sitemap.html', multilingual_content=multilingual_content, language=language)


@app.errorhandler(404)
def page_not_found(error):
    return render_template('pages-error-404.html'), 404


@app.errorhandler(500)
def page_not_found_500(error):
    return render_template('pages-error-500.html'), 500


if __name__ == '__main__':
    app.run(debug=True)