from flask import Flask, render_template, make_response, request, redirect, session, url_for, jsonify, \
    flash
from openpyxl import Workbook
from io import BytesIO
import mysql.connector
from openpyxl.packaging.manifest import mimetypes
from flask_mail import Mail, Message
import random
from datetime import datetime, timedelta
import os
# from captcha.image import ImageCaptcha

# Middleware import

# University Class


# ----------- FLASK APP CONFIGURATION -------------------
app = Flask(__name__)
app.config['SECRET_KEY'] = 'rootTanmay'
app.config['SESSION_TYPE'] = 'filesystem'
app.config['SESSION_PERMANENT'] = False
app.config['SESSION_USE_SIGNER'] = True
app.config['SESSION_KEY_PREFIX'] = 'session:'  # Optional, to prevent conflicts
# Session(app)
# ---------------------xxx-------------------------------


# -----------------------------------------------------------------------------------------------------------------
# ------------------------------------------- DATABASE CONFIGURATION ----------------------------------------------
# -----------------------------------------------------------------------------------------------------------------
# ------ HOST Configs are in classes/connection.py
from classes.connection import HostConfig, ConfigPaths

host = HostConfig.host
app_paths = ConfigPaths.paths.get(host)

if app_paths:
    for key, value in app_paths.items():
        app.config[key] = value

cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',  # --------  DATABASE CONNECTION
                              host=host,
                              database='ICSApplication')
cursor = cnx.cursor()
# --------------------------------------- XXXXXXXXXXXXXXXXXXXXXXXXXXXXXXX ----------------------------------------


# ------------ MAIL CONFIGURATION -------------------
# app.config['MAIL_SERVER'] = 'us2.smtp.mailhostbox.com'
# app.config['MAIL_PORT'] = 587
# app.config['MAIL_USERNAME'] = 'helpdesk@trti-maha.in'                     # --------  E-MAIL CONNECTION
# app.config['MAIL_PASSWORD'] = 'FOtIEzp9'
# app.config['MAIL_USE_TLS'] = True
# app.config['MAIL_USE_SSL'] = False
# app.config['MAIL_DEBUG'] = True  # Enable debugging
# mail = Mail(app)
app.config['MAIL_SERVER'] = 'smtp.zeptomail.in'
app.config['MAIL_PORT'] = 587
app.config['MAIL_USERNAME'] = 'noreply_fellowship@trti-maha.in'
app.config['MAIL_PASSWORD'] = 'PHtE6r1YFuzp2TJ69BkFsfewQ8+iPI8v/7hvKABA5IxGCKRVGU0G/t4jkWS/rUsvAPNAFPaYz948tb6c4r2HIGa4N2pIWmqyqK3sx/VYSPOZsbq6x00Vt1gackbeVI/udNVt0S3Vud/fNA=='
app.config['MAIL_USE_TLS'] = True
app.config['MAIL_USE_SSL'] = False
app.config['MAIL_DEBUG'] = True  # Enable debugging
app.config['MAIL_DEFAULT_SENDER'] = ('Your Name', 'noreply_fellowship@trti-maha.in')
mail = Mail(app)
# ---------------------xxx-------------------------------


# -----------------------------------------------------------------------------------------------------------------
# ------------------------------------ IMPORTS OF .PY MODULE FROM PYFiles -----------------------------------------
# -----------------------------------------------------------------------------------------------------------------

# ------------------------------- HOMEPAGE PY FILES AND ROUTES ----------------------------------------
# ---------------------------------------- IMPORT -----------------------------------------
from PyFiles.Homepage.homepage import homepage_blueprint, init_auth
from PyFiles.Homepage.footer_links import footer_links_blueprint, footer_links_auth
from PyFiles.Homepage.login_signup import login_blueprint, login_auth
# ---------------------------------------- APP MAIL CONFIG ---------------------------------
init_auth(app)
footer_links_auth(app)
login_auth(app, mail)
# ---------------------------------------- REGISTER BLUEPRINTS -----------------------------
app.register_blueprint(homepage_blueprint)
app.register_blueprint(footer_links_blueprint)
app.register_blueprint(login_blueprint)
# ------------------------------- END HOMEPAGE PY FILES AND ROUTES -------------------------------------

# ------------------------------- NEW USER PY FILES AND ROUTES ----------------------------------------
# ---------------------------------------- IMPORT -----------------------------------------
from PyFiles.Candidate.NewUser.ApplicationForm.section1 import section1_blueprint, section1_auth
from PyFiles.Candidate.NewUser.ApplicationForm.section2 import section2_blueprint, section2_auth
from PyFiles.Candidate.NewUser.ApplicationForm.section3 import section3_blueprint, section3_auth
from PyFiles.Candidate.NewUser.ApplicationForm.section4 import section4_blueprint, section4_auth
from PyFiles.Candidate.NewUser.ApplicationForm.section5 import section5_blueprint, section5_auth
# ---------------------------------------- APP MAIL CONFIG ---------------------------------
section1_auth(app)
section2_auth(app)
section3_auth(app)
section4_auth(app)
section5_auth(app)
# ---------------------------------------- REGISTER BLUEPRINTS -----------------------------
app.register_blueprint(section1_blueprint)
app.register_blueprint(section2_blueprint)
app.register_blueprint(section3_blueprint)
app.register_blueprint(section4_blueprint)
app.register_blueprint(section5_blueprint)

# ------------------------------- END NEW USER PY FILES AND ROUTES -------------------------------------

# ------------------------------- OLD USER PY FILES AND ROUTES ----------------------------------------
# ---------------------------------------- IMPORT -----------------------------------------
from PyFiles.Candidate.OldUser.old_user_form import old_user_blueprint, old_user_auth
from PyFiles.Candidate.OldUser.view_forms import viewform_blueprint, viewform_auth
# ---------------------------------------- APP MAIL CONFIG ---------------------------------
old_user_auth(app)
viewform_auth(app)

# ---------------------------------------- REGISTER BLUEPRINTS -----------------------------
app.register_blueprint(old_user_blueprint)
app.register_blueprint(viewform_blueprint)

# ------------------------------- END OLD USER PY FILES AND ROUTES -------------------------------------

# ------------------------------- Common USER PY FILES AND ROUTES ----------------------------------------
# ---------------------------------------- IMPORT -----------------------------------------
from PyFiles.Candidate.commonFiles.generate_pdf import generate_pdf_blueprint, generate_pdf_auth
from PyFiles.Candidate.commonFiles.my_profile import my_profile_blueprint, my_profile_auth
from PyFiles.Candidate.commonFiles.issues_raised import issue_raised_blueprint, issue_raised_auth

# ---------------------------------------- APP MAIL CONFIG ---------------------------------
generate_pdf_auth(app)
my_profile_auth(app, mail)
issue_raised_auth(app, mail)

# ---------------------------------------- REGISTER BLUEPRINTS -----------------------------
app.register_blueprint(generate_pdf_blueprint)
app.register_blueprint(my_profile_blueprint)
app.register_blueprint(issue_raised_blueprint)

# ------------------------------- END OLD USER PY FILES AND ROUTES -------------------------------------

# ------------------------------- Dashboard for Accepted USER PY FILES AND ROUTES ----------------------------------------
# ---------------------------------------- IMPORT -----------------------------------------
from PyFiles.Candidate.AcceptedUserDashboard.adhaar_seeding import adhaarseed_blueprint, adhaarseed_auth
from PyFiles.Candidate.AcceptedUserDashboard.manage_profile import manageprofile_blueprint, manageprofile_auth
from PyFiles.Candidate.AcceptedUserDashboard.awardletter import awardletter_blueprint, awardletter_auth
from PyFiles.Candidate.AcceptedUserDashboard.joining_report import joiningreport_blueprint, joiningreport_auth
from PyFiles.Candidate.AcceptedUserDashboard.undertakingdoc import undertakingdoc_blueprint, undertakingdoc_auth
from PyFiles.Candidate.AcceptedUserDashboard.presenty import presenty_blueprint, presenty_auth
from PyFiles.Candidate.AcceptedUserDashboard.halfyearlyrep import halfyearly_blueprint, halfyearly_auth
from PyFiles.Candidate.AcceptedUserDashboard.hracert import hracert_blueprint, hracert_auth
from PyFiles.Candidate.AcceptedUserDashboard.assessment_rep import assessment_blueprint, assessment_auth
from PyFiles.Candidate.AcceptedUserDashboard.withdraw import withdraw_blueprint, withdraw_auth
from PyFiles.Candidate.AcceptedUserDashboard.change_guide import changeguide_blueprint, changeguide_auth
from PyFiles.Candidate.AcceptedUserDashboard.change_research import changeres_blueprint, changeres_auth
from PyFiles.Candidate.AcceptedUserDashboard.phdaward import phdaward_blueprint, phdaward_auth
from PyFiles.Candidate.AcceptedUserDashboard.thesis import thesis_blueprint, thesis_auth
from PyFiles.Candidate.AcceptedUserDashboard.Installments.installments import installments_blueprint, installments_auth

# ---------------------------------------- APP MAIL CONFIG ---------------------------------
adhaarseed_auth(app)
manageprofile_auth(app, mail)
awardletter_auth(app)
joiningreport_auth(app)
undertakingdoc_auth(app)
presenty_auth(app)
halfyearly_auth(app)
hracert_auth(app)
assessment_auth(app)
withdraw_auth(app)
changeguide_auth(app)
changeres_auth(app)
phdaward_auth(app)
thesis_auth(app)
installments_auth(app)

# ---------------------------------------- REGISTER BLUEPRINTS -----------------------------
app.register_blueprint(adhaarseed_blueprint)
app.register_blueprint(manageprofile_blueprint)
app.register_blueprint(awardletter_blueprint)
app.register_blueprint(joiningreport_blueprint)
app.register_blueprint(undertakingdoc_blueprint)
app.register_blueprint(presenty_blueprint)
app.register_blueprint(halfyearly_blueprint)
app.register_blueprint(hracert_blueprint)
app.register_blueprint(assessment_blueprint)
app.register_blueprint(withdraw_blueprint)
app.register_blueprint(changeguide_blueprint)
app.register_blueprint(changeres_blueprint)
app.register_blueprint(phdaward_blueprint)
app.register_blueprint(thesis_blueprint)
app.register_blueprint(installments_blueprint)

# ------------------------------- END Dashboard for Accepted USER PY FILES AND ROUTES -------------------------------------

# ------------------------------- ADMIN DASHBOARD PY FILES AND ROUTES ----------------------------------------
# ---------------------------------------- IMPORT -----------------------------------------
from PyFiles.Admin.adminlogin import adminlogin_blueprint, adminlogin_auth
from PyFiles.Admin.index import index_blueprint, index_auth

from PyFiles.Admin.dashboardCounts.applicationCounts import applCounts_blueprint, applCounts_auth
from PyFiles.Admin.dashboardCounts.genderCounts import genderCounts_blueprint, gendercounts_auth
from PyFiles.Admin.dashboardCounts.pvtgCounts import pvtgCounts_blueprint, pvtgCounts_auth
from PyFiles.Admin.dashboardCounts.disableCounts import disabilityCount_blueprint, disabilityCount_auth

from PyFiles.Admin.generate_reports import generate_reports_blueprint, generate_reports_auth
from PyFiles.Admin.AdminLevels.admin_level_one import adminlevelone_blueprint, adminlevelone_auth
from PyFiles.Admin.AdminLevels.admin_level_two import adminleveltwo_blueprint, adminleveltwo_auth
from PyFiles.Admin.AdminLevels.admin_level_three import adminlevelthree_blueprint, adminlevelthree_auth

from PyFiles.Admin.PaymentSheet.payment_sheet import payment_sheet_blueprint, payment_sheet_auth
from PyFiles.Admin.PaymentSheet.fellowship_details import fellowshipdetails_blueprint, fellowshipdetails_auth
from PyFiles.Admin.budget import budget_blueprint, budget_auth
from PyFiles.Admin.payment_tracking import payment_tracking_blueprint, payment_tracking_auth
from PyFiles.Admin.fellowship_awarded import fellowship_awarded_blueprint, fellowship_awarded_auth
from PyFiles.Admin.withdrawed_appl import withdrawed_application_blueprint, withdrawed_application_auth
from PyFiles.Admin.bulkemails import bulkemails_blueprint, bulkemails_auth

from PyFiles.Admin.ManageStudents.studentmanage import managestud_blueprint, managestud_auth
from PyFiles.Admin.issue_raisedAdmin import issue_raisedAdmin_blueprint, issue_raisedAdmin_auth
from PyFiles.Admin.news import news_blueprint, news_auth

# ---------------------------------------- APP MAIL CONFIG ---------------------------------
adminlogin_auth(app)
index_auth(app)
applCounts_auth(app)
gendercounts_auth(app)
pvtgCounts_auth(app)
disabilityCount_auth(app)

generate_reports_auth(app)
adminlevelone_auth(app, mail)
adminleveltwo_auth(app, mail)
adminlevelthree_auth(app, mail)

payment_sheet_auth(app)
fellowshipdetails_auth(app)
budget_auth(app)
payment_tracking_auth(app)
fellowship_awarded_auth(app)
withdrawed_application_auth(app)
bulkemails_auth(app, mail)

managestud_auth(app)
issue_raisedAdmin_auth(app)
news_auth(app)

# ---------------------------------------- REGISTER BLUEPRINTS -----------------------------
app.register_blueprint(adminlogin_blueprint)
app.register_blueprint(index_blueprint)
app.register_blueprint(applCounts_blueprint)
app.register_blueprint(genderCounts_blueprint)
app.register_blueprint(pvtgCounts_blueprint)
app.register_blueprint(disabilityCount_blueprint)

app.register_blueprint(generate_reports_blueprint)
app.register_blueprint(adminlevelone_blueprint)
app.register_blueprint(adminleveltwo_blueprint)
app.register_blueprint(adminlevelthree_blueprint)

app.register_blueprint(payment_sheet_blueprint)
app.register_blueprint(fellowshipdetails_blueprint)
app.register_blueprint(budget_blueprint)
app.register_blueprint(payment_tracking_blueprint)
app.register_blueprint(fellowship_awarded_blueprint)
app.register_blueprint(withdrawed_application_blueprint)
app.register_blueprint(bulkemails_blueprint)

app.register_blueprint(managestud_blueprint)
app.register_blueprint(issue_raisedAdmin_blueprint)
app.register_blueprint(news_blueprint)

# ------------------------------- END ADMIN DASHBOARD PY FILES AND ROUTES -------------------------------------

# -------------------------------- EXPORT TO EXCELS -------------------------------------
from PyFiles.Admin.Export_to_Excel.adminlevels import adminlevels_blueprint, adminlevels_auth

adminlevels_auth(app)

app.register_blueprint(adminlevels_blueprint)

# -----------------------------------------------------------------------------------------------------------------
# ------------------------------------ END IMPORTS OF .PY MODULE FROM PYFiles -------------------------------------
# -----------------------------------------------------------------------------------------------------------------


@app.route('/set_session/<value>')
def set_session(value):
    session['language'] = value
    return redirect(request.referrer)


def old_user_incomplete_form(email, form_filled):                                     # -------------- CHECK IF USER HAS FILLED THE FORM
    email = session['email']
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor()
    sql = "SELECT form_filled FROM application_page WHERE email = %s AND form_filled = %s"
    cursor.execute(sql, (email,form_filled))
    result = cursor.fetchone()
    cursor.close()
    cnx.close()
    return result is not None


@app.route('/login_closed_2023')
def login_closed_2023():
    return render_template('loginClosed2023.html')


# -------------------


@app.route('/refresh_login_captcha', methods=['GET'])
def refresh_login_captcha():
    captcha_number = random.randrange(100000, 999999)
    img = ImageCaptcha(width=280, height=90)
    captcha_text = str(captcha_number)
    img.write(captcha_text, 'static/Images/captcha/user_captcha.png')

    # Update the captcha number in the session
    session['captcha_number'] = captcha_number

    # Redirect back to the login page after refreshing the Captcha
    return redirect(url_for('login'))



@app.route('/refresh_signup_captcha', methods=['GET'])
def refresh_signup_captcha():
    captcha_number = random.randrange(100000, 999999)
    img = ImageCaptcha(width=280, height=90)
    captcha_text = str(captcha_number)
    img.write(captcha_text, 'static/Images/captcha/user_captcha.png')

    # Update the captcha number in the session
    session['captcha_number'] = captcha_number

    # Redirect back to the login page after refreshing the Captcha
    return redirect(url_for('signup'))


def submit_edit_profile():                                      # ------------- SUBMIT EDIT PROFILE FOR MY PROFILE
    if request.method == 'POST':
        first_name = request.form['first_name']
        address = request.form['add_1']
        phone = request.form['mobile_number']
        email = session['email']
        cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                      host=host,
                                      database='ICSApplication')
        cursor = cnx.cursor(dictionary=True)
        # Construct the SQL update query
        sql = f"UPDATE application_page SET first_name = '{first_name}', add_1 = '{address}', mobile_number = '{phone}' WHERE email = '{email}'"
        cursor.execute(sql)
        cnx.commit()
        cursor.close()
        cnx.close()


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in {'png', 'jpg', 'jpeg', 'gif', 'bmp', 'pdf'}


#                                               ---- ADMIN LEVEL 1, 2, 3
# ------------------ ADMIN LEVEL 1 ----------------

# ------------------ END ADMIN LEVEL 1 ----------------





#---------------By Aditya---Rejected applications at level 1--------------------------
@app.route('/rejected_applications_l1', methods=['GET', 'POST'])
def rejected_at_level1():
    if not session.get('logged_in'):
    # Redirect to the admin login page if the user is not logged in
        return redirect(url_for('admin_login'))
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                    host=host,
                                    database='ICSApplication')
    cursor = cnx.cursor(dictionary=True)
    status = request.get_data()

    cursor.execute("SELECT * FROM application_page WHERE status='rejected'")
    data = cursor.fetchall()
    print(data)
    cursor.close()
    cnx.close()
    return render_template('rejected_applications_l1.html', data=data)

#---------------By Aditya---Rejected applications at level 2--------------------------
@app.route('/rejected_applications_l2', methods=['GET', 'POST'])
def rejected_at_level2():
    if not session.get('logged_in'):
    # Redirect to the admin login page if the user is not logged in
        return redirect(url_for('admin_login'))
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                    host=host,
                                    database='ICSApplication')
    cursor = cnx.cursor(dictionary=True)
    scrutiny_status = request.get_data()

    cursor.execute("SELECT * FROM application_page WHERE scrutiny_status='rejected'")
    data = cursor.fetchall()
    print(data)
    cursor.close()
    cnx.close()
    return render_template('rejected_applications_l2.html', data=data)


@app.route('/export_data', methods=['GET', 'POST'])
def export_data():
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor()
    cursor.execute("SELECT * FROM application_page")
    data = cursor.fetchall()
    # Create a workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Add headers to the worksheet
    ws.append(['Applicant Photo', 'Applicant Id', 'First Name', 'Middle Name', 'Last Name', 'Mobile Number', 'Email', 
               'Date Of Birth', 'Gender', 'Age', 'Caste', 'Your Caste', 'Marital Status', 'Dependents', 'Add 1', 'Add 2', 
               'Pincode', 'Village', 'Taluka', 'District', 'State', 'Phd Registration Date', 'Concerned University', 
               'Topic Of Phd', 'Name Of Guide', 'Name Of College', 'Stream', 'Board University', 'Admission Year', 
               'Passing Year', 'Percentage', 'Family Annual Income', 'Income Certificate Number', 'Issuing Authority',
               'Domicile', 'Domicile Certificate', 'Relation', 'Domicile Number', 'Caste', 'Caste Category', 
               'Caste Certificate Number', 'Issuing District', 'Caste Applicant Name', 'Caste Issuing Authority',
               'Salaried', 'Disability', 'Father Alive', 'Father Name', 'Mother Alive', 'Mother Name', 'Work In Government', 
               'Bank Name', 'Account Number', 'Ifsc Code', 'Account Holder Name', 'Documentfile1', 'Documentfile2',
               'Documentfile3', 'Documentfile4', 'Documentfile5', 'Documentfile6', 'Documentfile7', 'Documentfile8',
               'Documentfile9', 'Current Date', 'Formatted Datetime'])

    # Add data to the worksheet
    for row in data:
        ws.append(row)

    # Save the workbook in memory as bytes
    data = BytesIO()
    wb.save(data)
    data.seek(0)

    # Create a response object and attach the workbook as a file
    response = make_response(data.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=ApplicantData.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    return response


@app.route('/export')
def export():
    export_data()


#                                           ------- ADMIN PAYMENT SHEET



#----------------------  EXPORT TO EXCEL Payment Sheet (2023-2024) -------------------------------------------
@app.route('/export_payment_sheet', methods=['GET', 'POST'])
def export_payment_sheet():
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor()
    cursor.execute("SELECT number, full_name, email, faculty, duration_date_from, duration_date_to, total_months, fellowship," 
                    "to_fellowship, rate, amount, months, total_hra, count, pwd, total, city FROM payment_sheet")

    data = cursor.fetchall()
    # Create a workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    ws.append(['Applicant ID', 'Student Name', 'Email', 'Faculty',  'Duration Date From', 'Duration Date To', 'Total Months',
             'Fellowship', 'Total Fellowship', 'H.R.A Rate', 'H.R.A Amount', 'Months', 'Total H.R.A', 'Count Yearly', 'P.W.D', 'Total', 'City'])


    # Add data to the worksheet
    for row in data:
        ws.append(row)

    # Save the workbook in memory as bytes
    data = BytesIO()
    wb.save(data)
    data.seek(0)

    # Create a response object and attach the workbook as a files
    response = make_response(data.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=Payment_Sheet_2023_2024.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    return response



#----------------------  EXPORT TO EXCEL Payment Tracking Sheet-------------------------------------------
@app.route('/export_payment_tracking_sheet', methods=['GET', 'POST'])
def export_payment_tracking_sheet():
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor()
    cursor.execute("SELECT ap.*, ps.* FROM application_page ap JOIN payment_sheet ps ON ap.email = ps.email WHERE ap.final_approval = 'accepted' AND ap.joining_report IS NOT NULL  ")
    data = cursor.fetchall()
    # Create a workbook and add a worksheet
    wb = Workbook()
    ws = wb.active


    ws.append(['Applicant Id', 'First Name', 'Middle Name', 'Last Name', 'Mobile Number', 'Email', 'Date Of Birth', 
                'Gender', 'Age', 'Caste', 'Your Caste', 'Marital Status', 'Dependents', 'Add 1', 'Add 2', 'Pincode', 'Village', 
                'Taluka', 'District', 'State', 'Phd Registration Date', 'Concerned University', 'Topic Of Phd', 'Name Of Guide', 
                'Name Of College', 'Stream', 'Board University', 'Admission Year', 'Passing Year', 'Percentage', 'Family Annual Income', 
                'Income Certificate Number', 'Issuing Authority', 'Domicile', 'Domicile Certificate', 'Domicile Number', 'Caste Certf', 
                'Issuing District', 'Caste Issuing Authority', 'Salaried', 'Disability', 'Father Name', 'Mother Name', 'Work In Government', 
                'Bank Name', 'Account Number', 'Ifsc Code', 'Account Holder Name']	)

    # Add data to the worksheet
    for row in data:
        ws.append(row)

    # Save the workbook in memory as bytes
    data = BytesIO()
    wb.save(data)
    data.seek(0)

    # Create a response object and attach the workbook as a file
    response = make_response(data.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=Payment_Tracking_Sheet.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    return response











@app.route('/get_year_count', methods=['GET', 'POST'])
def get_year_count():
    if request.method == 'POST':
        data = request.form['selected_year']
        print(data)
        # selected_year = data.get('selected_year')
        count = f""" SELECT count(*) FROM application_page where phd_registration_year = {data} """
        print(count)
        count_accept = f""" SELECT count(*) FROM application_page where phd_registration_year = {data} and final_approval = 'accepted' """
        count_reject = f""" SELECT count(*) FROM application_page where phd_registration_year = {data} and final_approval = 'rejected' """
        # Execute the queries
        cursor.execute(count)
        total_result = cursor.fetchone()
        total_count = total_result[0] if total_result else 0

        cursor.execute(count_accept)
        accept_result = cursor.fetchone()
        accept_count = accept_result[0] if accept_result else 0

        cursor.execute(count_reject)
        reject_result = cursor.fetchone()
        reject_count = reject_result[0] if reject_result else 0
        result = cursor.fetchone()  # Fetch the result of the query
        count = result[0] if result else 0  # Extract the count from the result
        cnx.commit()
        return jsonify({'total_count': total_count, 'accept_count': accept_count, 'reject_count': reject_count})
    return 'INAVALID REQUEST'

#                                                     ---- ADMIN DASHBOARD




# ---------------------  DISABILITY FUNCTIONALITY ---------------------------------
# ---------------------------------------------------------------------------------
@app.route('/disability_report_yes', methods=['GET', 'POST'])
def disability_report_yes():
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor(dictionary=True)
    cursor.execute("SELECT * FROM application_page WHERE disability='Yes'")
    result = cursor.fetchall()
    cnx.commit()
    cursor.close()
    cnx.close()
    return render_template('disability_report_yes.html', result=result)


#----------------------  EXPORT TO EXCEL Applicants with Disability -------------------------------------------
@app.route('/export_disability_report_yes', methods=['GET', 'POST'])
def export_disability_report_yes():
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor()
    cursor.execute("SELECT * from application_page WHERE disability ='Yes' ")
    data = cursor.fetchall()
    # Create a workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Add headers to the worksheet
    ws.append(['Applicant Id', 'First Name', 'Middle Name', 'Last Name', 'Mobile Number', 'Email', 'Date Of Birth', 
                'Gender', 'Age', 'Caste', 'Your Caste', 'Marital Status', 'Dependents', 'Add 1', 'Add 2', 'Pincode', 'Village', 
                'Taluka', 'District', 'State', 'Phd Registration Date', 'Concerned University', 'Topic Of Phd', 'Name Of Guide', 
                'Name Of College', 'Stream', 'Board University', 'Admission Year', 'Passing Year', 'Percentage', 'Family Annual Income', 
                'Income Certificate Number', 'Issuing Authority', 'Domicile', 'Domicile Certificate', 'Domicile Number', 'Caste Certf', 
                'Issuing District', 'Caste Issuing Authority', 'Salaried', 'Disability', 'Father Name', 'Mother Name', 'Work In Government', 
                'Bank Name', 'Account Number', 'Ifsc Code', 'Account Holder Name']	)

    # Add data to the worksheet
    for row in data:
        ws.append(row)

    # Save the workbook in memory as bytes
    data = BytesIO()
    wb.save(data)
    data.seek(0)

    # Create a response object and attach the workbook as a file
    response = make_response(data.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=Applicants_with_disability.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    return response    


@app.route('/disability_report_no', methods=['GET', 'POST'])
def disability_report_no():
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor(dictionary=True)
    cursor.execute("SELECT * FROM application_page WHERE disability='No'")
    result = cursor.fetchall()
    cnx.commit()
    cursor.close()
    cnx.close()
    print(result)
    return render_template('disability_report_no.html', result=result)


#----------------------  EXPORT TO EXCEL Applicants WITHOUT Disability -------------------------------------------
@app.route('/export_disability_report_no', methods=['GET', 'POST'])
def export_disability_report_no():
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor()
    cursor.execute("SELECT * from application_page WHERE disability ='No' ")
    data = cursor.fetchall()
    # Create a workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Add headers to the worksheet
    ws.append(['Applicant Id', 'First Name', 'Middle Name', 'Last Name', 'Mobile Number', 'Email', 'Date Of Birth', 
                'Gender', 'Age', 'Caste', 'Your Caste', 'Marital Status', 'Dependents', 'Add 1', 'Add 2', 'Pincode', 'Village', 
                'Taluka', 'District', 'State', 'Phd Registration Date', 'Concerned University', 'Topic Of Phd', 'Name Of Guide', 
                'Name Of College', 'Stream', 'Board University', 'Admission Year', 'Passing Year', 'Percentage', 'Family Annual Income', 
                'Income Certificate Number', 'Issuing Authority', 'Domicile', 'Domicile Certificate', 'Domicile Number', 'Caste Certf', 
                'Issuing District', 'Caste Issuing Authority', 'Salaried', 'Disability', 'Father Name', 'Mother Name', 'Work In Government', 
                'Bank Name', 'Account Number', 'Ifsc Code', 'Account Holder Name']	)

    # Add data to the worksheet
    for row in data:
        ws.append(row)

    # Save the workbook in memory as bytes
    data = BytesIO()
    wb.save(data)
    data.seek(0)

    # Create a response object and attach the workbook as a file
    response = make_response(data.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=Applicants_without_disability.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    return response


def disability_yes_count_report():       # ----- To count users with NO disability  ------
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor()
    cursor.execute("SELECT COUNT(*) FROM application_page WHERE disability='Yes'")
    result = cursor.fetchone()
    print(result)
    return result[0]


def disability_no_count_report():       # ----- To count users with NO disability  ------
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor()
    cursor.execute("SELECT COUNT(*) FROM application_page WHERE disability='no'")
    result = cursor.fetchone()
    print(result)
    return result[0]


def type_disability_physically():       # ----- To count users with NO disability  ------
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor()
    cursor.execute("SELECT COUNT(*) FROM application_page WHERE type_of_disability='Physically Handicapped'")
    result = cursor.fetchone()
    print(result)
    return result[0]    


def type_disability_visually():       # ----- To count users with NO disability  ------
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor()
    cursor.execute("SELECT COUNT(*) FROM application_page WHERE type_of_disability='Visually Impaired'")
    result = cursor.fetchone()
    print(result)
    return result[0]


def type_disability_hearing():       # ----- To count users with NO disability  ------
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor()
    cursor.execute("SELECT COUNT(*) FROM application_page WHERE type_of_disability='Hearing Impaired'")
    result = cursor.fetchone()
    print(result)
    return result[0]


@app.route('/disabilty_report_physical', methods=['GET', 'POST'])
def disabilty_report_physical():
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor(dictionary=True)
    cursor.execute("SELECT * FROM application_page WHERE type_of_disability='Physically Handicapped'")
    result = cursor.fetchall()
    cnx.commit()
    cursor.close()
    cnx.close()
    print(result)
    return render_template('disabilty_report_physical.html', result=result)


#----------------------  EXPORT TO EXCEL Physically Handicapped Applicants -------------------------------------------
@app.route('/export_physically_handicapped_report', methods=['GET', 'POST'])
def export_physically_handicapped_report():
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor()
    cursor.execute("SELECT applicant_id, first_name, middle_name, last_name, mobile_number, email, date_of_birth,"
                   "gender, age, caste, your_caste, marital_status, add_1, add_2, pincode, village, taluka, district,"
                   "state, city,  phd_registration_date, concerned_university, topic_of_phd, "
                   "name_of_guide, name_of_college, stream, board_university, admission_year, passing_year, percentage,"
                   "family_annual_income, income_certificate_number, issuing_authority, domicile, domicile_certificate,"
                   "domicile_number, caste_certf, issuing_district, caste_issuing_authority, salaried, disability,"
                   "father_name, mother_name, work_in_government, bank_name, account_number, ifsc_code," 
                   "account_holder_name FROM application_page WHERE type_of_disability='Physically Handicapped'")
    data = cursor.fetchall()
    # Create a workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Add headers to the worksheet
    ws.append(['Applicant Id', 'First Name', 'Middle Name', 'Last Name', 'Mobile Number', 'Email', 'Date Of Birth', 
                'Gender', 'Age', 'Caste', 'Your Caste', 'Marital Status', 'Dependents', 'Add 1', 'Add 2', 'Pincode', 'Village', 
                'Taluka', 'District', 'State', 'Phd Registration Date', 'Concerned University', 'Topic Of Phd', 'Name Of Guide', 
                'Name Of College', 'Stream', 'Board University', 'Admission Year', 'Passing Year', 'Percentage', 'Family Annual Income', 
                'Income Certificate Number', 'Issuing Authority', 'Domicile', 'Domicile Certificate', 'Domicile Number', 'Caste Certf', 
                'Issuing District', 'Caste Issuing Authority', 'Salaried', 'Disability', 'Father Name', 'Mother Name', 'Work In Government', 
                'Bank Name', 'Account Number', 'Ifsc Code', 'Account Holder Name']	)

    # Add data to the worksheet
    for row in data:
        ws.append(row)

    # Save the workbook in memory as bytes
    data = BytesIO()
    wb.save(data)
    data.seek(0)

    # Create a response object and attach the workbook as a file
    response = make_response(data.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=Physically_handicapped_applicants.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    return response      


@app.route('/disabilty_report_visual', methods=['GET', 'POST'])
def disabilty_report_visual():
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor(dictionary=True)
    cursor.execute("SELECT * FROM application_page WHERE type_of_disability='Visually Impaired'")
    result = cursor.fetchall()
    cnx.commit()
    cursor.close()
    cnx.close()
    print(result)
    return render_template('disabilty_report_visual.html', result=result)


#----------------------  EXPORT TO EXCEL Visually Impaired Applicants -------------------------------------------
@app.route('/export_visually_impaired_report', methods=['GET', 'POST'])
def export_visually_impaired_report():
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor()
    cursor.execute("SELECT applicant_id, first_name, middle_name, last_name, mobile_number, email, date_of_birth,"
                   "gender, age, caste, your_caste, marital_status, add_1, add_2, pincode, village, taluka, district,"
                   "state, city,  phd_registration_date, concerned_university, topic_of_phd, "
                   "name_of_guide, name_of_college, stream, board_university, admission_year, passing_year, percentage,"
                   "family_annual_income, income_certificate_number, issuing_authority, domicile, domicile_certificate,"
                   "domicile_number, caste_certf, issuing_district, caste_issuing_authority, salaried, disability,"
                   "father_name, mother_name, work_in_government, bank_name, account_number, ifsc_code," 
                   "account_holder_name FROM application_page WHERE type_of_disability='Visually Impaired'")
    data = cursor.fetchall()
    # Create a workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Add headers to the worksheet
    ws.append(['Applicant Id', 'First Name', 'Middle Name', 'Last Name', 'Mobile Number', 'Email', 'Date Of Birth', 
                'Gender', 'Age', 'Caste', 'Your Caste', 'Marital Status', 'Dependents', 'Add 1', 'Add 2', 'Pincode', 'Village', 
                'Taluka', 'District', 'State', 'Phd Registration Date', 'Concerned University', 'Topic Of Phd', 'Name Of Guide', 
                'Name Of College', 'Stream', 'Board University', 'Admission Year', 'Passing Year', 'Percentage', 'Family Annual Income', 
                'Income Certificate Number', 'Issuing Authority', 'Domicile', 'Domicile Certificate', 'Domicile Number', 'Caste Certf', 
                'Issuing District', 'Caste Issuing Authority', 'Salaried', 'Disability', 'Father Name', 'Mother Name', 'Work In Government', 
                'Bank Name', 'Account Number', 'Ifsc Code', 'Account Holder Name']	)

    # Add data to the worksheet
    for row in data:
        ws.append(row)

    # Save the workbook in memory as bytes
    data = BytesIO()
    wb.save(data)
    data.seek(0)

    # Create a response object and attach the workbook as a file
    response = make_response(data.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=Visually_impaired_applicants.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    return response         


@app.route('/disabilty_report_hearing', methods=['GET', 'POST'])
def disabilty_report_hearing():
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor(dictionary=True)
    cursor.execute("SELECT * FROM application_page WHERE type_of_disability='Hearing Impaired'")
    result = cursor.fetchall()
    cnx.commit()
    cursor.close()
    cnx.close()
    print(result)
    return render_template('disabilty_report_hearing.html', result=result)


#----------------------  EXPORT TO EXCEL Hearing Impaired Applicants -------------------------------------------
@app.route('/export_hearing_impaired_report', methods=['GET', 'POST'])
def export_hearing_impaired_report():
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor()
    cursor.execute("SELECT applicant_id, first_name, middle_name, last_name, mobile_number, email, date_of_birth,"
                   "gender, age, caste, your_caste, marital_status, add_1, add_2, pincode, village, taluka, district,"
                   "state, city,  phd_registration_date, concerned_university, topic_of_phd, "
                   "name_of_guide, name_of_college, stream, board_university, admission_year, passing_year, percentage,"
                   "family_annual_income, income_certificate_number, issuing_authority, domicile, domicile_certificate,"
                   "domicile_number, caste_certf, issuing_district, caste_issuing_authority, salaried, disability,"
                   "father_name, mother_name, work_in_government, bank_name, account_number, ifsc_code," 
                   "account_holder_name FROM application_page WHERE type_of_disability='Hearing Impaired'")
    data = cursor.fetchall()
    # Create a workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Add headers to the worksheet
    ws.append(['Applicant Id', 'First Name', 'Middle Name', 'Last Name', 'Mobile Number', 'Email', 'Date Of Birth', 
                'Gender', 'Age', 'Caste', 'Your Caste', 'Marital Status', 'Dependents', 'Add 1', 'Add 2', 'Pincode', 'Village', 
                'Taluka', 'District', 'State', 'Phd Registration Date', 'Concerned University', 'Topic Of Phd', 'Name Of Guide', 
                'Name Of College', 'Stream', 'Board University', 'Admission Year', 'Passing Year', 'Percentage', 'Family Annual Income', 
                'Income Certificate Number', 'Issuing Authority', 'Domicile', 'Domicile Certificate', 'Domicile Number', 'Caste Certf', 
                'Issuing District', 'Caste Issuing Authority', 'Salaried', 'Disability', 'Father Name', 'Mother Name', 'Work In Government', 
                'Bank Name', 'Account Number', 'Ifsc Code', 'Account Holder Name']	)
    # Add data to the worksheet
    for row in data:
        ws.append(row)

    # Save the workbook in memory as bytes
    data = BytesIO()
    wb.save(data)
    data.seek(0)

    # Create a response object and attach the workbook as a file
    response = make_response(data.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=Hearing_impaired_applicants.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    return response     



@app.route('/male_record_report', methods=['GET', 'POST'])
def male_record_report():
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor(dictionary=True)
    cursor.execute("SELECT * FROM application_page WHERE gender='male'")
    result = cursor.fetchall()
    cnx.commit()
    cursor.close()
    cnx.close()
    #print(result)
    return render_template('male_record_report.html', result=result)


#----------------------  EXPORT TO EXCEL Male Applications -------------------------------------------
@app.route('/export_male_applications', methods=['GET', 'POST'])
def export_male_applications():
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor()
    cursor.execute("SELECT applicant_id, first_name, middle_name, last_name, mobile_number, email, date_of_birth,"
                   "gender, age, caste, your_caste, marital_status, add_1, add_2, pincode, village, taluka, district,"
                   "state, city,  phd_registration_date, concerned_university, topic_of_phd, "
                   "name_of_guide, name_of_college, stream, board_university, admission_year, passing_year, percentage,"
                   "family_annual_income, income_certificate_number, issuing_authority, domicile, domicile_certificate,"
                   "domicile_number, caste_certf, issuing_district, caste_issuing_authority, salaried, disability,"
                   "father_name, mother_name, work_in_government, bank_name, account_number, ifsc_code," 
                   "account_holder_name FROM application_page WHERE gender ='male' ")
    data = cursor.fetchall()
    # Create a workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Add headers to the worksheet
    ws.append(['Applicant Id', 'First Name', 'Middle Name', 'Last Name', 'Mobile Number', 'Email', 'Date Of Birth', 
                'Gender', 'Age', 'Caste', 'Your Caste', 'Marital Status', 'Dependents', 'Add 1', 'Add 2', 'Pincode', 'Village', 
                'Taluka', 'District', 'State', 'Phd Registration Date', 'Concerned University', 'Topic Of Phd', 'Name Of Guide', 
                'Name Of College', 'Stream', 'Board University', 'Admission Year', 'Passing Year', 'Percentage', 'Family Annual Income', 
                'Income Certificate Number', 'Issuing Authority', 'Domicile', 'Domicile Certificate', 'Domicile Number', 'Caste Certf', 
                'Issuing District', 'Caste Issuing Authority', 'Salaried', 'Disability', 'Father Name', 'Mother Name', 'Work In Government', 
                'Bank Name', 'Account Number', 'Ifsc Code', 'Account Holder Name']	)

    # Add data to the worksheet
    for row in data:
        ws.append(row)

    # Save the workbook in memory as bytes
    data = BytesIO()
    wb.save(data)
    data.seek(0)

    # Create a response object and attach the workbook as a file
    response = make_response(data.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=Male_Applicants_Data.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    return response    


@app.route('/female_record_report', methods=['GET', 'POST'])
def female_record_report():
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor(dictionary=True)
    cursor.execute("SELECT * FROM application_page WHERE gender='female'")
    result = cursor.fetchall()
    cnx.commit()
    cursor.close()
    cnx.close()
    print(result)
    return render_template('female_record_report.html', result=result)


#----------------------  EXPORT TO EXCEL Female Applications -------------------------------------------
@app.route('/export_female_applications', methods=['GET', 'POST'])
def export_female_applications():
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor()
    cursor.execute("SELECT applicant_id, first_name, middle_name, last_name, mobile_number, email, date_of_birth,"
                   "gender, age, caste, your_caste, marital_status, add_1, add_2, pincode, village, taluka, district,"
                   "state, city,  phd_registration_date, concerned_university, topic_of_phd, "
                   "name_of_guide, name_of_college, stream, board_university, admission_year, passing_year, percentage,"
                   "family_annual_income, income_certificate_number, issuing_authority, domicile, domicile_certificate,"
                   "domicile_number, caste_certf, issuing_district, caste_issuing_authority, salaried, disability,"
                   "father_name, mother_name, work_in_government, bank_name, account_number, ifsc_code," 
                   "account_holder_name FROM application_page WHERE gender ='female' ")
    data = cursor.fetchall()
    # Create a workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Add headers to the worksheet
    ws.append(['Applicant Id', 'First Name', 'Middle Name', 'Last Name', 'Mobile Number', 'Email', 'Date Of Birth', 
                'Gender', 'Age', 'Caste', 'Your Caste', 'Marital Status', 'Dependents', 'Add 1', 'Add 2', 'Pincode', 'Village', 
                'Taluka', 'District', 'State', 'Phd Registration Date', 'Concerned University', 'Topic Of Phd', 'Name Of Guide', 
                'Name Of College', 'Stream', 'Board University', 'Admission Year', 'Passing Year', 'Percentage', 'Family Annual Income', 
                'Income Certificate Number', 'Issuing Authority', 'Domicile', 'Domicile Certificate', 'Domicile Number', 'Caste Certf', 
                'Issuing District', 'Caste Issuing Authority', 'Salaried', 'Disability', 'Father Name', 'Mother Name', 'Work In Government', 
                'Bank Name', 'Account Number', 'Ifsc Code', 'Account Holder Name']	)

    # Add data to the worksheet
    for row in data:
        ws.append(row)

    # Save the workbook in memory as bytes
    data = BytesIO()
    wb.save(data)
    data.seek(0)

    # Create a response object and attach the workbook as a file
    response = make_response(data.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=Female_Applicants_Data.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    return response


def rejected_applications():    # ----- To count accepted applications  ------
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor()
    cursor.execute(" SELECT COUNT(*) FROM application_page WHERE final_approval='rejected' ")
    result = cursor.fetchone()
    print(result)
    return result[0]


def katkari_count_report():       # ----- To count male users  ------
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor()
    cursor.execute(" SELECT COUNT(*) FROM application_page WHERE your_caste='katkari' ")
    result = cursor.fetchone()
    print(result)
    return result[0]


def katkari_count21():       # ----- To count male users  ------
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor()
    cursor.execute(" SELECT COUNT(*) FROM application_page WHERE your_caste='katkari' and phd_registration_year='2021' ")
    result = cursor.fetchone()
    print(result)
    return result[0]


def katkari_count22():       # ----- To count male users  ------
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor()
    cursor.execute(" SELECT COUNT(*) FROM application_page WHERE your_caste='katkari' and phd_registration_year='2022' ")
    result = cursor.fetchone()
    print(result)
    return result[0]


def katkari_count23():       # ----- To count male users  ------
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor()
    cursor.execute(" SELECT COUNT(*) FROM application_page WHERE your_caste='katkari' and phd_registration_year='2023' ")
    result = cursor.fetchone()
    print(result)
    return result[0]


def kolam_count_report():     # ----- To count female users  ------
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor()
    cursor.execute(" SELECT COUNT(*) FROM application_page WHERE your_caste='kolam' ")
    result = cursor.fetchone()
    print(result)
    return result[0]


def kolam_count21():     # ----- To count female users  ------
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor()
    cursor.execute(" SELECT COUNT(*) FROM application_page WHERE your_caste='kolam' and phd_registration_year='2021' ")
    result = cursor.fetchone()
    print(result)
    return result[0]


def kolam_count22():     # ----- To count female users  ------
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor()
    cursor.execute(" SELECT COUNT(*) FROM application_page WHERE your_caste='kolam' and phd_registration_year='2022' ")
    result = cursor.fetchone()
    print(result)
    return result[0]


def kolam_count23():     # ----- To count female users  ------
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor()
    cursor.execute(" SELECT COUNT(*) FROM application_page WHERE your_caste='kolam' and phd_registration_year='2023' ")
    result = cursor.fetchone()
    print(result)
    return result[0]


def madia_count_report():       # ----- To count trans_gender users  ------
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor()
    cursor.execute(" SELECT COUNT(*) FROM application_page WHERE your_caste='madia' ")
    result = cursor.fetchone()
    print(result)
    return result[0]


def madia_count21():       # ----- To count trans_gender users  ------
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor()
    cursor.execute(" SELECT COUNT(*) FROM application_page WHERE your_caste='madia' and phd_registration_year='2021' ")
    result = cursor.fetchone()
    print(result)
    return result[0]


def madia_count22():       # ----- To count trans_gender users  ------
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor()
    cursor.execute(" SELECT COUNT(*) FROM application_page WHERE your_caste='madia' and phd_registration_year='2022' ")
    result = cursor.fetchone()
    print(result)
    return result[0]


def madia_count23():       # ----- To count trans_gender users  ------
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor()
    cursor.execute(" SELECT COUNT(*) FROM application_page WHERE your_caste='madia' and phd_registration_year='2023' ")
    result = cursor.fetchone()
    print(result)
    return result[0]


# ----------------------------------------------------------------------------
# ------------------- BEGIN APPLICATIONS 2023 REPORTS ------------------------------
# ----------------------------------------------------------------------------
@app.route('/total_application_report', methods=['GET', 'POST'])
def total_application_report():
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor(dictionary=True)
    cursor.execute(" SELECT * FROM application_page where phd_registration_year = 2023 ")
    result = cursor.fetchall()
    cnx.commit()
    cursor.close()
    cnx.close()
    # print(result)
    return render_template('total_application_report.html', result=result)


@app.route('/completed_form', methods=['GET', 'POST'])
def completed_form():
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor(dictionary=True)
    cursor.execute(" SELECT * FROM application_page where phd_registration_year = 2023 and form_filled='1' ")
    result = cursor.fetchall()
    cnx.commit()
    cursor.close()
    cnx.close()
    # print(result)
    return render_template('completed_form.html', result=result)


@app.route('/incompleted_form', methods=['GET', 'POST'])
def incompleted_form():
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor(dictionary=True)
    cursor.execute(" SELECT * FROM application_page where phd_registration_year = 2023 and form_filled='0' ")
    # cursor.execute(" SELECT * FROM application_page where email = 'tupotbhare@gmail.com' ")
    result = cursor.fetchall()
    cnx.commit()
    cursor.close()
    cnx.close()
    # print(result)
    return render_template('incompleted_form.html', result=result)


#----------------------  EXPORT TO EXCEL Total applications 2023 -------------------------------------------
@app.route('/export_total_applications_23', methods=['GET', 'POST'])
def export_total_applications_23():
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor()
    cursor.execute("  SELECT applicant_id, adhaar_number, first_name, last_name, middle_name, mobile_number, email, "
                   "date_of_birth, gender, age, caste, your_caste, subcaste, pvtg, pvtg_caste, marital_status," 
                    "dependents, state, district, taluka, village, city, add_1, add_2, pincode," 
                    "ssc_passing_year, ssc_percentage, ssc_school_name, ssc_stream, ssc_attempts, ssc_total," 
                    "hsc_passing_year, hsc_percentage, hsc_school_name, hsc_stream, hsc_attempts, hsc_total," 
                    "graduation_passing_year, graduation_percentage, graduation_school_name, grad_stream," 
                    "grad_attempts, grad_total, phd_passing_year, phd_percentage, phd_school_name, pg_stream, "
                    "pg_attempts, pg_total, have_you_qualified, name_of_college, other_college_name, "
                    "name_of_guide, topic_of_phd, concerned_university, department_name, faculty, "
                    "phd_registration_date, phd_registration_year, phd_registration_age, family_annual_income," 
                    "income_certificate_number, issuing_authority, income_issuing_district, income_issuing_taluka," 
                    "domicile, domicile_certificate, domicile_number, validity_certificate, validity_cert_number," 
                    "validity_issuing_district, validity_issuing_taluka, validity_issuing_authority, caste_certf, "
                    "caste_certf_number, issuing_district, caste_issuing_authority, salaried, disability, "
                    "type_of_disability, father_name, mother_name, work_in_government, gov_department," 
                    "gov_position, bank_name, account_number, ifsc_code, account_holder_name, micr FROM application_page"
                   " where phd_registration_year = '2023' ")

    data = cursor.fetchall()
    # Create a workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Add headers to the worksheet
    # ws.append(['applicant_id','email','first_name','last_name','application_date'])

    ws.append(['Applicant Id', 'Adhaar Card Number', 'First Name', 'Middle Name', 'Last Name', 'Mobile Number', 'Email',
               'Date Of Birth', 'Gender', 'Age', 'Caste/Tribe', 'Sub Caste', 'Are you PVTG', 'PVTG Caste/Tribe',
               'Marital Status', 'dependents', 'state', 'district', 'taluka', 'village', 'city', 'add_1', 'add_2',
               'pincode', 'SSC Passing Year', 'SSC Percentage', 'SSC School Name', 'SSC Stream', 'SSC Attempts',
                'SSC Total', 'HSC Passing Year', 'HSC Percentage', 'HSC School Name', 'HSC Stream', 'HSC Attempts', 'HSC Total',
                'Graduation Passing Year', 'Graduation Percentage', 'Graduation School Name', 'Graduation Stream', 'Graduation Attempts',
                'Graduation Total', 'PhD Passing Year', 'PhD Percentage', 'PhD School Name', 'PG Stream', 'PG Attempts', 'PG Total',
                'Have you Qualified', 'Name of College', 'Other College Name', 'Name of Guide', 'Topic of PhD',
                'Concerned University', 'Department Name', 'Faculty', 'PhD Registration Date', 'PhD Registration Year',
                'PhD Registration Age', 'Family Annual Income', 'Income Certificate Number', 'Issuing Authority',
                'Income Issuing District', 'Income Issuing Taluka', 'Domicile', 'Domicile Certificate', 'Domicile Number',
                'Validity Certificate', 'Validity Cert Number', 'Validity Issuing District', 'Validity Issuing Taluka',
                'Validity Issuing Authority', 'Caste Certificate', 'Caste Certf Number', 'Issuing District', 'Caste Issuing Authority',
                'Salaried', 'Disability', 'Type of Disability', 'Father Name', 'Mother Name', 'Work in Government', 'Government Department',
                'Government Position', 'Bank Name', 'Account Number', 'IFSC Code', 'Account Holder Name', 'MICR' ])

    # Add data to the worksheet
    for row in data:
        ws.append(row)

    # Save the workbook in memory as bytes
    data = BytesIO()
    wb.save(data)
    data.seek(0)

    # Create a response object and attach the workbook as a file
    response = make_response(data.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=Total_Applications_2023.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    return response


@app.route('/completed_form_export', methods=['GET', 'POST'])
def completed_form_export():
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor()
    cursor.execute("  SELECT applicant_id, adhaar_number, first_name, last_name, middle_name, mobile_number, email, "
                   "date_of_birth, gender, age, caste, your_caste, subcaste, pvtg, pvtg_caste, marital_status," 
                    "dependents, state, district, taluka, village, city, add_1, add_2, pincode," 
                    "ssc_passing_year, ssc_percentage, ssc_school_name, ssc_stream, ssc_attempts, ssc_total," 
                    "hsc_passing_year, hsc_percentage, hsc_school_name, hsc_stream, hsc_attempts, hsc_total," 
                    "graduation_passing_year, graduation_percentage, graduation_school_name, grad_stream," 
                    "grad_attempts, grad_total, phd_passing_year, phd_percentage, phd_school_name, pg_stream, "
                    "pg_attempts, pg_total, have_you_qualified, name_of_college, other_college_name, "
                    "name_of_guide, topic_of_phd, concerned_university, department_name, faculty, "
                    "phd_registration_date, phd_registration_year, phd_registration_age, family_annual_income," 
                    "income_certificate_number, issuing_authority, income_issuing_district, income_issuing_taluka," 
                    "domicile, domicile_certificate, domicile_number, validity_certificate, validity_cert_number," 
                    "validity_issuing_district, validity_issuing_taluka, validity_issuing_authority, caste_certf, "
                    "caste_certf_number, issuing_district, caste_issuing_authority, salaried, disability, "
                    "type_of_disability, father_name, mother_name, work_in_government, gov_department," 
                    "gov_position, bank_name, account_number, ifsc_code, account_holder_name, micr FROM application_page"
                   " where phd_registration_year = '2023' and form_filled='1' ")

    data = cursor.fetchall()
    # Create a workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Add headers to the worksheet
    # ws.append(['applicant_id','email','first_name','last_name','application_date'])

    ws.append(['Applicant Id', 'Adhaar Card Number', 'First Name', 'Middle Name', 'Last Name', 'Mobile Number', 'Email',
               'Date Of Birth', 'Gender', 'Age', 'Caste/Tribe', 'Your Caste', 'Sub Caste', 'Are you PVTG', 'PVTG Caste/Tribe',
               'Marital Status', 'dependents', 'state', 'district', 'taluka', 'village', 'city', 'add_1', 'add_2',
               'pincode', 'SSC Passing Year', 'SSC Percentage', 'SSC School Name', 'SSC Stream', 'SSC Attempts',
                'SSC Total', 'HSC Passing Year', 'HSC Percentage', 'HSC School Name', 'HSC Stream', 'HSC Attempts', 'HSC Total',
                'Graduation Passing Year', 'Graduation Percentage', 'Graduation School Name', 'Graduation Stream', 'Graduation Attempts',
                'Graduation Total', 'PhD Passing Year', 'PhD Percentage', 'PhD School Name', 'PG Stream', 'PG Attempts', 'PG Total',
                'Have you Qualified', 'Name of College', 'Other College Name', 'Name of Guide', 'Topic of PhD',
                'Concerned University', 'Department Name', 'Faculty', 'PhD Registration Date', 'PhD Registration Year',
                'PhD Registration Age', 'Family Annual Income', 'Income Certificate Number', 'Issuing Authority',
                'Income Issuing District', 'Income Issuing Taluka', 'Domicile', 'Domicile Certificate', 'Domicile Number',
                'Validity Certificate', 'Validity Cert Number', 'Validity Issuing District', 'Validity Issuing Taluka',
                'Validity Issuing Authority', 'Caste Certificate', 'Caste Certf Number', 'Issuing District', 'Caste Issuing Authority',
                'Salaried', 'Disability', 'Type of Disability', 'Father Name', 'Mother Name', 'Work in Government', 'Government Department',
                'Government Position', 'Bank Name', 'Account Number', 'IFSC Code', 'Account Holder Name', 'MICR' ])

    # Add data to the worksheet
    for row in data:
        ws.append(row)

    # Save the workbook in memory as bytes
    data = BytesIO()
    wb.save(data)
    data.seek(0)

    # Create a response object and attach the workbook as a file
    response = make_response(data.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=Completed Form Fellowship 2023.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    return response


@app.route('/incompleted_form_export', methods=['GET', 'POST'])
def incompleted_form_export():
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor()
    cursor.execute("  SELECT applicant_id, adhaar_number, first_name, last_name, middle_name, mobile_number, email, "
                   "date_of_birth, gender, age, caste, your_caste, subcaste, pvtg, pvtg_caste, marital_status," 
                    "dependents, state, district, taluka, village, city, add_1, add_2, pincode," 
                    "ssc_passing_year, ssc_percentage, ssc_school_name, ssc_stream, ssc_attempts, ssc_total," 
                    "hsc_passing_year, hsc_percentage, hsc_school_name, hsc_stream, hsc_attempts, hsc_total," 
                    "graduation_passing_year, graduation_percentage, graduation_school_name, grad_stream," 
                    "grad_attempts, grad_total, phd_passing_year, phd_percentage, phd_school_name, pg_stream, "
                    "pg_attempts, pg_total, have_you_qualified, name_of_college, other_college_name, "
                    "name_of_guide, topic_of_phd, concerned_university, department_name, faculty, "
                    "phd_registration_date, phd_registration_year, phd_registration_age, family_annual_income," 
                    "income_certificate_number, issuing_authority, income_issuing_district, income_issuing_taluka," 
                    "domicile, domicile_certificate, domicile_number, validity_certificate, validity_cert_number," 
                    "validity_issuing_district, validity_issuing_taluka, validity_issuing_authority, caste_certf, "
                    "caste_certf_number, issuing_district, caste_issuing_authority, salaried, disability, "
                    "type_of_disability, father_name, mother_name, work_in_government, gov_department," 
                    "gov_position, bank_name, account_number, ifsc_code, account_holder_name, micr FROM application_page"
                   " where phd_registration_year = '2023' and form_filled='0' ")

    data = cursor.fetchall()
    # Create a workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Add headers to the worksheet
    # ws.append(['applicant_id','email','first_name','last_name','application_date'])

    ws.append(['Applicant Id', 'Adhaar Card Number', 'First Name', 'Middle Name', 'Last Name', 'Mobile Number', 'Email',
               'Date Of Birth', 'Gender', 'Age', 'Caste/Tribe', 'Sub Caste', 'Are you PVTG', 'PVTG Caste/Tribe',
               'Marital Status', 'dependents', 'state', 'district', 'taluka', 'village', 'city', 'add_1', 'add_2',
               'pincode', 'SSC Passing Year', 'SSC Percentage', 'SSC School Name', 'SSC Stream', 'SSC Attempts',
                'SSC Total', 'HSC Passing Year', 'HSC Percentage', 'HSC School Name', 'HSC Stream', 'HSC Attempts', 'HSC Total',
                'Graduation Passing Year', 'Graduation Percentage', 'Graduation School Name', 'Graduation Stream', 'Graduation Attempts',
                'Graduation Total', 'PhD Passing Year', 'PhD Percentage', 'PhD School Name', 'PG Stream', 'PG Attempts', 'PG Total',
                'Have you Qualified', 'Name of College', 'Other College Name', 'Name of Guide', 'Topic of PhD',
                'Concerned University', 'Department Name', 'Faculty', 'PhD Registration Date', 'PhD Registration Year',
                'PhD Registration Age', 'Family Annual Income', 'Income Certificate Number', 'Issuing Authority',
                'Income Issuing District', 'Income Issuing Taluka', 'Domicile', 'Domicile Certificate', 'Domicile Number',
                'Validity Certificate', 'Validity Cert Number', 'Validity Issuing District', 'Validity Issuing Taluka',
                'Validity Issuing Authority', 'Caste Certificate', 'Caste Certf Number', 'Issuing District', 'Caste Issuing Authority',
                'Salaried', 'Disability', 'Type of Disability', 'Father Name', 'Mother Name', 'Work in Government', 'Government Department',
                'Government Position', 'Bank Name', 'Account Number', 'IFSC Code', 'Account Holder Name', 'MICR' ])

    # Add data to the worksheet
    for row in data:
        ws.append(row)

    # Save the workbook in memory as bytes
    data = BytesIO()
    wb.save(data)
    data.seek(0)

    # Create a response object and attach the workbook as a file
    response = make_response(data.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=Total_Applications_2023.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    return response


@app.route('/total_accepted_report', methods=['GET', 'POST'])
def total_accepted_report():
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor(dictionary=True)
    cursor.execute(" SELECT * FROM application_page WHERE final_approval='accepted' and phd_registration_year='2023' ")
    result = cursor.fetchall()
    cnx.commit()
    cursor.close()
    cnx.close()
    # print(result)
    return render_template('total_accepted_report.html', result=result)


#----------------------  EXPORT TO EXCEL Accepted Applications 2023 -------------------------------------------
@app.route('/export_accepted_applications_23', methods=['GET', 'POST'])
def export_accepted_applications_23():
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor()
    cursor.execute("  SELECT applicant_id, adhaar_number, first_name, last_name, middle_name, mobile_number, email, "
                   "date_of_birth, gender, age, caste, your_caste, subcaste, pvtg, pvtg_caste, marital_status," 
                    "dependents, state, district, taluka, village, city, add_1, add_2, pincode," 
                    "ssc_passing_year, ssc_percentage, ssc_school_name, ssc_stream, ssc_attempts, ssc_total," 
                    "hsc_passing_year, hsc_percentage, hsc_school_name, hsc_stream, hsc_attempts, hsc_total," 
                    "graduation_passing_year, graduation_percentage, graduation_school_name, grad_stream," 
                    "grad_attempts, grad_total, phd_passing_year, phd_percentage, phd_school_name, pg_stream, "
                    "pg_attempts, pg_total, have_you_qualified, name_of_college, other_college_name, "
                    "name_of_guide, topic_of_phd, concerned_university, department_name, faculty, "
                    "phd_registration_date, phd_registration_year, phd_registration_age, family_annual_income," 
                    "income_certificate_number, issuing_authority, income_issuing_district, income_issuing_taluka," 
                    "domicile, domicile_certificate, domicile_number, validity_certificate, validity_cert_number," 
                    "validity_issuing_district, validity_issuing_taluka, validity_issuing_authority, caste_certf, "
                    "caste_certf_number, issuing_district, caste_issuing_authority, salaried, disability, "
                    "type_of_disability, father_name, mother_name, work_in_government, gov_department," 
                    "gov_position, bank_name, account_number, ifsc_code, account_holder_name, micr FROM application_page "
                    " WHERE final_approval = 'accepted' and phd_registration_year='2023' ")
    data = cursor.fetchall()
    # Create a workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Add headers to the worksheet
    ws.append(['Applicant Id', 'Adhaar Card Number', 'First Name', 'Middle Name', 'Last Name', 'Mobile Number', 'Email',
               'Date Of Birth', 'Gender', 'Age', 'Caste/Tribe', 'Sub Caste', 'Are you PVTG', 'PVTG Caste/Tribe',
               'Marital Status', 'dependents', 'state', 'district', 'taluka', 'village', 'city', 'add_1', 'add_2',
               'pincode', 'SSC Passing Year', 'SSC Percentage', 'SSC School Name', 'SSC Stream', 'SSC Attempts',
                'SSC Total', 'HSC Passing Year', 'HSC Percentage', 'HSC School Name', 'HSC Stream', 'HSC Attempts', 'HSC Total',
                'Graduation Passing Year', 'Graduation Percentage', 'Graduation School Name', 'Graduation Stream', 'Graduation Attempts',
                'Graduation Total', 'PhD Passing Year', 'PhD Percentage', 'PhD School Name', 'PG Stream', 'PG Attempts', 'PG Total',
                'Have you Qualified', 'Name of College', 'Other College Name', 'Name of Guide', 'Topic of PhD',
                'Concerned University', 'Department Name', 'Faculty', 'PhD Registration Date', 'PhD Registration Year',
                'PhD Registration Age', 'Family Annual Income', 'Income Certificate Number', 'Issuing Authority',
                'Income Issuing District', 'Income Issuing Taluka', 'Domicile', 'Domicile Certificate', 'Domicile Number',
                'Validity Certificate', 'Validity Cert Number', 'Validity Issuing District', 'Validity Issuing Taluka',
                'Validity Issuing Authority', 'Caste Certificate', 'Caste Certf Number', 'Issuing District', 'Caste Issuing Authority',
                'Salaried', 'Disability', 'Type of Disability', 'Father Name', 'Mother Name', 'Work in Government', 'Government Department',
                'Government Position', 'Bank Name', 'Account Number', 'IFSC Code', 'Account Holder Name', 'MICR' ])

    # Add data to the worksheet
    for row in data:
        ws.append(row)

    # Save the workbook in memory as bytes
    data = BytesIO()
    wb.save(data)
    data.seek(0)

    # Create a response object and attach the workbook as a file
    response = make_response(data.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=Accepted_Applications_2023.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    return response    


@app.route('/total_rejected_report', methods=['GET', 'POST'])
def total_rejected_report():
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor(dictionary=True)
    cursor.execute("SELECT * FROM application_page WHERE final_approval='rejected' and phd_registration_year='2023'")
    result = cursor.fetchall()
    cnx.commit()
    cursor.close()
    cnx.close()
    # print(result)
    return render_template('total_rejected_report.html', result=result)


#----------------------  EXPORT TO EXCEL Rejected Applications 2023 -------------------------------------------
@app.route('/export_rejected_applications_23', methods=['GET', 'POST'])
def export_rejected_applications_23():
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor()
    cursor.execute("  SELECT applicant_id, adhaar_number, first_name, last_name, middle_name, mobile_number, email, "
                   "date_of_birth, gender, age, caste, your_caste, subcaste, pvtg, pvtg_caste, marital_status," 
                    "dependents, state, district, taluka, village, city, add_1, add_2, pincode," 
                    "ssc_passing_year, ssc_percentage, ssc_school_name, ssc_stream, ssc_attempts, ssc_total," 
                    "hsc_passing_year, hsc_percentage, hsc_school_name, hsc_stream, hsc_attempts, hsc_total," 
                    "graduation_passing_year, graduation_percentage, graduation_school_name, grad_stream," 
                    "grad_attempts, grad_total, phd_passing_year, phd_percentage, phd_school_name, pg_stream, "
                    "pg_attempts, pg_total, have_you_qualified, name_of_college, other_college_name, "
                    "name_of_guide, topic_of_phd, concerned_university, department_name, faculty, "
                    "phd_registration_date, phd_registration_year, phd_registration_age, family_annual_income," 
                    "income_certificate_number, issuing_authority, income_issuing_district, income_issuing_taluka," 
                    "domicile, domicile_certificate, domicile_number, validity_certificate, validity_cert_number," 
                    "validity_issuing_district, validity_issuing_taluka, validity_issuing_authority, caste_certf, "
                    "caste_certf_number, issuing_district, caste_issuing_authority, salaried, disability, "
                    "type_of_disability, father_name, mother_name, work_in_government, gov_department," 
                    "gov_position, bank_name, account_number, ifsc_code, account_holder_name, micr FROM application_page "
                    " WHERE final_approval = 'rejected' and phd_registration_year='2023' ")
    data = cursor.fetchall()
    # Create a workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Add headers to the worksheet
    ws.append(['Applicant Id', 'Adhaar Card Number', 'First Name', 'Middle Name', 'Last Name', 'Mobile Number', 'Email',
               'Date Of Birth', 'Gender', 'Age', 'Caste/Tribe', 'Sub Caste', 'Are you PVTG', 'PVTG Caste/Tribe',
               'Marital Status', 'dependents', 'state', 'district', 'taluka', 'village', 'city', 'add_1', 'add_2',
               'pincode', 'SSC Passing Year', 'SSC Percentage', 'SSC School Name', 'SSC Stream', 'SSC Attempts',
                'SSC Total', 'HSC Passing Year', 'HSC Percentage', 'HSC School Name', 'HSC Stream', 'HSC Attempts', 'HSC Total',
                'Graduation Passing Year', 'Graduation Percentage', 'Graduation School Name', 'Graduation Stream', 'Graduation Attempts',
                'Graduation Total', 'PhD Passing Year', 'PhD Percentage', 'PhD School Name', 'PG Stream', 'PG Attempts', 'PG Total',
                'Have you Qualified', 'Name of College', 'Other College Name', 'Name of Guide', 'Topic of PhD',
                'Concerned University', 'Department Name', 'Faculty', 'PhD Registration Date', 'PhD Registration Year',
                'PhD Registration Age', 'Family Annual Income', 'Income Certificate Number', 'Issuing Authority',
                'Income Issuing District', 'Income Issuing Taluka', 'Domicile', 'Domicile Certificate', 'Domicile Number',
                'Validity Certificate', 'Validity Cert Number', 'Validity Issuing District', 'Validity Issuing Taluka',
                'Validity Issuing Authority', 'Caste Certificate', 'Caste Certf Number', 'Issuing District', 'Caste Issuing Authority',
                'Salaried', 'Disability', 'Type of Disability', 'Father Name', 'Mother Name', 'Work in Government', 'Government Department',
                'Government Position', 'Bank Name', 'Account Number', 'IFSC Code', 'Account Holder Name', 'MICR' ])

    # Add data to the worksheet
    for row in data:
        ws.append(row)

    # Save the workbook in memory as bytes
    data = BytesIO()
    wb.save(data)
    data.seek(0)

    # Create a response object and attach the workbook as a file
    response = make_response(data.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=Rejected_Applications_2023.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    return response    

# ----------------------------------------------------------------------------
# ------------------- END APPLICATIONS 2023 REPORTS ------------------------------
# ----------------------------------------------------------------------------

# ----------------------------------------------------------------------------
# ------------------- BEGIN APPLICATIONS 2022 REPORTS ------------------------------
# ----------------------------------------------------------------------------
@app.route('/total_application_report_22', methods=['GET', 'POST'])
def total_application_report_22():
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor(dictionary=True)
    cursor.execute("SELECT * FROM application_page WHERE phd_registration_year='2022' ")
    result = cursor.fetchall()
    cnx.commit()
    cursor.close()
    cnx.close()
    # print(result)
    return render_template('total_application_report_22.html', result=result)


#----------------------  EXPORT TO EXCEL Total applications 2022 -------------------------------------------
@app.route('/export_total_applications_22', methods=['GET', 'POST'])
def export_total_applications_22():
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor()
    cursor.execute("  SELECT applicant_id, adhaar_number, first_name, last_name, middle_name, mobile_number, email, "
                   "date_of_birth, gender, age, caste, your_caste, subcaste, pvtg, pvtg_caste, marital_status," 
                    "dependents, state, district, taluka, village, city, add_1, add_2, pincode," 
                    "ssc_passing_year, ssc_percentage, ssc_school_name, ssc_stream, ssc_attempts, ssc_total," 
                    "hsc_passing_year, hsc_percentage, hsc_school_name, hsc_stream, hsc_attempts, hsc_total," 
                    "graduation_passing_year, graduation_percentage, graduation_school_name, grad_stream," 
                    "grad_attempts, grad_total, phd_passing_year, phd_percentage, phd_school_name, pg_stream, "
                    "pg_attempts, pg_total, have_you_qualified, name_of_college, other_college_name, "
                    "name_of_guide, topic_of_phd, concerned_university, department_name, faculty, "
                    "phd_registration_date, phd_registration_year, phd_registration_age, family_annual_income," 
                    "income_certificate_number, issuing_authority, income_issuing_district, income_issuing_taluka," 
                    "domicile, domicile_certificate, domicile_number, validity_certificate, validity_cert_number," 
                    "validity_issuing_district, validity_issuing_taluka, validity_issuing_authority, caste_certf, "
                    "caste_certf_number, issuing_district, caste_issuing_authority, salaried, disability, "
                    "type_of_disability, father_name, mother_name, work_in_government, gov_department," 
                    "gov_position, bank_name, account_number, ifsc_code, account_holder_name, micr FROM application_page"
                   " WHERE phd_registration_year='2022' ")
    data = cursor.fetchall()
    # Create a workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Add headers to the worksheet
    ws.append(['Applicant Id', 'Adhaar Card Number', 'First Name', 'Middle Name', 'Last Name', 'Mobile Number', 'Email',
               'Date Of Birth', 'Gender', 'Age', 'Caste/Tribe','Your Caste', 'Sub Caste', 'Are you PVTG', 'PVTG Caste/Tribe',
               'Marital Status', 'dependents', 'state', 'district', 'taluka', 'village', 'city', 'add_1', 'add_2',
               'pincode', 'SSC Passing Year', 'SSC Percentage', 'SSC School Name', 'SSC Stream', 'SSC Attempts',
                'SSC Total', 'HSC Passing Year', 'HSC Percentage', 'HSC School Name', 'HSC Stream', 'HSC Attempts', 'HSC Total',
                'Graduation Passing Year', 'Graduation Percentage', 'Graduation School Name', 'Graduation Stream', 'Graduation Attempts',
                'Graduation Total', 'PhD Passing Year', 'PhD Percentage', 'PhD School Name', 'PG Stream', 'PG Attempts', 'PG Total',
                'Have you Qualified', 'Name of College', 'Other College Name', 'Name of Guide', 'Topic of PhD',
                'Concerned University', 'Department Name', 'Faculty', 'PhD Registration Date', 'PhD Registration Year',
                'PhD Registration Age', 'Family Annual Income', 'Income Certificate Number', 'Issuing Authority',
                'Income Issuing District', 'Income Issuing Taluka', 'Domicile', 'Domicile Certificate', 'Domicile Number',
                'Validity Certificate', 'Validity Cert Number', 'Validity Issuing District', 'Validity Issuing Taluka',
                'Validity Issuing Authority', 'Caste Certificate', 'Caste Certf Number', 'Issuing District', 'Caste Issuing Authority',
                'Salaried', 'Disability', 'Type of Disability', 'Father Name', 'Mother Name', 'Work in Government', 'Government Department',
                'Government Position', 'Bank Name', 'Account Number', 'IFSC Code', 'Account Holder Name', 'MICR' ])

    # Add data to the worksheet
    for row in data:
        ws.append(row)

    # Save the workbook in memory as bytes
    data = BytesIO()
    wb.save(data)
    data.seek(0)

    # Create a response object and attach the workbook as a file
    response = make_response(data.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=Total_Applications_2022.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    return response    


@app.route('/total_accepted_report_22', methods=['GET', 'POST'])
def total_accepted_report_22():
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor(dictionary=True)
    cursor.execute("SELECT * FROM application_page WHERE final_approval='accepted' and phd_registration_year='2022' ")
    result = cursor.fetchall()
    cnx.commit()
    cursor.close()
    cnx.close()
    # print(result)
    return render_template('total_accepted_report_22.html', result=result)


#----------------------  EXPORT TO EXCEL Accepted Applications 2022 -------------------------------------------
@app.route('/export_accepted_applications_22', methods=['GET', 'POST'])
def export_accepted_applications_22():
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor()
    cursor.execute("  SELECT applicant_id, adhaar_number, first_name, last_name, middle_name, mobile_number, email, "
                   "date_of_birth, gender, age, caste, your_caste, subcaste, pvtg, pvtg_caste, marital_status," 
                    "dependents, state, district, taluka, village, city, add_1, add_2, pincode," 
                    "ssc_passing_year, ssc_percentage, ssc_school_name, ssc_stream, ssc_attempts, ssc_total," 
                    "hsc_passing_year, hsc_percentage, hsc_school_name, hsc_stream, hsc_attempts, hsc_total," 
                    "graduation_passing_year, graduation_percentage, graduation_school_name, grad_stream," 
                    "grad_attempts, grad_total, phd_passing_year, phd_percentage, phd_school_name, pg_stream, "
                    "pg_attempts, pg_total, have_you_qualified, name_of_college, other_college_name, "
                    "name_of_guide, topic_of_phd, concerned_university, department_name, faculty, "
                    "phd_registration_date, phd_registration_year, phd_registration_age, family_annual_income," 
                    "income_certificate_number, issuing_authority, income_issuing_district, income_issuing_taluka," 
                    "domicile, domicile_certificate, domicile_number, validity_certificate, validity_cert_number," 
                    "validity_issuing_district, validity_issuing_taluka, validity_issuing_authority, caste_certf, "
                    "caste_certf_number, issuing_district, caste_issuing_authority, salaried, disability, "
                    "type_of_disability, father_name, mother_name, work_in_government, gov_department," 
                    "gov_position, bank_name, account_number, ifsc_code, account_holder_name, micr FROM application_page"
                   " WHERE phd_registration_year='2022' and final_approval='accepted' ")
    data = cursor.fetchall()
    # Create a workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Add headers to the worksheet
    ws.append(['Applicant Id', 'Adhaar Card Number', 'First Name', 'Middle Name', 'Last Name', 'Mobile Number', 'Email',
               'Date Of Birth', 'Gender', 'Age', 'Caste/Tribe', 'Your Caste', 'Sub Caste', 'Are you PVTG', 'PVTG Caste/Tribe',
               'Marital Status', 'dependents', 'state', 'district', 'taluka', 'village', 'city', 'add_1', 'add_2',
               'pincode', 'SSC Passing Year', 'SSC Percentage', 'SSC School Name', 'SSC Stream', 'SSC Attempts',
                'SSC Total', 'HSC Passing Year', 'HSC Percentage', 'HSC School Name', 'HSC Stream', 'HSC Attempts', 'HSC Total',
                'Graduation Passing Year', 'Graduation Percentage', 'Graduation School Name', 'Graduation Stream', 'Graduation Attempts',
                'Graduation Total', 'PhD Passing Year', 'PhD Percentage', 'PhD School Name', 'PG Stream', 'PG Attempts', 'PG Total',
                'Have you Qualified', 'Name of College', 'Other College Name', 'Name of Guide', 'Topic of PhD',
                'Concerned University', 'Department Name', 'Faculty', 'PhD Registration Date', 'PhD Registration Year',
                'PhD Registration Age', 'Family Annual Income', 'Income Certificate Number', 'Issuing Authority',
                'Income Issuing District', 'Income Issuing Taluka', 'Domicile', 'Domicile Certificate', 'Domicile Number',
                'Validity Certificate', 'Validity Cert Number', 'Validity Issuing District', 'Validity Issuing Taluka',
                'Validity Issuing Authority', 'Caste Certificate', 'Caste Certf Number', 'Issuing District', 'Caste Issuing Authority',
                'Salaried', 'Disability', 'Type of Disability', 'Father Name', 'Mother Name', 'Work in Government', 'Government Department',
                'Government Position', 'Bank Name', 'Account Number', 'IFSC Code', 'Account Holder Name', 'MICR' ])

    # Add data to the worksheet
    for row in data:
        ws.append(row)

    # Save the workbook in memory as bytes
    data = BytesIO()
    wb.save(data)
    data.seek(0)

    # Create a response object and attach the workbook as a file
    response = make_response(data.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=Accepted_Applications_2022.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    return response     


@app.route('/total_rejected_report_22', methods=['GET', 'POST'])
def total_rejected_report_22():
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor(dictionary=True)
    cursor.execute("SELECT * FROM application_page WHERE phd_registration_year='2022' and final_approval='rejected' ")
    result = cursor.fetchall()
    cnx.commit()
    cursor.close()
    cnx.close()
    # print(result)
    return render_template('total_rejected_report_22.html', result=result)


#----------------------  EXPORT TO EXCEL Rejected Applications 2022 -------------------------------------------
@app.route('/export_rejected_applications_22', methods=['GET', 'POST'])
def export_rejected_applications_22():
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor()
    cursor.execute("  SELECT applicant_id, adhaar_number, first_name, last_name, middle_name, mobile_number, email, "
                   "date_of_birth, gender, age, caste, your_caste, subcaste, pvtg, pvtg_caste, marital_status," 
                    "dependents, state, district, taluka, village, city, add_1, add_2, pincode," 
                    "ssc_passing_year, ssc_percentage, ssc_school_name, ssc_stream, ssc_attempts, ssc_total," 
                    "hsc_passing_year, hsc_percentage, hsc_school_name, hsc_stream, hsc_attempts, hsc_total," 
                    "graduation_passing_year, graduation_percentage, graduation_school_name, grad_stream," 
                    "grad_attempts, grad_total, phd_passing_year, phd_percentage, phd_school_name, pg_stream, "
                    "pg_attempts, pg_total, have_you_qualified, name_of_college, other_college_name, "
                    "name_of_guide, topic_of_phd, concerned_university, department_name, faculty, "
                    "phd_registration_date, phd_registration_year, phd_registration_age, family_annual_income," 
                    "income_certificate_number, issuing_authority, income_issuing_district, income_issuing_taluka," 
                    "domicile, domicile_certificate, domicile_number, validity_certificate, validity_cert_number," 
                    "validity_issuing_district, validity_issuing_taluka, validity_issuing_authority, caste_certf, "
                    "caste_certf_number, issuing_district, caste_issuing_authority, salaried, disability, "
                    "type_of_disability, father_name, mother_name, work_in_government, gov_department," 
                    "gov_position, bank_name, account_number, ifsc_code, account_holder_name, micr FROM application_page"
                   " WHERE phd_registration_year='2022' and final_approval='rejected' ")
    data = cursor.fetchall()
    # Create a workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Add headers to the worksheet
    ws.append(['Applicant Id', 'Adhaar Card Number', 'First Name', 'Middle Name', 'Last Name', 'Mobile Number', 'Email',
               'Date Of Birth', 'Gender', 'Age', 'Caste/Tribe', 'Your Caste', 'Sub Caste', 'Are you PVTG', 'PVTG Caste/Tribe',
               'Marital Status', 'dependents', 'state', 'district', 'taluka', 'village', 'city', 'add_1', 'add_2',
               'pincode', 'SSC Passing Year', 'SSC Percentage', 'SSC School Name', 'SSC Stream', 'SSC Attempts',
                'SSC Total', 'HSC Passing Year', 'HSC Percentage', 'HSC School Name', 'HSC Stream', 'HSC Attempts', 'HSC Total',
                'Graduation Passing Year', 'Graduation Percentage', 'Graduation School Name', 'Graduation Stream', 'Graduation Attempts',
                'Graduation Total', 'PhD Passing Year', 'PhD Percentage', 'PhD School Name', 'PG Stream', 'PG Attempts', 'PG Total',
                'Have you Qualified', 'Name of College', 'Other College Name', 'Name of Guide', 'Topic of PhD',
                'Concerned University', 'Department Name', 'Faculty', 'PhD Registration Date', 'PhD Registration Year',
                'PhD Registration Age', 'Family Annual Income', 'Income Certificate Number', 'Issuing Authority',
                'Income Issuing District', 'Income Issuing Taluka', 'Domicile', 'Domicile Certificate', 'Domicile Number',
                'Validity Certificate', 'Validity Cert Number', 'Validity Issuing District', 'Validity Issuing Taluka',
                'Validity Issuing Authority', 'Caste Certificate', 'Caste Certf Number', 'Issuing District', 'Caste Issuing Authority',
                'Salaried', 'Disability', 'Type of Disability', 'Father Name', 'Mother Name', 'Work in Government', 'Government Department',
                'Government Position', 'Bank Name', 'Account Number', 'IFSC Code', 'Account Holder Name', 'MICR' ])

    # Add data to the worksheet
    for row in data:
        ws.append(row)

    # Save the workbook in memory as bytes
    data = BytesIO()
    wb.save(data)
    data.seek(0)

    # Create a response object and attach the workbook as a file
    response = make_response(data.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=Rejected_Applications_2022.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    return response    
         

# ----------------------------------------------------------------------------
# ------------------- END APPLICATIONS 2022 REPORTS ------------------------------
# ----------------------------------------------------------------------------

# ----------------------------------------------------------------------------
# ------------------- BEGIN APPLICATIONS 2021 REPORTS ------------------------------
# ----------------------------------------------------------------------------
@app.route('/total_application_report_21', methods=['GET', 'POST'])
def total_application_report_21():
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication') 
    cursor = cnx.cursor(dictionary=True)
    cursor.execute("SELECT * FROM application_page WHERE phd_registration_year='2021'")
    result = cursor.fetchall()
    cnx.commit()
    cursor.close()
    cnx.close()
    # print(result)
    return render_template('total_application_report_21.html', result=result)


#----------------------  EXPORT TO EXCEL Total applications 2021 ----------------
@app.route('/export_total_applications_21', methods=['GET', 'POST'])
def export_total_applications_21():
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor()
    cursor.execute("  SELECT applicant_id, adhaar_number, first_name, last_name, middle_name, mobile_number, email, "
                   "date_of_birth, gender, age, caste, your_caste, subcaste, pvtg, pvtg_caste, marital_status," 
                    "dependents, state, district, taluka, village, city, add_1, add_2, pincode," 
                    "ssc_passing_year, ssc_percentage, ssc_school_name, ssc_stream, ssc_attempts, ssc_total," 
                    "hsc_passing_year, hsc_percentage, hsc_school_name, hsc_stream, hsc_attempts, hsc_total," 
                    "graduation_passing_year, graduation_percentage, graduation_school_name, grad_stream," 
                    "grad_attempts, grad_total, phd_passing_year, phd_percentage, phd_school_name, pg_stream, "
                    "pg_attempts, pg_total, have_you_qualified, name_of_college, other_college_name, "
                    "name_of_guide, topic_of_phd, concerned_university, department_name, faculty, "
                    "phd_registration_date, phd_registration_year, phd_registration_age, family_annual_income," 
                    "income_certificate_number, issuing_authority, income_issuing_district, income_issuing_taluka," 
                    "domicile, domicile_certificate, domicile_number, validity_certificate, validity_cert_number," 
                    "validity_issuing_district, validity_issuing_taluka, validity_issuing_authority, caste_certf, "
                    "caste_certf_number, issuing_district, caste_issuing_authority, salaried, disability, "
                    "type_of_disability, father_name, mother_name, work_in_government, gov_department," 
                    "gov_position, bank_name, account_number, ifsc_code, account_holder_name, micr FROM application_page"
                   " WHERE phd_registration_year='2021' ")
    data = cursor.fetchall()
    # Create a workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Add headers to the worksheet
    ws.append(['Applicant Id', 'Adhaar Card Number', 'First Name', 'Middle Name', 'Last Name', 'Mobile Number', 'Email',
               'Date Of Birth', 'Gender', 'Age', 'Caste/Tribe', 'Your Caste', 'Sub Caste', 'Are you PVTG', 'PVTG Caste/Tribe',
               'Marital Status', 'dependents', 'state', 'district', 'taluka', 'village', 'city', 'add_1', 'add_2',
               'pincode', 'SSC Passing Year', 'SSC Percentage', 'SSC School Name', 'SSC Stream', 'SSC Attempts',
                'SSC Total', 'HSC Passing Year', 'HSC Percentage', 'HSC School Name', 'HSC Stream', 'HSC Attempts', 'HSC Total',
                'Graduation Passing Year', 'Graduation Percentage', 'Graduation School Name', 'Graduation Stream', 'Graduation Attempts',
                'Graduation Total', 'PhD Passing Year', 'PhD Percentage', 'PhD School Name', 'PG Stream', 'PG Attempts', 'PG Total',
                'Have you Qualified', 'Name of College', 'Other College Name', 'Name of Guide', 'Topic of PhD',
                'Concerned University', 'Department Name', 'Faculty', 'PhD Registration Date', 'PhD Registration Year',
                'PhD Registration Age', 'Family Annual Income', 'Income Certificate Number', 'Issuing Authority',
                'Income Issuing District', 'Income Issuing Taluka', 'Domicile', 'Domicile Certificate', 'Domicile Number',
                'Validity Certificate', 'Validity Cert Number', 'Validity Issuing District', 'Validity Issuing Taluka',
                'Validity Issuing Authority', 'Caste Certificate', 'Caste Certf Number', 'Issuing District', 'Caste Issuing Authority',
                'Salaried', 'Disability', 'Type of Disability', 'Father Name', 'Mother Name', 'Work in Government', 'Government Department',
                'Government Position', 'Bank Name', 'Account Number', 'IFSC Code', 'Account Holder Name', 'MICR' ])

    # Add data to the worksheet
    for row in data:
        ws.append(row)

    # Save the workbook in memory as bytes
    data = BytesIO()
    wb.save(data)
    data.seek(0)

    # Create a response object and attach the workbook as a file
    response = make_response(data.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=Total_Applications_2021.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    return response     


@app.route('/total_accepted_report_21', methods=['GET', 'POST'])
def total_accepted_report_21():
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor(dictionary=True)
    cursor.execute("SELECT * FROM application_page WHERE phd_registration_year='2021' and final_approval='accepted' ")
    result = cursor.fetchall()
    cnx.commit()
    cursor.close()
    cnx.close()
    # print(result)
    return render_template('total_accepted_report_21.html', result=result)


#----------------------  EXPORT TO EXCEL Accepted Applications 2021 -------------------------------------------
@app.route('/export_accepted_applications_21', methods=['GET', 'POST'])
def export_accepted_applications_21():
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor()
    cursor.execute("  SELECT applicant_id, adhaar_number, first_name, last_name, middle_name, mobile_number, email, "
                   "date_of_birth, gender, age, caste, your_caste, subcaste, pvtg, pvtg_caste, marital_status," 
                    "dependents, state, district, taluka, village, city, add_1, add_2, pincode," 
                    "ssc_passing_year, ssc_percentage, ssc_school_name, ssc_stream, ssc_attempts, ssc_total," 
                    "hsc_passing_year, hsc_percentage, hsc_school_name, hsc_stream, hsc_attempts, hsc_total," 
                    "graduation_passing_year, graduation_percentage, graduation_school_name, grad_stream," 
                    "grad_attempts, grad_total, phd_passing_year, phd_percentage, phd_school_name, pg_stream, "
                    "pg_attempts, pg_total, have_you_qualified, name_of_college, other_college_name, "
                    "name_of_guide, topic_of_phd, concerned_university, department_name, faculty, "
                    "phd_registration_date, phd_registration_year, phd_registration_age, family_annual_income," 
                    "income_certificate_number, issuing_authority, income_issuing_district, income_issuing_taluka," 
                    "domicile, domicile_certificate, domicile_number, validity_certificate, validity_cert_number," 
                    "validity_issuing_district, validity_issuing_taluka, validity_issuing_authority, caste_certf, "
                    "caste_certf_number, issuing_district, caste_issuing_authority, salaried, disability, "
                    "type_of_disability, father_name, mother_name, work_in_government, gov_department," 
                    "gov_position, bank_name, account_number, ifsc_code, account_holder_name, micr FROM application_page"
                   " WHERE phd_registration_year='2021' and final_approval='accepted' ")
    data = cursor.fetchall()
    # Create a workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Add headers to the worksheet
    ws.append(['Applicant Id', 'Adhaar Card Number', 'First Name', 'Middle Name', 'Last Name', 'Mobile Number', 'Email',
               'Date Of Birth', 'Gender', 'Age', 'Caste/Tribe', 'Your Caste', 'Sub Caste', 'Are you PVTG', 'PVTG Caste/Tribe',
               'Marital Status', 'dependents', 'state', 'district', 'taluka', 'village', 'city', 'add_1', 'add_2',
               'pincode', 'SSC Passing Year', 'SSC Percentage', 'SSC School Name', 'SSC Stream', 'SSC Attempts',
                'SSC Total', 'HSC Passing Year', 'HSC Percentage', 'HSC School Name', 'HSC Stream', 'HSC Attempts', 'HSC Total',
                'Graduation Passing Year', 'Graduation Percentage', 'Graduation School Name', 'Graduation Stream', 'Graduation Attempts',
                'Graduation Total', 'PhD Passing Year', 'PhD Percentage', 'PhD School Name', 'PG Stream', 'PG Attempts', 'PG Total',
                'Have you Qualified', 'Name of College', 'Other College Name', 'Name of Guide', 'Topic of PhD',
                'Concerned University', 'Department Name', 'Faculty', 'PhD Registration Date', 'PhD Registration Year',
                'PhD Registration Age', 'Family Annual Income', 'Income Certificate Number', 'Issuing Authority',
                'Income Issuing District', 'Income Issuing Taluka', 'Domicile', 'Domicile Certificate', 'Domicile Number',
                'Validity Certificate', 'Validity Cert Number', 'Validity Issuing District', 'Validity Issuing Taluka',
                'Validity Issuing Authority', 'Caste Certificate', 'Caste Certf Number', 'Issuing District', 'Caste Issuing Authority',
                'Salaried', 'Disability', 'Type of Disability', 'Father Name', 'Mother Name', 'Work in Government', 'Government Department',
                'Government Position', 'Bank Name', 'Account Number', 'IFSC Code', 'Account Holder Name', 'MICR' ])

    # Add data to the worksheet
    for row in data:
        ws.append(row)

    # Save the workbook in memory as bytes
    data = BytesIO()
    wb.save(data)
    data.seek(0)

    # Create a response object and attach the workbook as a file
    response = make_response(data.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=Accepted_Applications_2021.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    return response      


@app.route('/total_rejected_report_21', methods=['GET', 'POST'])
def total_rejected_report_21():
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor(dictionary=True)
    cursor.execute("SELECT * FROM application_page WHERE phd_registration_year='2021' and final_approval='rejected'")
    result = cursor.fetchall()
    cnx.commit()
    cursor.close()
    cnx.close()
    # print(result)
    return render_template('total_rejected_report_21.html', result=result)


#----------------------  EXPORT TO EXCEL Rejected Applications 2021 -------------------------------------------
@app.route('/export_rejected_applications_21', methods=['GET', 'POST'])
def export_rejected_applications_21():
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor()
    cursor.execute("  SELECT applicant_id, adhaar_number, first_name, last_name, middle_name, mobile_number, email, "
                   "date_of_birth, gender, age, caste, your_caste, subcaste, pvtg, pvtg_caste, marital_status," 
                    "dependents, state, district, taluka, village, city, add_1, add_2, pincode," 
                    "ssc_passing_year, ssc_percentage, ssc_school_name, ssc_stream, ssc_attempts, ssc_total," 
                    "hsc_passing_year, hsc_percentage, hsc_school_name, hsc_stream, hsc_attempts, hsc_total," 
                    "graduation_passing_year, graduation_percentage, graduation_school_name, grad_stream," 
                    "grad_attempts, grad_total, phd_passing_year, phd_percentage, phd_school_name, pg_stream, "
                    "pg_attempts, pg_total, have_you_qualified, name_of_college, other_college_name, "
                    "name_of_guide, topic_of_phd, concerned_university, department_name, faculty, "
                    "phd_registration_date, phd_registration_year, phd_registration_age, family_annual_income," 
                    "income_certificate_number, issuing_authority, income_issuing_district, income_issuing_taluka," 
                    "domicile, domicile_certificate, domicile_number, validity_certificate, validity_cert_number," 
                    "validity_issuing_district, validity_issuing_taluka, validity_issuing_authority, caste_certf, "
                    "caste_certf_number, issuing_district, caste_issuing_authority, salaried, disability, "
                    "type_of_disability, father_name, mother_name, work_in_government, gov_department," 
                    "gov_position, bank_name, account_number, ifsc_code, account_holder_name, micr FROM application_page"
                   " WHERE phd_registration_year='2021' and final_approval='rejected' ")
    data = cursor.fetchall()
    # Create a workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Add headers to the worksheet
    ws.append(['Applicant Id', 'Adhaar Card Number', 'First Name', 'Middle Name', 'Last Name', 'Mobile Number', 'Email',
               'Date Of Birth', 'Gender', 'Age', 'Caste/Tribe', 'Your Caste', 'Sub Caste', 'Are you PVTG', 'PVTG Caste/Tribe',
               'Marital Status', 'dependents', 'state', 'district', 'taluka', 'village', 'city', 'add_1', 'add_2',
               'pincode', 'SSC Passing Year', 'SSC Percentage', 'SSC School Name', 'SSC Stream', 'SSC Attempts',
                'SSC Total', 'HSC Passing Year', 'HSC Percentage', 'HSC School Name', 'HSC Stream', 'HSC Attempts', 'HSC Total',
                'Graduation Passing Year', 'Graduation Percentage', 'Graduation School Name', 'Graduation Stream', 'Graduation Attempts',
                'Graduation Total', 'PhD Passing Year', 'PhD Percentage', 'PhD School Name', 'PG Stream', 'PG Attempts', 'PG Total',
                'Have you Qualified', 'Name of College', 'Other College Name', 'Name of Guide', 'Topic of PhD',
                'Concerned University', 'Department Name', 'Faculty', 'PhD Registration Date', 'PhD Registration Year',
                'PhD Registration Age', 'Family Annual Income', 'Income Certificate Number', 'Issuing Authority',
                'Income Issuing District', 'Income Issuing Taluka', 'Domicile', 'Domicile Certificate', 'Domicile Number',
                'Validity Certificate', 'Validity Cert Number', 'Validity Issuing District', 'Validity Issuing Taluka',
                'Validity Issuing Authority', 'Caste Certificate', 'Caste Certf Number', 'Issuing District', 'Caste Issuing Authority',
                'Salaried', 'Disability', 'Type of Disability', 'Father Name', 'Mother Name', 'Work in Government', 'Government Department',
                'Government Position', 'Bank Name', 'Account Number', 'IFSC Code', 'Account Holder Name', 'MICR' ])

    # Add data to the worksheet
    for row in data:
        ws.append(row)

    # Save the workbook in memory as bytes
    data = BytesIO()
    wb.save(data)
    data.seek(0)

    # Create a response object and attach the workbook as a file
    response = make_response(data.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=Rejected_Applications_2021.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    return response     

# ----------------------------------------------------------------------------
# ------------------- END APPLICATIONS 2021 REPORTS ------------------------------
# ----------------------------------------------------------------------------

@app.route('/priority_people_report', methods=['GET', 'POST'])
def priority_people_report():
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor(dictionary=True)
    cursor.execute("SELECT * FROM application_page WHERE your_caste IN ('katkari', 'kolam', 'madia')")
    result = cursor.fetchall()
    cnx.commit()
    cursor.close()
    cnx.close()
    # print(result)
    return render_template('priority_people_report.html', result=result)


# ---------------------  Export Disability Report Yes list to Excel  -----------------------
@app.route('/export_priority_caste_report', methods=['GET', 'POST'])
def export_priority_caste_report():
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor()
    cursor.execute("SELECT applicant_id, first_name, middle_name, last_name, mobile_number, email, date_of_birth,"
                   "gender, age, caste, your_caste, marital_status, add_1, add_2, pincode, village, taluka, district,"
                   "state, city,  phd_registration_date, concerned_university, topic_of_phd, "
                   "name_of_guide, name_of_college, stream, board_university, admission_year, passing_year, percentage,"
                   "family_annual_income, income_certificate_number, issuing_authority, domicile, domicile_certificate,"
                   "domicile_number, caste_certf, issuing_district, caste_issuing_authority, salaried, disability,"
                   "father_name, mother_name, work_in_government, bank_name, account_number, ifsc_code," 
                   "account_holder_name FROM application_page WHERE your_caste IN ('katkari', 'kolam', 'madia')")    
    data = cursor.fetchall()
    # Create a workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Add headers to the worksheet
    ws.append(['Applicant Id', 'First Name', 'Middle Name', 'Last Name', 'Mobile Number', 'Email', 'Date Of Birth', 
                'Gender', 'Age', 'Caste', 'Your Caste', 'Marital Status', 'Dependents', 'Add 1', 'Add 2', 'Pincode', 'Village', 
                'Taluka', 'District', 'State', 'Phd Registration Date', 'Concerned University', 'Topic Of Phd', 'Name Of Guide', 
                'Name Of College', 'Stream', 'Board University', 'Admission Year', 'Passing Year', 'Percentage', 'Family Annual Income', 
                'Income Certificate Number', 'Issuing Authority', 'Domicile', 'Domicile Certificate', 'Domicile Number', 'Caste Certf', 
                'Issuing District', 'Caste Issuing Authority', 'Salaried', 'Disability', 'Father Name', 'Mother Name', 'Work In Government', 
                'Bank Name', 'Account Number', 'Ifsc Code', 'Account Holder Name'])

    # Add data to the worksheet
    for row in data:
        ws.append(row)

    # Save the workbook in memory as bytes
    data = BytesIO()
    wb.save(data)
    data.seek(0)

    # Create a response object and attach the workbook as a file
    response = make_response(data.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=Priority_Caste_Report.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    return response


@app.route('/all_news', methods=['GET', 'POST'])
def all_news():
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor(dictionary=True)
    cursor.execute("SELECT * FROM news_and_updates")
    result = cursor.fetchall()
    cnx.commit()
    cursor.close()
    cnx.close()
    # print(result)
    return render_template('all_news_updates.html', result=result)


#---------------------  Export News in Excel  -----------------------
@app.route('/export_news', methods=['GET', 'POST'])
def export_news():
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor()
    cursor.execute("SELECT id, user, title, subtitle FROM news_and_updates")
    data = cursor.fetchall()
    # Create a workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Add headers to the worksheet
    ws.append(['Id','User','Title','Subtitle'])

    # Add data to the worksheet
    for row in data:
        ws.append(row)

    # Save the workbook in memory as bytes
    data = BytesIO()
    wb.save(data)
    data.seek(0)

    # Create a response object and attach the workbook as a file
    response = make_response(data.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=News_and_Updates.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    return response

#---------------------  Export Rejected Users list in Excel  -----------------------
@app.route('/export_rejected_applications', methods=['GET', 'POST'])
def export_rejected_applications():
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor()
    cursor.execute("SELECT first_name,middle_name,last_name, mobile_number, email FROM application_page WHERE final_approval = 'rejected'")
    data = cursor.fetchall()
    # Create a workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Add headers to the worksheet
    ws.append(['first_name','middle_name','last_name','mobile_number','email','day','month','year','age'])

    # Add data to the worksheet
    for row in data:
        ws.append(row)

    # Save the workbook in memory as bytes
    data = BytesIO()
    wb.save(data)
    data.seek(0)

    # Create a response object and attach the workbook as a file
    response = make_response(data.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=Rejected_Users_Data.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    return response




@app.route('/preview', methods=['GET', 'POST'])
def preview():
    email = session['email']
    cnx = mysql.connector.connect(user='icswebapp', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor(dictionary=True)
    # Check if the user's email is in old_users for 2021 or 2022
    cursor.execute("SELECT year FROM signup WHERE email = %s AND year IN ('2020', '2021', '2022')", (email,))
    old_user_data = cursor.fetchone()
    concerned_university = request.get_data('concerned_university')
    cursor.execute("SELECT college_name FROM universities WHERE affiliated_universities = %s ", (concerned_university,))
    data_college = cursor.fetchone()
    cursor.execute('SELECT * FROM districts')
    district_list = cursor.fetchall()
    if old_user_data:
        session['user_type'] = 'old_user' 
        university_data = university_college()
        # User exists in old_users for 2021 or 2022, use this data
        return render_template('preview.html', records=old_user_data, editable=True, university_data=university_data, data_college=data_college, district_list = district_list)
    else:
        session['user_type'] = 'new_user'
        # User does not exist in old_users for 2021 or 2022, fetch from application_page
        cursor.execute("SELECT * FROM application_page WHERE email = %s", (email,))
        application_page_data = cursor.fetchone()
        university_data = university_college()
        return render_template('preview.html', records=application_page_data, editable=True, university_data=university_data, data_college=data_college, district_list = district_list)

    cursor.close()
    cnx.close()
    return render_template('preview.html', records=records, editable=True)


@app.route('/old_user_preview', methods=['GET', 'POST'])
def old_user_preview():
    if session.pop('logged_in_from_login', None):
        flash('Logged in Successfully', 'success')
    email = session['email']
    cnx = mysql.connector.connect(user='icswebapp', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor(dictionary=True)
    # Check if the user's email is in old_users for 2021 or 2022
    cursor.execute("SELECT year FROM signup WHERE email = %s AND year IN ('2020', '2021', '2022')", (email,))
    old_user = cursor.fetchone()
    concerned_university = request.get_data('concerned_university')
    cursor.execute("SELECT college_name FROM universities WHERE affiliated_universities = %s ", (concerned_university,))
    data_college = cursor.fetchone()
    cursor.execute('SELECT * FROM districts')
    district_list = cursor.fetchall()
    if old_user:
        session['user_type'] = 'old_user'
        university_data = university_college()
        cursor.execute("SELECT * FROM old_users WHERE email = %s", (email,))
        old_user_data = cursor.fetchone()
        # User exists in old_users for 2021 or 2022, use this data
        return render_template('old_user_preview.html', records=old_user_data, editable=True, university_data=university_data, data_college=data_college, district_list = district_list)
    else:
        session['user_type'] = 'new_user'
        # User does not exist in old_users for 2021 or 2022, fetch from application_page
        cursor.execute("SELECT * FROM application_page WHERE email = %s", (email,))
        application_page_data = cursor.fetchone()
        university_data = university_college()
        return render_template('old_user_preview.html', records=application_page_data, editable=True, university_data=university_data, data_college=data_college, district_list = district_list)

    cursor.close()
    cnx.close()
    return render_template('old_user_preview.html', records=records, editable=True, adhaar_error=adhaar_error)


def get_file_extension(filename):
    # Using os.path.splitext
    _, extension = os.path.splitext(filename)
    return extension.lower()



def save_file(file, firstname, lastname):
    if file:
        filename = f"{firstname}_{lastname}_{file.filename}"
        file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
        # return os.path.join(app.config['UPLOAD_FOLDER'], filename)
        return '/static/uploads/image_retrive/' + filename
    else:
        return "Save File"


def save_bulk_email_file(file):
    if file:
        filename = f"{file.filename}"
        file.save(os.path.join(app.config['EMAIL_DOCS'], filename))
        return '/static/uploads/sendbulkemails/' + filename
    else:
        return "Save File"


def allowed_file(filename):
    ALLOWED_EXTENSIONS = {'pdf', 'txt', 'jpg', 'jpeg', 'png'}
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def insert_into_old_users(email, applicant_id, phd_registration_date, date_of_birth, age):
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor()
    # Update the specific record with the matching email
    with cursor:
        sql = "UPDATE old_users SET applicant_id = %s, phd_registration_date = %s, date_of_birth = %s, age = %s WHERE email = %s"
        cursor.execute(sql, (applicant_id, phd_registration_date, date_of_birth, age, email))
        cnx.commit()
        cursor.close()
        cnx.close()



@app.errorhandler(404)
def page_not_found(error):
    requested_url = request.url
    flash(f'The page you are  for ("{requested_url}") is under Development, Please try again later.', 'warning')
    return redirect(request.referrer)


@app.errorhandler(500)
def page_not_found_500(error):
    requested_url = request.url
    flash(f'The page you are looking for ("{requested_url}") is under Development, Please try again later.', 'warning')
    return render_template(request.referrer)


@app.route('/developer_dashboard', methods=['GET', 'POST'])
def developer_dashboard():
    return render_template('developer_dashboard.html')


@app.route('/admin_DevDash')
def admin_devdash():
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor(dictionary=True)
    sql = """ SELECT * FROM admin """
    cursor.execute(sql)
    result = cursor.fetchall()

    cursor.close()
    cnx.close()
    return render_template('admin_devdash.html', result=result)


@app.route('/applicationPage_DevDash')
def applicationPage_DevDash():
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor(dictionary=True)
    sql = """ SELECT * FROM application_page """
    cursor.execute(sql)
    result = cursor.fetchall()

    cursor.close()
    cnx.close()
    return render_template('applicationPage_DevDash.html', result=result)


@app.route('/send_mail_incomplete_applications')
def send_mail_incomplete_applications():
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor(dictionary=True)
    sql = """ SELECT email FROM application_page where phd_registration_year>='2023' and form_filled='0' """
    # sql = """ SELECT email FROM application_page where email in ('tupotbhare@gmail.com', 'Tanmay@icspune.com') """
    cursor.execute(sql)
    result = cursor.fetchall()
    # message = 'This is a Test Email Please Ignore'
    # subject = 'Please Ignore - TEST EMAIL for Fellowship'
    message = 'Dear Applicant, This is a reminder to complete your application form for "Fellowship". Please complete and submit the form by 31st May 2024 at 6PM. Thank you for your attention to this matter.'
    subject = 'Regarding Incomplete Form for Fellowship'
    email_list = [row['email'] for row in result]  # Extracting emails from the result
    email_string = ', '.join(email_list)
    send_mail_incomplete(message, subject, email_list=email_list)
    sent_datetime = datetime.now()
    sent_date = sent_datetime.strftime('%Y-%m-%d')
    sent_time = sent_datetime.strftime('%H:%M:%S')
    sent_day = sent_datetime.strftime('%A')

    sql_insert = """INSERT INTO email_record_incompleteform (emails, message, subject, date, time, day) 
                        VALUES (%s, %s, %s, %s, %s, %s)"""
    cursor.execute(sql_insert, (email_string, message, subject, sent_date, sent_time, sent_day))
    # Commit changes to the database
    cnx.commit()
    # Close cursor and database connection
    cursor.close()
    cnx.close()
    return redirect(url_for('fetch_incomplete_form_emails'))


@app.route('/fetch_incomplete_form_emails', methods=['GET', 'POST'])
def fetch_incomplete_form_emails():
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor(dictionary=True)
    sql = """ SELECT * FROM email_record_incompleteform """
    cursor.execute(sql)
    result = cursor.fetchall()
    cnx.commit()
    # Close cursor and database connection
    cursor.close()
    cnx.close()
    return render_template('fetch_incomplete_form_emails.html', result=result)


def send_mail_incomplete(message, subject, email_list, filename=None, file_data=None):
    msg = Message(subject=subject, sender='helpdesk@trti-maha.in', recipients=email_list)
    msg_body = message
    msg.html = msg_body

    # Attach file if provided
    if filename and file_data:
        # Determine the MIME type based on the file extension
        mime_type, _ = mimetypes.guess_type(filename)
        # If MIME type is not detected, set a default value
        if not mime_type:
            mime_type = 'application/octet-stream'
        print(filename)
        # Add the attachment to the email message
        msg.attach(filename, mime_type, file_data)
    mail.send(msg)





@app.route('/delete_field/<int:id>/<field_value>', methods=['POST'])
def delete_field(id, field_value):
    try:
        cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                      host=host,
                                      database='ICSApplication')
        cursor = cnx.cursor(dictionary=True)
        # Build the SQL query to delete records where email matches the specified value
        sql = f"UPDATE application_page SET {field_value} = NULL WHERE id = %s"
        # Execute the SQL query with the specified email value
        cursor.execute(sql, (id,))
        # Commit the transaction
        cnx.commit()
        # Close the cursor
        cursor.close()
        # Redirect the user to the edit_student_admin_management route with the corresponding id
        return redirect(url_for('edit_student_admin_management', id=id))
    except Exception as e:
        print("Error deleting record:", e)
        return None


@app.route('/update_field/<int:id>/<field_name>', methods=['GET', 'POST'])
def update_field(id, field_name):
    try:
        cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                      host=host,
                                      database='ICSApplication')
        cursor = cnx.cursor(dictionary=True)
        new_value = request.form['new_value']
        # Build the SQL query to update the specified field value
        sql = f"UPDATE application_page SET {field_name} = %s WHERE id = %s"
        # Execute the SQL query with the specified field value and id
        cursor.execute(sql, (new_value, id))
        # Commit the transaction
        cnx.commit()
        # Close the cursor
        cursor.close()
        # Redirect the user to the edit_student_admin_management route with the corresponding id
        return redirect(url_for('edit_student_admin_management', id=id))
    except Exception as e:
        print("Error updating record:", e)
        return None


@app.route('/export_accepted_students_level1', methods=['GET', 'POST'])
def export_accepted_students_level1():
    if not session.get('logged_in'):
        # Redirect to the admin login page if the user is not logged in
        return redirect(url_for('admin_login'))
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor()
    cursor.execute("  SELECT applicant_id, adhaar_number, first_name, last_name, middle_name, mobile_number, email, "
                   "date_of_birth, gender, age, caste, your_caste, subcaste, pvtg, pvtg_caste, marital_status," 
                    "dependents, state, district, taluka, village, city, add_1, add_2, pincode," 
                    "ssc_passing_year, ssc_percentage, ssc_school_name, ssc_stream, ssc_attempts, ssc_total," 
                    "hsc_passing_year, hsc_percentage, hsc_school_name, hsc_stream, hsc_attempts, hsc_total," 
                    "graduation_passing_year, graduation_percentage, graduation_school_name, grad_stream," 
                    "grad_attempts, grad_total, phd_passing_year, phd_percentage, phd_school_name, pg_stream, "
                    "pg_attempts, pg_total, have_you_qualified, name_of_college, other_college_name, "
                    "name_of_guide, topic_of_phd, concerned_university, department_name, faculty, "
                    "phd_registration_date, phd_registration_year, phd_registration_age, family_annual_income," 
                    "income_certificate_number, issuing_authority, income_issuing_district, income_issuing_taluka," 
                    "domicile, domicile_certificate, domicile_number, validity_certificate, validity_cert_number," 
                    "validity_issuing_district, validity_issuing_taluka, validity_issuing_authority, caste_certf, "
                    "caste_certf_number, issuing_district, caste_issuing_authority, salaried, disability, "
                    "type_of_disability, father_name, mother_name, work_in_government, gov_department," 
                    "gov_position, bank_name, account_number, ifsc_code, account_holder_name, micr FROM application_page"
                   " WHERE phd_registration_year>='2023' and status='accepted' ")
    data = cursor.fetchall()
    # Create a workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Add headers to the worksheet
    # ws.append(['applicant_id','email','first_name','last_name','application_date'])

    ws.append(['Applicant Id', 'Adhaar Card Number', 'First Name', 'Middle Name', 'Last Name', 'Mobile Number', 'Email',
               'Date Of Birth', 'Gender', 'Age', 'Caste/Tribe', 'Your Caste', 'Sub Caste', 'Are you PVTG', 'PVTG Caste/Tribe',
               'Marital Status', 'dependents', 'state', 'district', 'taluka', 'village', 'city', 'add_1', 'add_2',
               'pincode', 'SSC Passing Year', 'SSC Percentage', 'SSC School Name', 'SSC Stream', 'SSC Attempts',
                'SSC Total', 'HSC Passing Year', 'HSC Percentage', 'HSC School Name', 'HSC Stream', 'HSC Attempts', 'HSC Total',
                'Graduation Passing Year', 'Graduation Percentage', 'Graduation School Name', 'Graduation Stream', 'Graduation Attempts',
                'Graduation Total', 'PhD Passing Year', 'PhD Percentage', 'PhD School Name', 'PG Stream', 'PG Attempts', 'PG Total',
                'Have you Qualified', 'Name of College', 'Other College Name', 'Name of Guide', 'Topic of PhD',
                'Concerned University', 'Department Name', 'Faculty', 'PhD Registration Date', 'PhD Registration Year',
                'PhD Registration Age', 'Family Annual Income', 'Income Certificate Number', 'Issuing Authority',
                'Income Issuing District', 'Income Issuing Taluka', 'Domicile', 'Domicile Certificate', 'Domicile Number',
                'Validity Certificate', 'Validity Cert Number', 'Validity Issuing District', 'Validity Issuing Taluka',
                'Validity Issuing Authority', 'Caste Certificate', 'Caste Certf Number', 'Issuing District', 'Caste Issuing Authority',
                'Salaried', 'Disability', 'Type of Disability', 'Father Name', 'Mother Name', 'Work in Government', 'Government Department',
                'Government Position', 'Bank Name', 'Account Number', 'IFSC Code', 'Account Holder Name', 'MICR' ])

    # Add data to the worksheet
    for row in data:
        ws.append(row)
    # Save the workbook in memory as bytes
    data = BytesIO()
    wb.save(data)
    data.seek(0)
    # Create a response object and attach the workbook as a file
    response = make_response(data.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=Level 1 Accepted Students.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    return response


@app.route('/export_pending_students_level1', methods=['GET', 'POST'])
def export_pending_students_level1():
    if not session.get('logged_in'):
        # Redirect to the admin login page if the user is not logged in
        return redirect(url_for('admin_login'))
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor()
    cursor.execute("  SELECT applicant_id, adhaar_number, first_name, last_name, middle_name, mobile_number, email, "
                   "date_of_birth, gender, age, caste, your_caste, subcaste, pvtg, pvtg_caste, marital_status," 
                    "dependents, state, district, taluka, village, city, add_1, add_2, pincode," 
                    "ssc_passing_year, ssc_percentage, ssc_school_name, ssc_stream, ssc_attempts, ssc_total," 
                    "hsc_passing_year, hsc_percentage, hsc_school_name, hsc_stream, hsc_attempts, hsc_total," 
                    "graduation_passing_year, graduation_percentage, graduation_school_name, grad_stream," 
                    "grad_attempts, grad_total, phd_passing_year, phd_percentage, phd_school_name, pg_stream, "
                    "pg_attempts, pg_total, have_you_qualified, name_of_college, other_college_name, "
                    "name_of_guide, topic_of_phd, concerned_university, department_name, faculty, "
                    "phd_registration_date, phd_registration_year, phd_registration_age, family_annual_income," 
                    "income_certificate_number, issuing_authority, income_issuing_district, income_issuing_taluka," 
                    "domicile, domicile_certificate, domicile_number, validity_certificate, validity_cert_number," 
                    "validity_issuing_district, validity_issuing_taluka, validity_issuing_authority, caste_certf, "
                    "caste_certf_number, issuing_district, caste_issuing_authority, salaried, disability, "
                    "type_of_disability, father_name, mother_name, work_in_government, gov_department," 
                    "gov_position, bank_name, account_number, ifsc_code, account_holder_name, micr FROM application_page"
                   " WHERE phd_registration_year>='2023' and status='pending' ")
    data = cursor.fetchall()
    # Create a workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Add headers to the worksheet
    # ws.append(['applicant_id','email','first_name','last_name','application_date'])

    ws.append(['Applicant Id', 'Adhaar Card Number', 'First Name', 'Middle Name', 'Last Name', 'Mobile Number', 'Email',
               'Date Of Birth', 'Gender', 'Age', 'Caste/Tribe', 'Your Caste', 'Sub Caste', 'Are you PVTG', 'PVTG Caste/Tribe',
               'Marital Status', 'dependents', 'state', 'district', 'taluka', 'village', 'city', 'add_1', 'add_2',
               'pincode', 'SSC Passing Year', 'SSC Percentage', 'SSC School Name', 'SSC Stream', 'SSC Attempts',
                'SSC Total', 'HSC Passing Year', 'HSC Percentage', 'HSC School Name', 'HSC Stream', 'HSC Attempts', 'HSC Total',
                'Graduation Passing Year', 'Graduation Percentage', 'Graduation School Name', 'Graduation Stream', 'Graduation Attempts',
                'Graduation Total', 'PhD Passing Year', 'PhD Percentage', 'PhD School Name', 'PG Stream', 'PG Attempts', 'PG Total',
                'Have you Qualified', 'Name of College', 'Other College Name', 'Name of Guide', 'Topic of PhD',
                'Concerned University', 'Department Name', 'Faculty', 'PhD Registration Date', 'PhD Registration Year',
                'PhD Registration Age', 'Family Annual Income', 'Income Certificate Number', 'Issuing Authority',
                'Income Issuing District', 'Income Issuing Taluka', 'Domicile', 'Domicile Certificate', 'Domicile Number',
                'Validity Certificate', 'Validity Cert Number', 'Validity Issuing District', 'Validity Issuing Taluka',
                'Validity Issuing Authority', 'Caste Certificate', 'Caste Certf Number', 'Issuing District', 'Caste Issuing Authority',
                'Salaried', 'Disability', 'Type of Disability', 'Father Name', 'Mother Name', 'Work in Government', 'Government Department',
                'Government Position', 'Bank Name', 'Account Number', 'IFSC Code', 'Account Holder Name', 'MICR' ])

    # Add data to the worksheet
    for row in data:
        ws.append(row)
    # Save the workbook in memory as bytes
    data = BytesIO()
    wb.save(data)
    data.seek(0)
    # Create a response object and attach the workbook as a file
    response = make_response(data.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=Level 1 Pending Students.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    return response


@app.route('/export_rejected_students_level1', methods=['GET', 'POST'])
def export_rejected_students_level1():
    if not session.get('logged_in'):
        # Redirect to the admin login page if the user is not logged in
        return redirect(url_for('admin_login'))
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor()
    cursor.execute("  SELECT applicant_id, adhaar_number, first_name, last_name, middle_name, mobile_number, email, "
                   "date_of_birth, gender, age, caste, your_caste, subcaste, pvtg, pvtg_caste, marital_status," 
                    "dependents, state, district, taluka, village, city, add_1, add_2, pincode," 
                    "ssc_passing_year, ssc_percentage, ssc_school_name, ssc_stream, ssc_attempts, ssc_total," 
                    "hsc_passing_year, hsc_percentage, hsc_school_name, hsc_stream, hsc_attempts, hsc_total," 
                    "graduation_passing_year, graduation_percentage, graduation_school_name, grad_stream," 
                    "grad_attempts, grad_total, phd_passing_year, phd_percentage, phd_school_name, pg_stream, "
                    "pg_attempts, pg_total, have_you_qualified, name_of_college, other_college_name, "
                    "name_of_guide, topic_of_phd, concerned_university, department_name, faculty, "
                    "phd_registration_date, phd_registration_year, phd_registration_age, family_annual_income," 
                    "income_certificate_number, issuing_authority, income_issuing_district, income_issuing_taluka," 
                    "domicile, domicile_certificate, domicile_number, validity_certificate, validity_cert_number," 
                    "validity_issuing_district, validity_issuing_taluka, validity_issuing_authority, caste_certf, "
                    "caste_certf_number, issuing_district, caste_issuing_authority, salaried, disability, "
                    "type_of_disability, father_name, mother_name, work_in_government, gov_department," 
                    "gov_position, bank_name, account_number, ifsc_code, account_holder_name, micr FROM application_page"
                   " WHERE phd_registration_year>='2023' and status='rejected' ")
    data = cursor.fetchall()
    # Create a workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Add headers to the worksheet
    # ws.append(['applicant_id','email','first_name','last_name','application_date'])

    ws.append(['Applicant Id', 'Adhaar Card Number', 'First Name', 'Middle Name', 'Last Name', 'Mobile Number', 'Email',
               'Date Of Birth', 'Gender', 'Age', 'Caste/Tribe', 'Your Caste', 'Sub Caste', 'Are you PVTG', 'PVTG Caste/Tribe',
               'Marital Status', 'dependents', 'state', 'district', 'taluka', 'village', 'city', 'add_1', 'add_2',
               'pincode', 'SSC Passing Year', 'SSC Percentage', 'SSC School Name', 'SSC Stream', 'SSC Attempts',
                'SSC Total', 'HSC Passing Year', 'HSC Percentage', 'HSC School Name', 'HSC Stream', 'HSC Attempts', 'HSC Total',
                'Graduation Passing Year', 'Graduation Percentage', 'Graduation School Name', 'Graduation Stream', 'Graduation Attempts',
                'Graduation Total', 'PhD Passing Year', 'PhD Percentage', 'PhD School Name', 'PG Stream', 'PG Attempts', 'PG Total',
                'Have you Qualified', 'Name of College', 'Other College Name', 'Name of Guide', 'Topic of PhD',
                'Concerned University', 'Department Name', 'Faculty', 'PhD Registration Date', 'PhD Registration Year',
                'PhD Registration Age', 'Family Annual Income', 'Income Certificate Number', 'Issuing Authority',
                'Income Issuing District', 'Income Issuing Taluka', 'Domicile', 'Domicile Certificate', 'Domicile Number',
                'Validity Certificate', 'Validity Cert Number', 'Validity Issuing District', 'Validity Issuing Taluka',
                'Validity Issuing Authority', 'Caste Certificate', 'Caste Certf Number', 'Issuing District', 'Caste Issuing Authority',
                'Salaried', 'Disability', 'Type of Disability', 'Father Name', 'Mother Name', 'Work in Government', 'Government Department',
                'Government Position', 'Bank Name', 'Account Number', 'IFSC Code', 'Account Holder Name', 'MICR' ])

    # Add data to the worksheet
    for row in data:
        ws.append(row)
    # Save the workbook in memory as bytes
    data = BytesIO()
    wb.save(data)
    data.seek(0)
    # Create a response object and attach the workbook as a file
    response = make_response(data.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=Level 1 Rejected Students.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    return response


@app.route('/export_pvtg_students_level1', methods=['GET', 'POST'])
def export_pvtg_students_level1():
    if not session.get('logged_in'):
        # Redirect to the admin login page if the user is not logged in
        return redirect(url_for('admin_login'))
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor()
    cursor.execute("  SELECT applicant_id, adhaar_number, first_name, last_name, middle_name, mobile_number, email, "
                   "date_of_birth, gender, age, caste, your_caste, subcaste, pvtg, pvtg_caste, marital_status," 
                    "dependents, state, district, taluka, village, city, add_1, add_2, pincode," 
                    "ssc_passing_year, ssc_percentage, ssc_school_name, ssc_stream, ssc_attempts, ssc_total," 
                    "hsc_passing_year, hsc_percentage, hsc_school_name, hsc_stream, hsc_attempts, hsc_total," 
                    "graduation_passing_year, graduation_percentage, graduation_school_name, grad_stream," 
                    "grad_attempts, grad_total, phd_passing_year, phd_percentage, phd_school_name, pg_stream, "
                    "pg_attempts, pg_total, have_you_qualified, name_of_college, other_college_name, "
                    "name_of_guide, topic_of_phd, concerned_university, department_name, faculty, "
                    "phd_registration_date, phd_registration_year, phd_registration_age, family_annual_income," 
                    "income_certificate_number, issuing_authority, income_issuing_district, income_issuing_taluka," 
                    "domicile, domicile_certificate, domicile_number, validity_certificate, validity_cert_number," 
                    "validity_issuing_district, validity_issuing_taluka, validity_issuing_authority, caste_certf, "
                    "caste_certf_number, issuing_district, caste_issuing_authority, salaried, disability, "
                    "type_of_disability, father_name, mother_name, work_in_government, gov_department," 
                    "gov_position, bank_name, account_number, ifsc_code, account_holder_name, micr FROM application_page"
                   " WHERE phd_registration_year>='2023' and your_caste IN ('katkari', 'kolam', 'madia') ")
    data = cursor.fetchall()
    # Create a workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Add headers to the worksheet
    # ws.append(['applicant_id','email','first_name','last_name','application_date'])

    ws.append(['Applicant Id', 'Adhaar Card Number', 'First Name', 'Middle Name', 'Last Name', 'Mobile Number', 'Email',
               'Date Of Birth', 'Gender', 'Age', 'Caste/Tribe', 'Your Caste', 'Sub Caste', 'Are you PVTG', 'PVTG Caste/Tribe',
               'Marital Status', 'dependents', 'state', 'district', 'taluka', 'village', 'city', 'add_1', 'add_2',
               'pincode', 'SSC Passing Year', 'SSC Percentage', 'SSC School Name', 'SSC Stream', 'SSC Attempts',
                'SSC Total', 'HSC Passing Year', 'HSC Percentage', 'HSC School Name', 'HSC Stream', 'HSC Attempts', 'HSC Total',
                'Graduation Passing Year', 'Graduation Percentage', 'Graduation School Name', 'Graduation Stream', 'Graduation Attempts',
                'Graduation Total', 'PhD Passing Year', 'PhD Percentage', 'PhD School Name', 'PG Stream', 'PG Attempts', 'PG Total',
                'Have you Qualified', 'Name of College', 'Other College Name', 'Name of Guide', 'Topic of PhD',
                'Concerned University', 'Department Name', 'Faculty', 'PhD Registration Date', 'PhD Registration Year',
                'PhD Registration Age', 'Family Annual Income', 'Income Certificate Number', 'Issuing Authority',
                'Income Issuing District', 'Income Issuing Taluka', 'Domicile', 'Domicile Certificate', 'Domicile Number',
                'Validity Certificate', 'Validity Cert Number', 'Validity Issuing District', 'Validity Issuing Taluka',
                'Validity Issuing Authority', 'Caste Certificate', 'Caste Certf Number', 'Issuing District', 'Caste Issuing Authority',
                'Salaried', 'Disability', 'Type of Disability', 'Father Name', 'Mother Name', 'Work in Government', 'Government Department',
                'Government Position', 'Bank Name', 'Account Number', 'IFSC Code', 'Account Holder Name', 'MICR' ])

    # Add data to the worksheet
    for row in data:
        ws.append(row)
    # Save the workbook in memory as bytes
    data = BytesIO()
    wb.save(data)
    data.seek(0)
    # Create a response object and attach the workbook as a file
    response = make_response(data.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=Level 1 PVTG Students.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    return response


@app.route('/export_disabled_students_level1', methods=['GET', 'POST'])
def export_disabled_students_level1():
    if not session.get('logged_in'):
        # Redirect to the admin login page if the user is not logged in
        return redirect(url_for('admin_login'))
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor()
    cursor.execute("  SELECT applicant_id, adhaar_number, first_name, last_name, middle_name, mobile_number, email, "
                   "date_of_birth, gender, age, caste, your_caste, subcaste, pvtg, pvtg_caste, marital_status," 
                    "dependents, state, district, taluka, village, city, add_1, add_2, pincode," 
                    "ssc_passing_year, ssc_percentage, ssc_school_name, ssc_stream, ssc_attempts, ssc_total," 
                    "hsc_passing_year, hsc_percentage, hsc_school_name, hsc_stream, hsc_attempts, hsc_total," 
                    "graduation_passing_year, graduation_percentage, graduation_school_name, grad_stream," 
                    "grad_attempts, grad_total, phd_passing_year, phd_percentage, phd_school_name, pg_stream, "
                    "pg_attempts, pg_total, have_you_qualified, name_of_college, other_college_name, "
                    "name_of_guide, topic_of_phd, concerned_university, department_name, faculty, "
                    "phd_registration_date, phd_registration_year, phd_registration_age, family_annual_income," 
                    "income_certificate_number, issuing_authority, income_issuing_district, income_issuing_taluka," 
                    "domicile, domicile_certificate, domicile_number, validity_certificate, validity_cert_number," 
                    "validity_issuing_district, validity_issuing_taluka, validity_issuing_authority, caste_certf, "
                    "caste_certf_number, issuing_district, caste_issuing_authority, salaried, disability, "
                    "type_of_disability, father_name, mother_name, work_in_government, gov_department," 
                    "gov_position, bank_name, account_number, ifsc_code, account_holder_name, micr FROM application_page"
                   " WHERE phd_registration_year>='2023' and disability='Yes' ")
    data = cursor.fetchall()
    # Create a workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Add headers to the worksheet
    # ws.append(['applicant_id','email','first_name','last_name','application_date'])

    ws.append(['Applicant Id', 'Adhaar Card Number', 'First Name', 'Middle Name', 'Last Name', 'Mobile Number', 'Email',
               'Date Of Birth', 'Gender', 'Age', 'Caste/Tribe', 'Your Caste', 'Sub Caste', 'Are you PVTG', 'PVTG Caste/Tribe',
               'Marital Status', 'dependents', 'state', 'district', 'taluka', 'village', 'city', 'add_1', 'add_2',
               'pincode', 'SSC Passing Year', 'SSC Percentage', 'SSC School Name', 'SSC Stream', 'SSC Attempts',
                'SSC Total', 'HSC Passing Year', 'HSC Percentage', 'HSC School Name', 'HSC Stream', 'HSC Attempts', 'HSC Total',
                'Graduation Passing Year', 'Graduation Percentage', 'Graduation School Name', 'Graduation Stream', 'Graduation Attempts',
                'Graduation Total', 'PhD Passing Year', 'PhD Percentage', 'PhD School Name', 'PG Stream', 'PG Attempts', 'PG Total',
                'Have you Qualified', 'Name of College', 'Other College Name', 'Name of Guide', 'Topic of PhD',
                'Concerned University', 'Department Name', 'Faculty', 'PhD Registration Date', 'PhD Registration Year',
                'PhD Registration Age', 'Family Annual Income', 'Income Certificate Number', 'Issuing Authority',
                'Income Issuing District', 'Income Issuing Taluka', 'Domicile', 'Domicile Certificate', 'Domicile Number',
                'Validity Certificate', 'Validity Cert Number', 'Validity Issuing District', 'Validity Issuing Taluka',
                'Validity Issuing Authority', 'Caste Certificate', 'Caste Certf Number', 'Issuing District', 'Caste Issuing Authority',
                'Salaried', 'Disability', 'Type of Disability', 'Father Name', 'Mother Name', 'Work in Government', 'Government Department',
                'Government Position', 'Bank Name', 'Account Number', 'IFSC Code', 'Account Holder Name', 'MICR' ])

    # Add data to the worksheet
    for row in data:
        ws.append(row)
    # Save the workbook in memory as bytes
    data = BytesIO()
    wb.save(data)
    data.seek(0)
    # Create a response object and attach the workbook as a file
    response = make_response(data.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=Level 1 Disabled Students.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    return response
# ---------------------------------------------------------------------------------------
# --------------------------------- LEVEL 2 ADMIN BUTTON ROUTES -------------------------
# ---------------------------------------------------------------------------------------

@app.route('/export_accepted_students_level2', methods=['GET', 'POST'])
def export_accepted_students_level2():
    if not session.get('logged_in'):
        # Redirect to the admin login page if the user is not logged in
        return redirect(url_for('admin_login'))
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor()
    cursor.execute("  SELECT applicant_id, adhaar_number, first_name, last_name, middle_name, mobile_number, email, "
                   "date_of_birth, gender, age, caste, your_caste, subcaste, pvtg, pvtg_caste, marital_status," 
                    "dependents, state, district, taluka, village, city, add_1, add_2, pincode," 
                    "ssc_passing_year, ssc_percentage, ssc_school_name, ssc_stream, ssc_attempts, ssc_total," 
                    "hsc_passing_year, hsc_percentage, hsc_school_name, hsc_stream, hsc_attempts, hsc_total," 
                    "graduation_passing_year, graduation_percentage, graduation_school_name, grad_stream," 
                    "grad_attempts, grad_total, phd_passing_year, phd_percentage, phd_school_name, pg_stream, "
                    "pg_attempts, pg_total, have_you_qualified, name_of_college, other_college_name, "
                    "name_of_guide, topic_of_phd, concerned_university, department_name, faculty, "
                    "phd_registration_date, phd_registration_year, phd_registration_age, family_annual_income," 
                    "income_certificate_number, issuing_authority, income_issuing_district, income_issuing_taluka," 
                    "domicile, domicile_certificate, domicile_number, validity_certificate, validity_cert_number," 
                    "validity_issuing_district, validity_issuing_taluka, validity_issuing_authority, caste_certf, "
                    "caste_certf_number, issuing_district, caste_issuing_authority, salaried, disability, "
                    "type_of_disability, father_name, mother_name, work_in_government, gov_department," 
                    "gov_position, bank_name, account_number, ifsc_code, account_holder_name, micr FROM application_page"
                   " WHERE phd_registration_year>='2023' and scrutiny_status='accepted' and status='accepted' ")
    data = cursor.fetchall()
    # Create a workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Add headers to the worksheet
    # ws.append(['applicant_id','email','first_name','last_name','application_date'])

    ws.append(['Applicant Id', 'Adhaar Card Number', 'First Name', 'Middle Name', 'Last Name', 'Mobile Number', 'Email',
               'Date Of Birth', 'Gender', 'Age', 'Caste/Tribe', 'Your Caste', 'Sub Caste', 'Are you PVTG', 'PVTG Caste/Tribe',
               'Marital Status', 'dependents', 'state', 'district', 'taluka', 'village', 'city', 'add_1', 'add_2',
               'pincode', 'SSC Passing Year', 'SSC Percentage', 'SSC School Name', 'SSC Stream', 'SSC Attempts',
                'SSC Total', 'HSC Passing Year', 'HSC Percentage', 'HSC School Name', 'HSC Stream', 'HSC Attempts', 'HSC Total',
                'Graduation Passing Year', 'Graduation Percentage', 'Graduation School Name', 'Graduation Stream', 'Graduation Attempts',
                'Graduation Total', 'PhD Passing Year', 'PhD Percentage', 'PhD School Name', 'PG Stream', 'PG Attempts', 'PG Total',
                'Have you Qualified', 'Name of College', 'Other College Name', 'Name of Guide', 'Topic of PhD',
                'Concerned University', 'Department Name', 'Faculty', 'PhD Registration Date', 'PhD Registration Year',
                'PhD Registration Age', 'Family Annual Income', 'Income Certificate Number', 'Issuing Authority',
                'Income Issuing District', 'Income Issuing Taluka', 'Domicile', 'Domicile Certificate', 'Domicile Number',
                'Validity Certificate', 'Validity Cert Number', 'Validity Issuing District', 'Validity Issuing Taluka',
                'Validity Issuing Authority', 'Caste Certificate', 'Caste Certf Number', 'Issuing District', 'Caste Issuing Authority',
                'Salaried', 'Disability', 'Type of Disability', 'Father Name', 'Mother Name', 'Work in Government', 'Government Department',
                'Government Position', 'Bank Name', 'Account Number', 'IFSC Code', 'Account Holder Name', 'MICR' ])

    # Add data to the worksheet
    for row in data:
        ws.append(row)
    # Save the workbook in memory as bytes
    data = BytesIO()
    wb.save(data)
    data.seek(0)
    # Create a response object and attach the workbook as a file
    response = make_response(data.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=Level 2 Accepted Students.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    return response


@app.route('/export_pending_students_level2', methods=['GET', 'POST'])
def export_pending_students_level2():
    if not session.get('logged_in'):
        # Redirect to the admin login page if the user is not logged in
        return redirect(url_for('admin_login'))
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor()
    cursor.execute("  SELECT applicant_id, adhaar_number, first_name, last_name, middle_name, mobile_number, email, "
                   "date_of_birth, gender, age, caste, your_caste, subcaste, pvtg, pvtg_caste, marital_status," 
                    "dependents, state, district, taluka, village, city, add_1, add_2, pincode," 
                    "ssc_passing_year, ssc_percentage, ssc_school_name, ssc_stream, ssc_attempts, ssc_total," 
                    "hsc_passing_year, hsc_percentage, hsc_school_name, hsc_stream, hsc_attempts, hsc_total," 
                    "graduation_passing_year, graduation_percentage, graduation_school_name, grad_stream," 
                    "grad_attempts, grad_total, phd_passing_year, phd_percentage, phd_school_name, pg_stream, "
                    "pg_attempts, pg_total, have_you_qualified, name_of_college, other_college_name, "
                    "name_of_guide, topic_of_phd, concerned_university, department_name, faculty, "
                    "phd_registration_date, phd_registration_year, phd_registration_age, family_annual_income," 
                    "income_certificate_number, issuing_authority, income_issuing_district, income_issuing_taluka," 
                    "domicile, domicile_certificate, domicile_number, validity_certificate, validity_cert_number," 
                    "validity_issuing_district, validity_issuing_taluka, validity_issuing_authority, caste_certf, "
                    "caste_certf_number, issuing_district, caste_issuing_authority, salaried, disability, "
                    "type_of_disability, father_name, mother_name, work_in_government, gov_department," 
                    "gov_position, bank_name, account_number, ifsc_code, account_holder_name, micr FROM application_page"
                   " WHERE phd_registration_year>='2023' and scrutiny_status='pending' ")
    data = cursor.fetchall()
    # Create a workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Add headers to the worksheet
    # ws.append(['applicant_id','email','first_name','last_name','application_date'])

    ws.append(['Applicant Id', 'Adhaar Card Number', 'First Name', 'Middle Name', 'Last Name', 'Mobile Number', 'Email',
               'Date Of Birth', 'Gender', 'Age', 'Caste/Tribe', 'Your Caste', 'Sub Caste', 'Are you PVTG', 'PVTG Caste/Tribe',
               'Marital Status', 'dependents', 'state', 'district', 'taluka', 'village', 'city', 'add_1', 'add_2',
               'pincode', 'SSC Passing Year', 'SSC Percentage', 'SSC School Name', 'SSC Stream', 'SSC Attempts',
                'SSC Total', 'HSC Passing Year', 'HSC Percentage', 'HSC School Name', 'HSC Stream', 'HSC Attempts', 'HSC Total',
                'Graduation Passing Year', 'Graduation Percentage', 'Graduation School Name', 'Graduation Stream', 'Graduation Attempts',
                'Graduation Total', 'PhD Passing Year', 'PhD Percentage', 'PhD School Name', 'PG Stream', 'PG Attempts', 'PG Total',
                'Have you Qualified', 'Name of College', 'Other College Name', 'Name of Guide', 'Topic of PhD',
                'Concerned University', 'Department Name', 'Faculty', 'PhD Registration Date', 'PhD Registration Year',
                'PhD Registration Age', 'Family Annual Income', 'Income Certificate Number', 'Issuing Authority',
                'Income Issuing District', 'Income Issuing Taluka', 'Domicile', 'Domicile Certificate', 'Domicile Number',
                'Validity Certificate', 'Validity Cert Number', 'Validity Issuing District', 'Validity Issuing Taluka',
                'Validity Issuing Authority', 'Caste Certificate', 'Caste Certf Number', 'Issuing District', 'Caste Issuing Authority',
                'Salaried', 'Disability', 'Type of Disability', 'Father Name', 'Mother Name', 'Work in Government', 'Government Department',
                'Government Position', 'Bank Name', 'Account Number', 'IFSC Code', 'Account Holder Name', 'MICR' ])

    # Add data to the worksheet
    for row in data:
        ws.append(row)
    # Save the workbook in memory as bytes
    data = BytesIO()
    wb.save(data)
    data.seek(0)
    # Create a response object and attach the workbook as a file
    response = make_response(data.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=Level 2 Pending Students.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    return response


@app.route('/export_rejected_students_level2', methods=['GET', 'POST'])
def export_rejected_students_level2():
    if not session.get('logged_in'):
        # Redirect to the admin login page if the user is not logged in
        return redirect(url_for('admin_login'))
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor()
    cursor.execute("  SELECT applicant_id, adhaar_number, first_name, last_name, middle_name, mobile_number, email, "
                   "date_of_birth, gender, age, caste, your_caste, subcaste, pvtg, pvtg_caste, marital_status," 
                    "dependents, state, district, taluka, village, city, add_1, add_2, pincode," 
                    "ssc_passing_year, ssc_percentage, ssc_school_name, ssc_stream, ssc_attempts, ssc_total," 
                    "hsc_passing_year, hsc_percentage, hsc_school_name, hsc_stream, hsc_attempts, hsc_total," 
                    "graduation_passing_year, graduation_percentage, graduation_school_name, grad_stream," 
                    "grad_attempts, grad_total, phd_passing_year, phd_percentage, phd_school_name, pg_stream, "
                    "pg_attempts, pg_total, have_you_qualified, name_of_college, other_college_name, "
                    "name_of_guide, topic_of_phd, concerned_university, department_name, faculty, "
                    "phd_registration_date, phd_registration_year, phd_registration_age, family_annual_income," 
                    "income_certificate_number, issuing_authority, income_issuing_district, income_issuing_taluka," 
                    "domicile, domicile_certificate, domicile_number, validity_certificate, validity_cert_number," 
                    "validity_issuing_district, validity_issuing_taluka, validity_issuing_authority, caste_certf, "
                    "caste_certf_number, issuing_district, caste_issuing_authority, salaried, disability, "
                    "type_of_disability, father_name, mother_name, work_in_government, gov_department," 
                    "gov_position, bank_name, account_number, ifsc_code, account_holder_name, micr FROM application_page"
                   " WHERE phd_registration_year>='2023' and scrutiny_status='rejected' and status='accepted' ")
    data = cursor.fetchall()
    # Create a workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Add headers to the worksheet
    # ws.append(['applicant_id','email','first_name','last_name','application_date'])

    ws.append(['Applicant Id', 'Adhaar Card Number', 'First Name', 'Middle Name', 'Last Name', 'Mobile Number', 'Email',
               'Date Of Birth', 'Gender', 'Age', 'Caste/Tribe', 'Your Caste', 'Sub Caste', 'Are you PVTG', 'PVTG Caste/Tribe',
               'Marital Status', 'dependents', 'state', 'district', 'taluka', 'village', 'city', 'add_1', 'add_2',
               'pincode', 'SSC Passing Year', 'SSC Percentage', 'SSC School Name', 'SSC Stream', 'SSC Attempts',
                'SSC Total', 'HSC Passing Year', 'HSC Percentage', 'HSC School Name', 'HSC Stream', 'HSC Attempts', 'HSC Total',
                'Graduation Passing Year', 'Graduation Percentage', 'Graduation School Name', 'Graduation Stream', 'Graduation Attempts',
                'Graduation Total', 'PhD Passing Year', 'PhD Percentage', 'PhD School Name', 'PG Stream', 'PG Attempts', 'PG Total',
                'Have you Qualified', 'Name of College', 'Other College Name', 'Name of Guide', 'Topic of PhD',
                'Concerned University', 'Department Name', 'Faculty', 'PhD Registration Date', 'PhD Registration Year',
                'PhD Registration Age', 'Family Annual Income', 'Income Certificate Number', 'Issuing Authority',
                'Income Issuing District', 'Income Issuing Taluka', 'Domicile', 'Domicile Certificate', 'Domicile Number',
                'Validity Certificate', 'Validity Cert Number', 'Validity Issuing District', 'Validity Issuing Taluka',
                'Validity Issuing Authority', 'Caste Certificate', 'Caste Certf Number', 'Issuing District', 'Caste Issuing Authority',
                'Salaried', 'Disability', 'Type of Disability', 'Father Name', 'Mother Name', 'Work in Government', 'Government Department',
                'Government Position', 'Bank Name', 'Account Number', 'IFSC Code', 'Account Holder Name', 'MICR' ])

    # Add data to the worksheet
    for row in data:
        ws.append(row)
    # Save the workbook in memory as bytes
    data = BytesIO()
    wb.save(data)
    data.seek(0)
    # Create a response object and attach the workbook as a file
    response = make_response(data.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=Level 2 Rejected Students.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    return response


@app.route('/export_pvtg_students_level2', methods=['GET', 'POST'])
def export_pvtg_students_level2():
    if not session.get('logged_in'):
        # Redirect to the admin login page if the user is not logged in
        return redirect(url_for('admin_login'))
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor()
    cursor.execute("  SELECT applicant_id, adhaar_number, first_name, last_name, middle_name, mobile_number, email, "
                   "date_of_birth, gender, age, caste, your_caste, subcaste, pvtg, pvtg_caste, marital_status," 
                    "dependents, state, district, taluka, village, city, add_1, add_2, pincode," 
                    "ssc_passing_year, ssc_percentage, ssc_school_name, ssc_stream, ssc_attempts, ssc_total," 
                    "hsc_passing_year, hsc_percentage, hsc_school_name, hsc_stream, hsc_attempts, hsc_total," 
                    "graduation_passing_year, graduation_percentage, graduation_school_name, grad_stream," 
                    "grad_attempts, grad_total, phd_passing_year, phd_percentage, phd_school_name, pg_stream, "
                    "pg_attempts, pg_total, have_you_qualified, name_of_college, other_college_name, "
                    "name_of_guide, topic_of_phd, concerned_university, department_name, faculty, "
                    "phd_registration_date, phd_registration_year, phd_registration_age, family_annual_income," 
                    "income_certificate_number, issuing_authority, income_issuing_district, income_issuing_taluka," 
                    "domicile, domicile_certificate, domicile_number, validity_certificate, validity_cert_number," 
                    "validity_issuing_district, validity_issuing_taluka, validity_issuing_authority, caste_certf, "
                    "caste_certf_number, issuing_district, caste_issuing_authority, salaried, disability, "
                    "type_of_disability, father_name, mother_name, work_in_government, gov_department," 
                    "gov_position, bank_name, account_number, ifsc_code, account_holder_name, micr FROM application_page"
                   " WHERE phd_registration_year>='2023' and your_caste IN ('katkari', 'kolam', 'madia') ")
    data = cursor.fetchall()
    # Create a workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Add headers to the worksheet
    # ws.append(['applicant_id','email','first_name','last_name','application_date'])

    ws.append(['Applicant Id', 'Adhaar Card Number', 'First Name', 'Middle Name', 'Last Name', 'Mobile Number', 'Email',
               'Date Of Birth', 'Gender', 'Age', 'Caste/Tribe', 'Your Caste', 'Sub Caste', 'Are you PVTG', 'PVTG Caste/Tribe',
               'Marital Status', 'dependents', 'state', 'district', 'taluka', 'village', 'city', 'add_1', 'add_2',
               'pincode', 'SSC Passing Year', 'SSC Percentage', 'SSC School Name', 'SSC Stream', 'SSC Attempts',
                'SSC Total', 'HSC Passing Year', 'HSC Percentage', 'HSC School Name', 'HSC Stream', 'HSC Attempts', 'HSC Total',
                'Graduation Passing Year', 'Graduation Percentage', 'Graduation School Name', 'Graduation Stream', 'Graduation Attempts',
                'Graduation Total', 'PhD Passing Year', 'PhD Percentage', 'PhD School Name', 'PG Stream', 'PG Attempts', 'PG Total',
                'Have you Qualified', 'Name of College', 'Other College Name', 'Name of Guide', 'Topic of PhD',
                'Concerned University', 'Department Name', 'Faculty', 'PhD Registration Date', 'PhD Registration Year',
                'PhD Registration Age', 'Family Annual Income', 'Income Certificate Number', 'Issuing Authority',
                'Income Issuing District', 'Income Issuing Taluka', 'Domicile', 'Domicile Certificate', 'Domicile Number',
                'Validity Certificate', 'Validity Cert Number', 'Validity Issuing District', 'Validity Issuing Taluka',
                'Validity Issuing Authority', 'Caste Certificate', 'Caste Certf Number', 'Issuing District', 'Caste Issuing Authority',
                'Salaried', 'Disability', 'Type of Disability', 'Father Name', 'Mother Name', 'Work in Government', 'Government Department',
                'Government Position', 'Bank Name', 'Account Number', 'IFSC Code', 'Account Holder Name', 'MICR' ])

    # Add data to the worksheet
    for row in data:
        ws.append(row)
    # Save the workbook in memory as bytes
    data = BytesIO()
    wb.save(data)
    data.seek(0)
    # Create a response object and attach the workbook as a file
    response = make_response(data.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=Level 2 PVTG Students.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    return response


@app.route('/export_disabled_students_level2', methods=['GET', 'POST'])
def export_disabled_students_level2():
    if not session.get('logged_in'):
        # Redirect to the admin login page if the user is not logged in
        return redirect(url_for('admin_login'))
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor()
    cursor.execute("  SELECT applicant_id, adhaar_number, first_name, last_name, middle_name, mobile_number, email, "
                   "date_of_birth, gender, age, caste, your_caste, subcaste, pvtg, pvtg_caste, marital_status," 
                    "dependents, state, district, taluka, village, city, add_1, add_2, pincode," 
                    "ssc_passing_year, ssc_percentage, ssc_school_name, ssc_stream, ssc_attempts, ssc_total," 
                    "hsc_passing_year, hsc_percentage, hsc_school_name, hsc_stream, hsc_attempts, hsc_total," 
                    "graduation_passing_year, graduation_percentage, graduation_school_name, grad_stream," 
                    "grad_attempts, grad_total, phd_passing_year, phd_percentage, phd_school_name, pg_stream, "
                    "pg_attempts, pg_total, have_you_qualified, name_of_college, other_college_name, "
                    "name_of_guide, topic_of_phd, concerned_university, department_name, faculty, "
                    "phd_registration_date, phd_registration_year, phd_registration_age, family_annual_income," 
                    "income_certificate_number, issuing_authority, income_issuing_district, income_issuing_taluka," 
                    "domicile, domicile_certificate, domicile_number, validity_certificate, validity_cert_number," 
                    "validity_issuing_district, validity_issuing_taluka, validity_issuing_authority, caste_certf, "
                    "caste_certf_number, issuing_district, caste_issuing_authority, salaried, disability, "
                    "type_of_disability, father_name, mother_name, work_in_government, gov_department," 
                    "gov_position, bank_name, account_number, ifsc_code, account_holder_name, micr FROM application_page"
                   " WHERE phd_registration_year>='2023' and disability='Yes' ")
    data = cursor.fetchall()
    # Create a workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Add headers to the worksheet
    # ws.append(['applicant_id','email','first_name','last_name','application_date'])

    ws.append(['Applicant Id', 'Adhaar Card Number', 'First Name', 'Middle Name', 'Last Name', 'Mobile Number', 'Email',
               'Date Of Birth', 'Gender', 'Age', 'Caste/Tribe', 'Your Caste', 'Sub Caste', 'Are you PVTG', 'PVTG Caste/Tribe',
               'Marital Status', 'dependents', 'state', 'district', 'taluka', 'village', 'city', 'add_1', 'add_2',
               'pincode', 'SSC Passing Year', 'SSC Percentage', 'SSC School Name', 'SSC Stream', 'SSC Attempts',
                'SSC Total', 'HSC Passing Year', 'HSC Percentage', 'HSC School Name', 'HSC Stream', 'HSC Attempts', 'HSC Total',
                'Graduation Passing Year', 'Graduation Percentage', 'Graduation School Name', 'Graduation Stream', 'Graduation Attempts',
                'Graduation Total', 'PhD Passing Year', 'PhD Percentage', 'PhD School Name', 'PG Stream', 'PG Attempts', 'PG Total',
                'Have you Qualified', 'Name of College', 'Other College Name', 'Name of Guide', 'Topic of PhD',
                'Concerned University', 'Department Name', 'Faculty', 'PhD Registration Date', 'PhD Registration Year',
                'PhD Registration Age', 'Family Annual Income', 'Income Certificate Number', 'Issuing Authority',
                'Income Issuing District', 'Income Issuing Taluka', 'Domicile', 'Domicile Certificate', 'Domicile Number',
                'Validity Certificate', 'Validity Cert Number', 'Validity Issuing District', 'Validity Issuing Taluka',
                'Validity Issuing Authority', 'Caste Certificate', 'Caste Certf Number', 'Issuing District', 'Caste Issuing Authority',
                'Salaried', 'Disability', 'Type of Disability', 'Father Name', 'Mother Name', 'Work in Government', 'Government Department',
                'Government Position', 'Bank Name', 'Account Number', 'IFSC Code', 'Account Holder Name', 'MICR' ])

    # Add data to the worksheet
    for row in data:
        ws.append(row)
    # Save the workbook in memory as bytes
    data = BytesIO()
    wb.save(data)
    data.seek(0)
    # Create a response object and attach the workbook as a file
    response = make_response(data.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=Level 2 Disabled Students.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    return response
# ---------------------------------------------------------------------------------------
# --------------------------------- LEVEL 3 ADMIN BUTTON ROUTES -------------------------
# ---------------------------------------------------------------------------------------

@app.route('/export_accepted_students_level3', methods=['GET', 'POST'])
def export_accepted_students_level3():
    if not session.get('logged_in'):
        # Redirect to the admin login page if the user is not logged in
        return redirect(url_for('admin_login'))
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor()
    cursor.execute("  SELECT applicant_id, adhaar_number, first_name, last_name, middle_name, mobile_number, email, "
                   "date_of_birth, gender, age, caste, your_caste, subcaste, pvtg, pvtg_caste, marital_status," 
                    "dependents, state, district, taluka, village, city, add_1, add_2, pincode," 
                    "ssc_passing_year, ssc_percentage, ssc_school_name, ssc_stream, ssc_attempts, ssc_total," 
                    "hsc_passing_year, hsc_percentage, hsc_school_name, hsc_stream, hsc_attempts, hsc_total," 
                    "graduation_passing_year, graduation_percentage, graduation_school_name, grad_stream," 
                    "grad_attempts, grad_total, phd_passing_year, phd_percentage, phd_school_name, pg_stream, "
                    "pg_attempts, pg_total, have_you_qualified, name_of_college, other_college_name, "
                    "name_of_guide, topic_of_phd, concerned_university, department_name, faculty, "
                    "phd_registration_date, phd_registration_year, phd_registration_age, family_annual_income," 
                    "income_certificate_number, issuing_authority, income_issuing_district, income_issuing_taluka," 
                    "domicile, domicile_certificate, domicile_number, validity_certificate, validity_cert_number," 
                    "validity_issuing_district, validity_issuing_taluka, validity_issuing_authority, caste_certf, "
                    "caste_certf_number, issuing_district, caste_issuing_authority, salaried, disability, "
                    "type_of_disability, father_name, mother_name, work_in_government, gov_department," 
                    "gov_position, bank_name, account_number, ifsc_code, account_holder_name, micr FROM application_page"
                   " WHERE phd_registration_year>='2023' and final_approval='accepted' and scrutiny_status='accepted' ")
    data = cursor.fetchall()
    # Create a workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Add headers to the worksheet
    # ws.append(['applicant_id','email','first_name','last_name','application_date'])

    ws.append(['Applicant Id', 'Adhaar Card Number', 'First Name', 'Middle Name', 'Last Name', 'Mobile Number', 'Email',
               'Date Of Birth', 'Gender', 'Age', 'Caste/Tribe', 'Your Caste', 'Sub Caste', 'Are you PVTG', 'PVTG Caste/Tribe',
               'Marital Status', 'dependents', 'state', 'district', 'taluka', 'village', 'city', 'add_1', 'add_2',
               'pincode', 'SSC Passing Year', 'SSC Percentage', 'SSC School Name', 'SSC Stream', 'SSC Attempts',
                'SSC Total', 'HSC Passing Year', 'HSC Percentage', 'HSC School Name', 'HSC Stream', 'HSC Attempts', 'HSC Total',
                'Graduation Passing Year', 'Graduation Percentage', 'Graduation School Name', 'Graduation Stream', 'Graduation Attempts',
                'Graduation Total', 'PhD Passing Year', 'PhD Percentage', 'PhD School Name', 'PG Stream', 'PG Attempts', 'PG Total',
                'Have you Qualified', 'Name of College', 'Other College Name', 'Name of Guide', 'Topic of PhD',
                'Concerned University', 'Department Name', 'Faculty', 'PhD Registration Date', 'PhD Registration Year',
                'PhD Registration Age', 'Family Annual Income', 'Income Certificate Number', 'Issuing Authority',
                'Income Issuing District', 'Income Issuing Taluka', 'Domicile', 'Domicile Certificate', 'Domicile Number',
                'Validity Certificate', 'Validity Cert Number', 'Validity Issuing District', 'Validity Issuing Taluka',
                'Validity Issuing Authority', 'Caste Certificate', 'Caste Certf Number', 'Issuing District', 'Caste Issuing Authority',
                'Salaried', 'Disability', 'Type of Disability', 'Father Name', 'Mother Name', 'Work in Government', 'Government Department',
                'Government Position', 'Bank Name', 'Account Number', 'IFSC Code', 'Account Holder Name', 'MICR' ])

    # Add data to the worksheet
    for row in data:
        ws.append(row)
    # Save the workbook in memory as bytes
    data = BytesIO()
    wb.save(data)
    data.seek(0)
    # Create a response object and attach the workbook as a file
    response = make_response(data.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=Level 3 Accepted Students.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    return response


@app.route('/export_pending_students_level3', methods=['GET', 'POST'])
def export_pending_students_level3():
    if not session.get('logged_in'):
        # Redirect to the admin login page if the user is not logged in
        return redirect(url_for('admin_login'))
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor()
    cursor.execute("  SELECT applicant_id, adhaar_number, first_name, last_name, middle_name, mobile_number, email, "
                   "date_of_birth, gender, age, caste, your_caste, subcaste, pvtg, pvtg_caste, marital_status," 
                    "dependents, state, district, taluka, village, city, add_1, add_2, pincode," 
                    "ssc_passing_year, ssc_percentage, ssc_school_name, ssc_stream, ssc_attempts, ssc_total," 
                    "hsc_passing_year, hsc_percentage, hsc_school_name, hsc_stream, hsc_attempts, hsc_total," 
                    "graduation_passing_year, graduation_percentage, graduation_school_name, grad_stream," 
                    "grad_attempts, grad_total, phd_passing_year, phd_percentage, phd_school_name, pg_stream, "
                    "pg_attempts, pg_total, have_you_qualified, name_of_college, other_college_name, "
                    "name_of_guide, topic_of_phd, concerned_university, department_name, faculty, "
                    "phd_registration_date, phd_registration_year, phd_registration_age, family_annual_income," 
                    "income_certificate_number, issuing_authority, income_issuing_district, income_issuing_taluka," 
                    "domicile, domicile_certificate, domicile_number, validity_certificate, validity_cert_number," 
                    "validity_issuing_district, validity_issuing_taluka, validity_issuing_authority, caste_certf, "
                    "caste_certf_number, issuing_district, caste_issuing_authority, salaried, disability, "
                    "type_of_disability, father_name, mother_name, work_in_government, gov_department," 
                    "gov_position, bank_name, account_number, ifsc_code, account_holder_name, micr FROM application_page"
                   " WHERE phd_registration_year>='2023' and final_approval='pending' and scrutiny_status='accepted' ")
    data = cursor.fetchall()
    # Create a workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Add headers to the worksheet
    # ws.append(['applicant_id','email','first_name','last_name','application_date'])

    ws.append(['Applicant Id', 'Adhaar Card Number', 'First Name', 'Middle Name', 'Last Name', 'Mobile Number', 'Email',
               'Date Of Birth', 'Gender', 'Age', 'Caste/Tribe', 'Your Caste', 'Sub Caste', 'Are you PVTG', 'PVTG Caste/Tribe',
               'Marital Status', 'dependents', 'state', 'district', 'taluka', 'village', 'city', 'add_1', 'add_2',
               'pincode', 'SSC Passing Year', 'SSC Percentage', 'SSC School Name', 'SSC Stream', 'SSC Attempts',
                'SSC Total', 'HSC Passing Year', 'HSC Percentage', 'HSC School Name', 'HSC Stream', 'HSC Attempts', 'HSC Total',
                'Graduation Passing Year', 'Graduation Percentage', 'Graduation School Name', 'Graduation Stream', 'Graduation Attempts',
                'Graduation Total', 'PhD Passing Year', 'PhD Percentage', 'PhD School Name', 'PG Stream', 'PG Attempts', 'PG Total',
                'Have you Qualified', 'Name of College', 'Other College Name', 'Name of Guide', 'Topic of PhD',
                'Concerned University', 'Department Name', 'Faculty', 'PhD Registration Date', 'PhD Registration Year',
                'PhD Registration Age', 'Family Annual Income', 'Income Certificate Number', 'Issuing Authority',
                'Income Issuing District', 'Income Issuing Taluka', 'Domicile', 'Domicile Certificate', 'Domicile Number',
                'Validity Certificate', 'Validity Cert Number', 'Validity Issuing District', 'Validity Issuing Taluka',
                'Validity Issuing Authority', 'Caste Certificate', 'Caste Certf Number', 'Issuing District', 'Caste Issuing Authority',
                'Salaried', 'Disability', 'Type of Disability', 'Father Name', 'Mother Name', 'Work in Government', 'Government Department',
                'Government Position', 'Bank Name', 'Account Number', 'IFSC Code', 'Account Holder Name', 'MICR' ])

    # Add data to the worksheet
    for row in data:
        ws.append(row)
    # Save the workbook in memory as bytes
    data = BytesIO()
    wb.save(data)
    data.seek(0)
    # Create a response object and attach the workbook as a file
    response = make_response(data.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=Level 3 Pending Students.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    return response


@app.route('/export_rejected_students_level3', methods=['GET', 'POST'])
def export_rejected_students_level3():
    if not session.get('logged_in'):
        # Redirect to the admin login page if the user is not logged in
        return redirect(url_for('admin_login'))
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor()
    cursor.execute("  SELECT applicant_id, adhaar_number, first_name, last_name, middle_name, mobile_number, email, "
                   "date_of_birth, gender, age, caste, your_caste, subcaste, pvtg, pvtg_caste, marital_status," 
                    "dependents, state, district, taluka, village, city, add_1, add_2, pincode," 
                    "ssc_passing_year, ssc_percentage, ssc_school_name, ssc_stream, ssc_attempts, ssc_total," 
                    "hsc_passing_year, hsc_percentage, hsc_school_name, hsc_stream, hsc_attempts, hsc_total," 
                    "graduation_passing_year, graduation_percentage, graduation_school_name, grad_stream," 
                    "grad_attempts, grad_total, phd_passing_year, phd_percentage, phd_school_name, pg_stream, "
                    "pg_attempts, pg_total, have_you_qualified, name_of_college, other_college_name, "
                    "name_of_guide, topic_of_phd, concerned_university, department_name, faculty, "
                    "phd_registration_date, phd_registration_year, phd_registration_age, family_annual_income," 
                    "income_certificate_number, issuing_authority, income_issuing_district, income_issuing_taluka," 
                    "domicile, domicile_certificate, domicile_number, validity_certificate, validity_cert_number," 
                    "validity_issuing_district, validity_issuing_taluka, validity_issuing_authority, caste_certf, "
                    "caste_certf_number, issuing_district, caste_issuing_authority, salaried, disability, "
                    "type_of_disability, father_name, mother_name, work_in_government, gov_department," 
                    "gov_position, bank_name, account_number, ifsc_code, account_holder_name, micr FROM application_page"
                   " WHERE phd_registration_year>='2023' and final_approval='rejected' and scrutiny_status='accepted' ")
    data = cursor.fetchall()
    # Create a workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Add headers to the worksheet
    # ws.append(['applicant_id','email','first_name','last_name','application_date'])

    ws.append(['Applicant Id', 'Adhaar Card Number', 'First Name', 'Middle Name', 'Last Name', 'Mobile Number', 'Email',
               'Date Of Birth', 'Gender', 'Age', 'Caste/Tribe', 'Your Caste', 'Sub Caste', 'Are you PVTG', 'PVTG Caste/Tribe',
               'Marital Status', 'dependents', 'state', 'district', 'taluka', 'village', 'city', 'add_1', 'add_2',
               'pincode', 'SSC Passing Year', 'SSC Percentage', 'SSC School Name', 'SSC Stream', 'SSC Attempts',
                'SSC Total', 'HSC Passing Year', 'HSC Percentage', 'HSC School Name', 'HSC Stream', 'HSC Attempts', 'HSC Total',
                'Graduation Passing Year', 'Graduation Percentage', 'Graduation School Name', 'Graduation Stream', 'Graduation Attempts',
                'Graduation Total', 'PhD Passing Year', 'PhD Percentage', 'PhD School Name', 'PG Stream', 'PG Attempts', 'PG Total',
                'Have you Qualified', 'Name of College', 'Other College Name', 'Name of Guide', 'Topic of PhD',
                'Concerned University', 'Department Name', 'Faculty', 'PhD Registration Date', 'PhD Registration Year',
                'PhD Registration Age', 'Family Annual Income', 'Income Certificate Number', 'Issuing Authority',
                'Income Issuing District', 'Income Issuing Taluka', 'Domicile', 'Domicile Certificate', 'Domicile Number',
                'Validity Certificate', 'Validity Cert Number', 'Validity Issuing District', 'Validity Issuing Taluka',
                'Validity Issuing Authority', 'Caste Certificate', 'Caste Certf Number', 'Issuing District', 'Caste Issuing Authority',
                'Salaried', 'Disability', 'Type of Disability', 'Father Name', 'Mother Name', 'Work in Government', 'Government Department',
                'Government Position', 'Bank Name', 'Account Number', 'IFSC Code', 'Account Holder Name', 'MICR' ])

    # Add data to the worksheet
    for row in data:
        ws.append(row)
    # Save the workbook in memory as bytes
    data = BytesIO()
    wb.save(data)
    data.seek(0)
    # Create a response object and attach the workbook as a file
    response = make_response(data.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=Level 3 Rejected Students.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    return response


@app.route('/export_pvtg_students_level3', methods=['GET', 'POST'])
def export_pvtg_students_level3():
    if not session.get('logged_in'):
        # Redirect to the admin login page if the user is not logged in
        return redirect(url_for('admin_login'))
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor()
    cursor.execute("  SELECT applicant_id, adhaar_number, first_name, last_name, middle_name, mobile_number, email, "
                   "date_of_birth, gender, age, caste, your_caste, subcaste, pvtg, pvtg_caste, marital_status," 
                    "dependents, state, district, taluka, village, city, add_1, add_2, pincode," 
                    "ssc_passing_year, ssc_percentage, ssc_school_name, ssc_stream, ssc_attempts, ssc_total," 
                    "hsc_passing_year, hsc_percentage, hsc_school_name, hsc_stream, hsc_attempts, hsc_total," 
                    "graduation_passing_year, graduation_percentage, graduation_school_name, grad_stream," 
                    "grad_attempts, grad_total, phd_passing_year, phd_percentage, phd_school_name, pg_stream, "
                    "pg_attempts, pg_total, have_you_qualified, name_of_college, other_college_name, "
                    "name_of_guide, topic_of_phd, concerned_university, department_name, faculty, "
                    "phd_registration_date, phd_registration_year, phd_registration_age, family_annual_income," 
                    "income_certificate_number, issuing_authority, income_issuing_district, income_issuing_taluka," 
                    "domicile, domicile_certificate, domicile_number, validity_certificate, validity_cert_number," 
                    "validity_issuing_district, validity_issuing_taluka, validity_issuing_authority, caste_certf, "
                    "caste_certf_number, issuing_district, caste_issuing_authority, salaried, disability, "
                    "type_of_disability, father_name, mother_name, work_in_government, gov_department," 
                    "gov_position, bank_name, account_number, ifsc_code, account_holder_name, micr FROM application_page"
                   " WHERE phd_registration_year>='2023' and your_caste IN ('katkari', 'kolam', 'madia') ")
    data = cursor.fetchall()
    # Create a workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Add headers to the worksheet
    # ws.append(['applicant_id','email','first_name','last_name','application_date'])

    ws.append(['Applicant Id', 'Adhaar Card Number', 'First Name', 'Middle Name', 'Last Name', 'Mobile Number', 'Email',
               'Date Of Birth', 'Gender', 'Age', 'Caste/Tribe', 'Your Caste', 'Sub Caste', 'Are you PVTG', 'PVTG Caste/Tribe',
               'Marital Status', 'dependents', 'state', 'district', 'taluka', 'village', 'city', 'add_1', 'add_2',
               'pincode', 'SSC Passing Year', 'SSC Percentage', 'SSC School Name', 'SSC Stream', 'SSC Attempts',
                'SSC Total', 'HSC Passing Year', 'HSC Percentage', 'HSC School Name', 'HSC Stream', 'HSC Attempts', 'HSC Total',
                'Graduation Passing Year', 'Graduation Percentage', 'Graduation School Name', 'Graduation Stream', 'Graduation Attempts',
                'Graduation Total', 'PhD Passing Year', 'PhD Percentage', 'PhD School Name', 'PG Stream', 'PG Attempts', 'PG Total',
                'Have you Qualified', 'Name of College', 'Other College Name', 'Name of Guide', 'Topic of PhD',
                'Concerned University', 'Department Name', 'Faculty', 'PhD Registration Date', 'PhD Registration Year',
                'PhD Registration Age', 'Family Annual Income', 'Income Certificate Number', 'Issuing Authority',
                'Income Issuing District', 'Income Issuing Taluka', 'Domicile', 'Domicile Certificate', 'Domicile Number',
                'Validity Certificate', 'Validity Cert Number', 'Validity Issuing District', 'Validity Issuing Taluka',
                'Validity Issuing Authority', 'Caste Certificate', 'Caste Certf Number', 'Issuing District', 'Caste Issuing Authority',
                'Salaried', 'Disability', 'Type of Disability', 'Father Name', 'Mother Name', 'Work in Government', 'Government Department',
                'Government Position', 'Bank Name', 'Account Number', 'IFSC Code', 'Account Holder Name', 'MICR' ])

    # Add data to the worksheet
    for row in data:
        ws.append(row)
    # Save the workbook in memory as bytes
    data = BytesIO()
    wb.save(data)
    data.seek(0)
    # Create a response object and attach the workbook as a file
    response = make_response(data.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=Level 3 PVTG Students.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    return response


@app.route('/export_disabled_students_level3', methods=['GET', 'POST'])
def export_disabled_students_level3():
    if not session.get('logged_in'):
        # Redirect to the admin login page if the user is not logged in
        return redirect(url_for('admin_login'))
    cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                  host=host,
                                  database='ICSApplication')
    cursor = cnx.cursor()
    cursor.execute("  SELECT applicant_id, adhaar_number, first_name, last_name, middle_name, mobile_number, email, "
                   "date_of_birth, gender, age, caste, your_caste, subcaste, pvtg, pvtg_caste, marital_status," 
                    "dependents, state, district, taluka, village, city, add_1, add_2, pincode," 
                    "ssc_passing_year, ssc_percentage, ssc_school_name, ssc_stream, ssc_attempts, ssc_total," 
                    "hsc_passing_year, hsc_percentage, hsc_school_name, hsc_stream, hsc_attempts, hsc_total," 
                    "graduation_passing_year, graduation_percentage, graduation_school_name, grad_stream," 
                    "grad_attempts, grad_total, phd_passing_year, phd_percentage, phd_school_name, pg_stream, "
                    "pg_attempts, pg_total, have_you_qualified, name_of_college, other_college_name, "
                    "name_of_guide, topic_of_phd, concerned_university, department_name, faculty, "
                    "phd_registration_date, phd_registration_year, phd_registration_age, family_annual_income," 
                    "income_certificate_number, issuing_authority, income_issuing_district, income_issuing_taluka," 
                    "domicile, domicile_certificate, domicile_number, validity_certificate, validity_cert_number," 
                    "validity_issuing_district, validity_issuing_taluka, validity_issuing_authority, caste_certf, "
                    "caste_certf_number, issuing_district, caste_issuing_authority, salaried, disability, "
                    "type_of_disability, father_name, mother_name, work_in_government, gov_department," 
                    "gov_position, bank_name, account_number, ifsc_code, account_holder_name, micr FROM application_page"
                   " WHERE phd_registration_year>='2023' and disability='Yes' ")
    data = cursor.fetchall()
    # Create a workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Add headers to the worksheet
    # ws.append(['applicant_id','email','first_name','last_name','application_date'])

    ws.append(['Applicant Id', 'Adhaar Card Number', 'First Name', 'Middle Name', 'Last Name', 'Mobile Number', 'Email',
               'Date Of Birth', 'Gender', 'Age', 'Caste/Tribe', 'Your Caste', 'Sub Caste', 'Are you PVTG', 'PVTG Caste/Tribe',
               'Marital Status', 'dependents', 'state', 'district', 'taluka', 'village', 'city', 'add_1', 'add_2',
               'pincode', 'SSC Passing Year', 'SSC Percentage', 'SSC School Name', 'SSC Stream', 'SSC Attempts',
                'SSC Total', 'HSC Passing Year', 'HSC Percentage', 'HSC School Name', 'HSC Stream', 'HSC Attempts', 'HSC Total',
                'Graduation Passing Year', 'Graduation Percentage', 'Graduation School Name', 'Graduation Stream', 'Graduation Attempts',
                'Graduation Total', 'PhD Passing Year', 'PhD Percentage', 'PhD School Name', 'PG Stream', 'PG Attempts', 'PG Total',
                'Have you Qualified', 'Name of College', 'Other College Name', 'Name of Guide', 'Topic of PhD',
                'Concerned University', 'Department Name', 'Faculty', 'PhD Registration Date', 'PhD Registration Year',
                'PhD Registration Age', 'Family Annual Income', 'Income Certificate Number', 'Issuing Authority',
                'Income Issuing District', 'Income Issuing Taluka', 'Domicile', 'Domicile Certificate', 'Domicile Number',
                'Validity Certificate', 'Validity Cert Number', 'Validity Issuing District', 'Validity Issuing Taluka',
                'Validity Issuing Authority', 'Caste Certificate', 'Caste Certf Number', 'Issuing District', 'Caste Issuing Authority',
                'Salaried', 'Disability', 'Type of Disability', 'Father Name', 'Mother Name', 'Work in Government', 'Government Department',
                'Government Position', 'Bank Name', 'Account Number', 'IFSC Code', 'Account Holder Name', 'MICR' ])

    # Add data to the worksheet
    for row in data:
        ws.append(row)
    # Save the workbook in memory as bytes
    data = BytesIO()
    wb.save(data)
    data.seek(0)
    # Create a response object and attach the workbook as a file
    response = make_response(data.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=Level 3 Disabled Students.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    return response




if __name__ == '__main__':
    app.run(debug = True)


# 27th March 2024 (10:04)