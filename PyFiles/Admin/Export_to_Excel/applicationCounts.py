from io import BytesIO
import mysql.connector
from openpyxl.workbook import Workbook
from classes.connection import HostConfig, ConfigPaths
import os
from flask import Blueprint, render_template, session, request, redirect, url_for, flash, make_response


app_counts_blueprint = Blueprint('app_counts', __name__)


def app_counts_auth(app):
    # ------ HOST Configs are in classes/connection.py
    host = HostConfig.host
    app_paths = ConfigPaths.paths.get(host)
    if app_paths:
        for key, value in app_paths.items():
            app.config[key] = value

    # @app_counts_blueprint.route('/total_application_report_22', methods=['GET', 'POST'])
    # def total_application_report_22():
    #     cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
    #                                   host=host,
    #                                   database='ICSApplication')
    #     cursor = cnx.cursor(dictionary=True)
    #     cursor.execute("SELECT * FROM application_page WHERE phd_registration_year='2022' ")
    #     result = cursor.fetchall()
    #     cnx.commit()
    #     cursor.close()
    #     cnx.close()
    #     # print(result)
    #     return render_template('Admin/Export_to_Excel/total_application_report.html', result=result)

    @app_counts_blueprint.route('/total_application_report', methods=['GET', 'POST'])
    def total_application_report():
        cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                      host=host,
                                      database='ICSApplication')
        cursor = cnx.cursor(dictionary=True)
        cursor.execute(" SELECT * FROM application_page where phd_registration_year = 2023 ")
        result = cursor.fetchall()
        cnx.commit()
        cursor.close()
        cnx.close()
        # print(result)
        return render_template('Admin/Export_to_Excel/total_application_report.html', result=result)

    @app_counts_blueprint.route('/export_total_applications_23', methods=['GET', 'POST'])
    def export_total_applications_23():
        cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                      host=host,
                                      database='ICSApplication')
        cursor = cnx.cursor()
        cursor.execute(
            "  SELECT applicant_id, adhaar_number, first_name, last_name, middle_name, mobile_number, email, "
            "date_of_birth, gender, age, caste, your_caste, subcaste, pvtg, pvtg_caste, marital_status,"
            "dependents, state, district, taluka, village, city, add_1, add_2, pincode,"
            "ssc_passing_year, ssc_percentage, ssc_school_name, ssc_stream, ssc_attempts, ssc_total,"
            "hsc_passing_year, hsc_percentage, hsc_school_name, hsc_stream, hsc_attempts, hsc_total,"
            "graduation_passing_year, graduation_percentage, graduation_school_name, grad_stream,"
            "grad_attempts, grad_total, phd_passing_year, phd_percentage, phd_school_name, pg_stream, "
            "pg_attempts, pg_total, have_you_qualified, name_of_college, other_college_name, "
            "name_of_guide, topic_of_phd, concerned_university, department_name, faculty, "
            "phd_registration_date, phd_registration_year, phd_registration_age, family_annual_income,"
            "income_certificate_number, issuing_authority, income_issuing_district, income_issuing_taluka,"
            "domicile, domicile_certificate, domicile_number, validity_certificate, validity_cert_number,"
            "validity_issuing_district, validity_issuing_taluka, validity_issuing_authority, caste_certf, "
            "caste_certf_number, issuing_district, caste_issuing_authority, salaried, disability, "
            "type_of_disability, father_name, mother_name, work_in_government, gov_department,"
            "gov_position, bank_name, account_number, ifsc_code, account_holder_name, micr FROM application_page"
            " where phd_registration_year = '2023' ")

        data = cursor.fetchall()
        # Create a workbook and add a worksheet
        wb = Workbook()
        ws = wb.active

        # Add headers to the worksheet
        # ws.append(['applicant_id','email','first_name','last_name','application_date'])

        ws.append(
            ['Applicant Id', 'Adhaar Card Number', 'First Name', 'Middle Name', 'Last Name', 'Mobile Number', 'Email',
             'Date Of Birth', 'Gender', 'Age', 'Caste/Tribe', 'Sub Caste', 'Are you PVTG', 'PVTG Caste/Tribe',
             'Marital Status', 'dependents', 'state', 'district', 'taluka', 'village', 'city', 'add_1', 'add_2',
             'pincode', 'SSC Passing Year', 'SSC Percentage', 'SSC School Name', 'SSC Stream', 'SSC Attempts',
             'SSC Total', 'HSC Passing Year', 'HSC Percentage', 'HSC School Name', 'HSC Stream', 'HSC Attempts',
             'HSC Total',
             'Graduation Passing Year', 'Graduation Percentage', 'Graduation School Name', 'Graduation Stream',
             'Graduation Attempts',
             'Graduation Total', 'PhD Passing Year', 'PhD Percentage', 'PhD School Name', 'PG Stream', 'PG Attempts',
             'PG Total',
             'Have you Qualified', 'Name of College', 'Other College Name', 'Name of Guide', 'Topic of PhD',
             'Concerned University', 'Department Name', 'Faculty', 'PhD Registration Date', 'PhD Registration Year',
             'PhD Registration Age', 'Family Annual Income', 'Income Certificate Number', 'Issuing Authority',
             'Income Issuing District', 'Income Issuing Taluka', 'Domicile', 'Domicile Certificate', 'Domicile Number',
             'Validity Certificate', 'Validity Cert Number', 'Validity Issuing District', 'Validity Issuing Taluka',
             'Validity Issuing Authority', 'Caste Certificate', 'Caste Certf Number', 'Issuing District',
             'Caste Issuing Authority',
             'Salaried', 'Disability', 'Type of Disability', 'Father Name', 'Mother Name', 'Work in Government',
             'Government Department',
             'Government Position', 'Bank Name', 'Account Number', 'IFSC Code', 'Account Holder Name', 'MICR'])

        # Add data to the worksheet
        for row in data:
            ws.append(row)

        # Save the workbook in memory as bytes
        data = BytesIO()
        wb.save(data)
        data.seek(0)

        # Create a response object and attach the workbook as a file
        response = make_response(data.getvalue())
        response.headers['Content-Disposition'] = 'attachment; filename=Total_Applications_2023.xlsx'
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

        return response

    @app_counts_blueprint.route('/completed_form', methods=['GET', 'POST'])
    def completed_form():
        cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                      host=host,
                                      database='ICSApplication')
        cursor = cnx.cursor(dictionary=True)
        cursor.execute(" SELECT * FROM application_page where phd_registration_year = 2023 and form_filled='1' ")
        result = cursor.fetchall()
        cnx.commit()
        cursor.close()
        cnx.close()
        # print(result)
        return render_template('Admin/Export_to_Excel/completed_form.html', result=result)

    @app_counts_blueprint.route('/completed_form_export', methods=['GET', 'POST'])
    def completed_form_export():
        cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                      host=host,
                                      database='ICSApplication')
        cursor = cnx.cursor()
        cursor.execute(
            "  SELECT applicant_id, adhaar_number, first_name, last_name, middle_name, mobile_number, email, "
            "date_of_birth, gender, age, caste, your_caste, subcaste, pvtg, pvtg_caste, marital_status,"
            "dependents, state, district, taluka, village, city, add_1, add_2, pincode,"
            "ssc_passing_year, ssc_percentage, ssc_school_name, ssc_stream, ssc_attempts, ssc_total,"
            "hsc_passing_year, hsc_percentage, hsc_school_name, hsc_stream, hsc_attempts, hsc_total,"
            "graduation_passing_year, graduation_percentage, graduation_school_name, grad_stream,"
            "grad_attempts, grad_total, phd_passing_year, phd_percentage, phd_school_name, pg_stream, "
            "pg_attempts, pg_total, have_you_qualified, name_of_college, other_college_name, "
            "name_of_guide, topic_of_phd, concerned_university, department_name, faculty, "
            "phd_registration_date, phd_registration_year, phd_registration_age, family_annual_income,"
            "income_certificate_number, issuing_authority, income_issuing_district, income_issuing_taluka,"
            "domicile, domicile_certificate, domicile_number, validity_certificate, validity_cert_number,"
            "validity_issuing_district, validity_issuing_taluka, validity_issuing_authority, caste_certf, "
            "caste_certf_number, issuing_district, caste_issuing_authority, salaried, disability, "
            "type_of_disability, father_name, mother_name, work_in_government, gov_department,"
            "gov_position, bank_name, account_number, ifsc_code, account_holder_name, micr FROM application_page"
            " where phd_registration_year = '2023' and form_filled='1' ")

        data = cursor.fetchall()
        # Create a workbook and add a worksheet
        wb = Workbook()
        ws = wb.active

        # Add headers to the worksheet
        # ws.append(['applicant_id','email','first_name','last_name','application_date'])

        ws.append(
            ['Applicant Id', 'Adhaar Card Number', 'First Name', 'Middle Name', 'Last Name', 'Mobile Number', 'Email',
             'Date Of Birth', 'Gender', 'Age', 'Caste/Tribe', 'Your Caste', 'Sub Caste', 'Are you PVTG',
             'PVTG Caste/Tribe',
             'Marital Status', 'dependents', 'state', 'district', 'taluka', 'village', 'city', 'add_1', 'add_2',
             'pincode', 'SSC Passing Year', 'SSC Percentage', 'SSC School Name', 'SSC Stream', 'SSC Attempts',
             'SSC Total', 'HSC Passing Year', 'HSC Percentage', 'HSC School Name', 'HSC Stream', 'HSC Attempts',
             'HSC Total',
             'Graduation Passing Year', 'Graduation Percentage', 'Graduation School Name', 'Graduation Stream',
             'Graduation Attempts',
             'Graduation Total', 'PhD Passing Year', 'PhD Percentage', 'PhD School Name', 'PG Stream', 'PG Attempts',
             'PG Total',
             'Have you Qualified', 'Name of College', 'Other College Name', 'Name of Guide', 'Topic of PhD',
             'Concerned University', 'Department Name', 'Faculty', 'PhD Registration Date', 'PhD Registration Year',
             'PhD Registration Age', 'Family Annual Income', 'Income Certificate Number', 'Issuing Authority',
             'Income Issuing District', 'Income Issuing Taluka', 'Domicile', 'Domicile Certificate', 'Domicile Number',
             'Validity Certificate', 'Validity Cert Number', 'Validity Issuing District', 'Validity Issuing Taluka',
             'Validity Issuing Authority', 'Caste Certificate', 'Caste Certf Number', 'Issuing District',
             'Caste Issuing Authority',
             'Salaried', 'Disability', 'Type of Disability', 'Father Name', 'Mother Name', 'Work in Government',
             'Government Department',
             'Government Position', 'Bank Name', 'Account Number', 'IFSC Code', 'Account Holder Name', 'MICR'])

        # Add data to the worksheet
        for row in data:
            ws.append(row)

        # Save the workbook in memory as bytes
        data = BytesIO()
        wb.save(data)
        data.seek(0)

        # Create a response object and attach the workbook as a file
        response = make_response(data.getvalue())
        response.headers['Content-Disposition'] = 'attachment; filename=Completed Form Fellowship 2023.xlsx'
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

        return response

    @app_counts_blueprint.route('/incompleted_form', methods=['GET', 'POST'])
    def incompleted_form():
        cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                      host=host,
                                      database='ICSApplication')
        cursor = cnx.cursor(dictionary=True)
        cursor.execute(" SELECT * FROM application_page where phd_registration_year = 2023 and form_filled='0' ")
        # cursor.execute(" SELECT * FROM application_page where email = 'tupotbhare@gmail.com' ")
        result = cursor.fetchall()
        cnx.commit()
        cursor.close()
        cnx.close()
        # print(result)
        return render_template('Admin/Export_to_Excel/incompleted_form.html', result=result)

    @app_counts_blueprint.route('/incompleted_form_export', methods=['GET', 'POST'])
    def incompleted_form_export():
        cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                      host=host,
                                      database='ICSApplication')
        cursor = cnx.cursor()
        cursor.execute(
            "  SELECT applicant_id, adhaar_number, first_name, last_name, middle_name, mobile_number, email, "
            "date_of_birth, gender, age, caste, your_caste, subcaste, pvtg, pvtg_caste, marital_status,"
            "dependents, state, district, taluka, village, city, add_1, add_2, pincode,"
            "ssc_passing_year, ssc_percentage, ssc_school_name, ssc_stream, ssc_attempts, ssc_total,"
            "hsc_passing_year, hsc_percentage, hsc_school_name, hsc_stream, hsc_attempts, hsc_total,"
            "graduation_passing_year, graduation_percentage, graduation_school_name, grad_stream,"
            "grad_attempts, grad_total, phd_passing_year, phd_percentage, phd_school_name, pg_stream, "
            "pg_attempts, pg_total, have_you_qualified, name_of_college, other_college_name, "
            "name_of_guide, topic_of_phd, concerned_university, department_name, faculty, "
            "phd_registration_date, phd_registration_year, phd_registration_age, family_annual_income,"
            "income_certificate_number, issuing_authority, income_issuing_district, income_issuing_taluka,"
            "domicile, domicile_certificate, domicile_number, validity_certificate, validity_cert_number,"
            "validity_issuing_district, validity_issuing_taluka, validity_issuing_authority, caste_certf, "
            "caste_certf_number, issuing_district, caste_issuing_authority, salaried, disability, "
            "type_of_disability, father_name, mother_name, work_in_government, gov_department,"
            "gov_position, bank_name, account_number, ifsc_code, account_holder_name, micr FROM application_page"
            " where phd_registration_year = '2023' and form_filled='0' ")

        data = cursor.fetchall()
        # Create a workbook and add a worksheet
        wb = Workbook()
        ws = wb.active

        # Add headers to the worksheet
        # ws.append(['applicant_id','email','first_name','last_name','application_date'])

        ws.append(
            ['Applicant Id', 'Adhaar Card Number', 'First Name', 'Middle Name', 'Last Name', 'Mobile Number', 'Email',
             'Date Of Birth', 'Gender', 'Age', 'Caste/Tribe', 'Sub Caste', 'Are you PVTG', 'PVTG Caste/Tribe',
             'Marital Status', 'dependents', 'state', 'district', 'taluka', 'village', 'city', 'add_1', 'add_2',
             'pincode', 'SSC Passing Year', 'SSC Percentage', 'SSC School Name', 'SSC Stream', 'SSC Attempts',
             'SSC Total', 'HSC Passing Year', 'HSC Percentage', 'HSC School Name', 'HSC Stream', 'HSC Attempts',
             'HSC Total',
             'Graduation Passing Year', 'Graduation Percentage', 'Graduation School Name', 'Graduation Stream',
             'Graduation Attempts',
             'Graduation Total', 'PhD Passing Year', 'PhD Percentage', 'PhD School Name', 'PG Stream', 'PG Attempts',
             'PG Total',
             'Have you Qualified', 'Name of College', 'Other College Name', 'Name of Guide', 'Topic of PhD',
             'Concerned University', 'Department Name', 'Faculty', 'PhD Registration Date', 'PhD Registration Year',
             'PhD Registration Age', 'Family Annual Income', 'Income Certificate Number', 'Issuing Authority',
             'Income Issuing District', 'Income Issuing Taluka', 'Domicile', 'Domicile Certificate', 'Domicile Number',
             'Validity Certificate', 'Validity Cert Number', 'Validity Issuing District', 'Validity Issuing Taluka',
             'Validity Issuing Authority', 'Caste Certificate', 'Caste Certf Number', 'Issuing District',
             'Caste Issuing Authority',
             'Salaried', 'Disability', 'Type of Disability', 'Father Name', 'Mother Name', 'Work in Government',
             'Government Department',
             'Government Position', 'Bank Name', 'Account Number', 'IFSC Code', 'Account Holder Name', 'MICR'])

        # Add data to the worksheet
        for row in data:
            ws.append(row)

        # Save the workbook in memory as bytes
        data = BytesIO()
        wb.save(data)
        data.seek(0)

        # Create a response object and attach the workbook as a file
        response = make_response(data.getvalue())
        response.headers['Content-Disposition'] = 'attachment; filename=Total_Applications_2023.xlsx'
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

        return response

    @app_counts_blueprint.route('/total_accepted_report', methods=['GET', 'POST'])
    def total_accepted_report():
        cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                      host=host,
                                      database='ICSApplication')
        cursor = cnx.cursor(dictionary=True)
        cursor.execute(
            " SELECT * FROM application_page WHERE final_approval='accepted' and phd_registration_year='2023' ")
        result = cursor.fetchall()
        cnx.commit()
        cursor.close()
        cnx.close()
        # print(result)
        return render_template('Admin/Export_to_Excel/total_accepted_report.html', result=result)

    @app_counts_blueprint.route('/total_rejected_report', methods=['GET', 'POST'])
    def total_rejected_report():
        cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                      host=host,
                                      database='ICSApplication')
        cursor = cnx.cursor(dictionary=True)
        cursor.execute(
            "SELECT * FROM application_page WHERE final_approval='rejected' and phd_registration_year='2023'")
        result = cursor.fetchall()
        cnx.commit()
        cursor.close()
        cnx.close()
        # print(result)
        return render_template('total_rejected_report.html', result=result)
