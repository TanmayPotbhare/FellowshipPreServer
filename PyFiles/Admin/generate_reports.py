import mysql.connector
from classes.connection import HostConfig, ConfigPaths
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from io import BytesIO
from datetime import datetime, timedelta, date
import os
from flask import Blueprint, render_template, session, request, redirect, url_for, flash, make_response
from authentication.middleware import auth

generate_reports_blueprint = Blueprint('generate_reports', __name__)


def generate_reports_auth(app):
    # ------ HOST Configs are in classes/connection.py
    host = HostConfig.host
    app_paths = ConfigPaths.paths.get(host)
    if app_paths:
        for key, value in app_paths.items():
            app.config[key] = value

    @generate_reports_blueprint.route('/generate_reports_admin', methods=['GET', 'POST'])
    def generate_reports_admin():
        if request.method == 'POST':
            start_date = request.form.get('start_date')
            end_date = request.form.get('end_date')
            year = request.form.get('year')
            filter_option = request.form.get('filter')
            stream = request.form.get('stream')
            query = build_query(start_date, end_date, year, filter_option, stream)
            data = fetch_data(query)

            if 'fetch_result' in request.form:
                # Code to fetch and display results
                return render_template('Admin/generate_reports_admin.html', data=data)

            elif 'export_to_excel' in request.form:
                # Code to export data to Excel
                wb = Workbook()
                ws = wb.active

                headers = ['Applicant Id', 'First Name', 'Middle Name', 'Last Name', 'Mobile Number', 'Email',
                           'Date Of Birth', 'Gender', 'Age', 'Caste', 'Your Caste', 'Marital Status', 'Dependents',
                           'Add 1',
                           'Add 2',
                           'Pincode', 'Village', 'Taluka', 'District', 'State', 'Phd Registration Date',
                           'Concerned University', 'Topic Of Phd',
                           'Name Of Guide', 'Name Of College', 'Stream', 'Board University', 'Admission Year',
                           'Passing Year',
                           'Percentage', 'Family Annual Income', 'Income Certificate Number', 'Issuing Authority',
                           'Domicile',
                           'Domicile Certificate', 'Domicile Number', 'Caste Certf', 'Issuing District',
                           'Caste Issuing Authority',
                           'Salaried', 'Disability', 'Father Name', 'Mother Name', 'Work In Government', 'Bank Name',
                           'Account Number',
                           'Ifsc Code', 'Account Holder Name']

                ws.append(headers)

                description = "List of Students"
                ws.insert_rows(1)
                header_row = ws[1]
                header_row[0].value = description
                for cell in header_row:
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center')
                    cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

                for row in data:
                    ws.append([row.get(header.lower().replace(' ', '_'), '') for header in headers])

                output = BytesIO()
                wb.save(output)
                output.seek(0)

                filename = f"{filter_option} {year} Students.xlsx" if year else f"{filter_option} Students.xlsx"

                response = make_response(output.getvalue())
                response.headers['Content-Disposition'] = f'attachment; filename={filename}'
                response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

                return response

        return render_template('Admin/generate_reports_admin.html')

    def build_query(start_date, end_date, year, filter_option, stream):
        query = "SELECT * FROM application_page"
        conditions = []

        if start_date and end_date:
            conditions.append(f"phd_registration_date BETWEEN '{start_date}' AND '{end_date}'")
        if year:
            conditions.append(f"phd_registration_year = '{year}'")
        if filter_option:
            if filter_option == 'total':
                # No specific filter for total
                pass
            elif filter_option == 'pending':
                conditions.append("status = 'pending'")
            elif filter_option == 'accepted':
                conditions.append("status = 'accepted'")
            elif filter_option == 'rejected':
                conditions.append("status = 'rejected'")
            elif filter_option == 'complete':
                conditions.append("form_filled = '1'")
            elif filter_option == 'incomplete':
                conditions.append("form_filled = '0'")
            elif filter_option == 'male':
                conditions.append("gender = 'male'")
            elif filter_option == 'female':
                conditions.append("gender = 'female'")
            elif filter_option == 'transgender':
                conditions.append("gender = 'transgender'")
            elif filter_option == 'Katari':
                conditions.append("your_caste = 'Katari'")
            elif filter_option == 'Kolam':
                conditions.append("your_caste = 'Kolam'")
            elif filter_option == 'Madia':
                conditions.append("your_caste = 'Madia'")
            elif filter_option == 'disabled':
                conditions.append("disability = 'Yes'")
            elif filter_option == 'notdisabled':
                conditions.append("disability = 'No'")
        if stream:
            if stream == 'science':
                conditions.append("faculty = 'science'")
            if stream == 'arts':
                conditions.append("faculty = 'arts'")
            if stream == 'commerce':
                conditions.append("faculty = 'commerce'")
            if stream == 'other':
                conditions.append("faculty = 'other'")

        if conditions:
            query += " WHERE " + " AND ".join(conditions)

        print('query', query)
        return query

    def fetch_data(query):
        cnx = mysql.connector.connect(user='root', password='A9CALcsd7lc%7ac',
                                      host=host,
                                      database='ICSApplication')
        cursor = cnx.cursor(dictionary=True)
        cursor.execute(query)
        results = cursor.fetchall()
        cursor.close()
        cnx.close()
        for result in results:
            for key, value in result.items():
                if isinstance(value, (date, datetime)):
                    result[key] = value.isoformat()
        return results